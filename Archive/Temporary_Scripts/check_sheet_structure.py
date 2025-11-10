"""Check the structure of property sheets in master file"""

import openpyxl
from pathlib import Path

file_path = Path('Portfolio_Reports/MASTER_Portfolio_Complete_Data.xlsx')
wb = openpyxl.load_workbook(file_path)

# Check McCord Park FL sheet
ws = wb['McCord Park FL']

print('McCord Park FL Sheet Structure:')
print('=' * 80)
print()

# Show first 15 rows
for row in range(1, 16):
    row_data = []
    for col in range(1, 10):
        val = ws.cell(row, col).value
        if val:
            row_data.append(f'Col{col}: {str(val)[:40]}')
    if row_data:
        print(f'Row {row}: {" | ".join(row_data)}')
    else:
        print(f'Row {row}: (empty)')

print()
print('=' * 80)
print()

# Check Orion McKinney
ws2 = wb['Orion McKinney']
print('Orion McKinney Sheet Structure:')
print('=' * 80)
print()

for row in range(1, 16):
    row_data = []
    for col in range(1, 10):
        val = ws2.cell(row, col).value
        if val:
            row_data.append(f'Col{col}: {str(val)[:40]}')
    if row_data:
        print(f'Row {row}: {" | ".join(row_data)}')
    else:
        print(f'Row {row}: (empty)')

wb.close()

