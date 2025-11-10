"""
Export summary of master file contents
"""

import pandas as pd
import openpyxl

master_file = 'Portfolio_Reports/MASTER_Portfolio_Complete_Data.xlsx'

print('=' * 80)
print('MASTER FILE COMPLETE ANALYSIS')
print('=' * 80)
print()

# Load workbook
wb = openpyxl.load_workbook(master_file)

print(f'Total Sheets: {len(wb.sheetnames)}')
print()

print('ALL SHEETS:')
for i, sheet_name in enumerate(wb.sheetnames, 1):
    ws = wb[sheet_name]
    print(f'{i}. {sheet_name} ({ws.max_row} rows × {ws.max_column} cols)')

print()
print('=' * 80)
print('PROPERTY OVERVIEW TAB CHECK')
print('=' * 80)
print()

if 'Property Overview' in wb.sheetnames:
    df = pd.read_excel(master_file, sheet_name='Property Overview')
    
    print(f'Rows: {len(df)}')
    print(f'Columns: {len(df.columns)}')
    print()
    
    print('Columns:')
    for col in df.columns:
        print(f'  - {col}')
    print()
    
    print('First 5 rows:')
    print(df.head().to_string())
    print()
else:
    print('⚠️ Property Overview sheet not found!')
    print()

wb.close()

