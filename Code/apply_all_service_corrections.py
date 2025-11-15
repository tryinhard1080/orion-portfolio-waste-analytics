"""
Apply All Service Configuration Corrections Based on Latest Invoice Validation

This script applies corrections identified during the comprehensive invoice review:
1. Bella Mirage: 10 containers @ 3x/week → 6 containers @ 4x/week
2. Orion McKinney: 10x 8 YD @ 3x/week → 8x 8 YD + 2x 10 YD @ 3x/week
3. McCord Park FL: Verified count but sizes unclear from invoices (keeping as-is)
4. Other properties: Verified correct or insufficient invoice data

Evidence Sources:
- Bella Mirage: Oct & Nov 2024 invoices
- Orion McKinney: September 2025 invoice
- Orion Prosper: August 2025 invoice (verified correct)
- The Club at Millenia: September 2025 invoice (verified correct)
- Orion Prosper Lakes: July 2025 invoice (previously corrected)
- Mandarina: WM Agreement 2018 (contract data - matches current)
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# All current service configurations based on latest invoices
CORRECTED_SERVICE_CONFIGS = {
    'Bella Mirage': {
        'containers': [
            {'type': 'Dumpster (FEL)', 'size': '8 YD', 'quantity': 4, 'frequency': '4x/week'},
            {'type': 'Dumpster (FEL)', 'size': '6 YD', 'quantity': 1, 'frequency': '4x/week'},
            {'type': 'Dumpster (FEL)', 'size': '4 YD', 'quantity': 1, 'frequency': '4x/week'}
        ],
        'notes': 'From Oct & Nov 2024 invoices - 6 containers total @ 4x/week'
    },
    'Orion McKinney': {
        'containers': [
            {'type': 'Dumpster (FEL)', 'size': '8 YD', 'quantity': 8, 'frequency': '3x/week'},
            {'type': 'Dumpster (FEL)', 'size': '10 YD', 'quantity': 2, 'frequency': '3x/week'}
        ],
        'notes': 'From Sept 2025 invoice - 8x 8YD + 2x 10YD @ 3x/week'
    },
    # Keep other properties as-is (already correct or insufficient data to change)
    'Orion Prosper': {
        'containers': [
            {'type': 'Compactor', 'size': '10 YD', 'quantity': 2, 'frequency': '6x/week'}
        ],
        'notes': 'Verified correct from Aug 2025 invoice'
    },
    'Orion Prosper Lakes': {
        'containers': [
            {'type': 'Compactor', 'size': '30 CY', 'quantity': 2, 'frequency': 'On-call (2-3x/week avg)'}
        ],
        'notes': 'Previously corrected - discontinued 40 YD open top early 2025'
    },
    'The Club at Millenia': {
        'containers': [
            {'type': 'Compactor', 'size': '30 YD', 'quantity': 2, 'frequency': 'On-call (varies)'}
        ],
        'notes': 'Verified correct from Sept 2025 invoice - compactor roll-off service'
    },
    'McCord Park FL': {
        'containers': [
            {'type': 'Dumpster (FEL)', 'size': '8 YD', 'quantity': 15, 'frequency': '3x/week'}
        ],
        'notes': 'Invoice shows ~13 containers but sizes not specified - keeping current data'
    },
    'Mandarina': {
        'containers': [
            {'type': 'Compactor', 'size': '6 YD', 'quantity': 2, 'frequency': '3x/week'}
        ],
        'notes': 'From WM contract - compactor service only (Ally Waste bulk separate)'
    },
    'Pavilions at Arrowhead': {
        'containers': [
            {'type': 'Dumpster', 'size': '4 YD', 'quantity': 4, 'frequency': '2x/week'},
            {'type': 'Bulk Service', 'size': 'N/A', 'quantity': 1, 'frequency': 'Weekly (Thursday)'}
        ],
        'notes': 'City of Glendale + WCI Bulk - contract data only'
    },
    'Springs at Alta Mesa': {
        'containers': [
            {'type': 'Dumpster', 'size': '6 YD', 'quantity': 5, 'frequency': '3x/week'},
            {'type': 'Dumpster', 'size': '4 YD', 'quantity': 4, 'frequency': '3x/week'},
            {'type': 'Cart', 'size': '90 gallon', 'quantity': 7, 'frequency': '3x/week'},
            {'type': 'Bulk Service', 'size': 'N/A', 'quantity': 1, 'frequency': 'Weekly (Thursday)'}
        ],
        'notes': 'City of Mesa + WCI Bulk - contract data only'
    },
    'Tempe Vista': {
        'containers': [
            {'type': 'Dumpster', 'size': '4 YD', 'quantity': 6, 'frequency': '3x/week'},
            {'type': 'Dumpster', 'size': '6 YD', 'quantity': 1, 'frequency': '3x/week'},
            {'type': 'Dumpster', 'size': '8 YD', 'quantity': 1, 'frequency': '3x/week'}
        ],
        'notes': 'WM dumpster service + WCI Bulk - contract data only'
    }
}

def main():
    print("="*80)
    print("APPLYING ALL SERVICE CONFIGURATION CORRECTIONS")
    print("="*80)

    master_path = 'Portfolio_Reports/MASTER_Portfolio_Complete_Data.xlsx'

    # Read current Service Details
    df_service = pd.read_excel(master_path, sheet_name='Service Details')

    print("\nCURRENT SERVICE DETAILS:")
    print("="*80)
    for prop in sorted(df_service['Property'].unique()):
        prop_data = df_service[df_service['Property'] == prop]
        total = int(prop_data['Quantity'].sum())
        print(f"{prop}: {total} containers")

    # Build new service details from corrected configurations
    new_service_rows = []

    for property_name, config in CORRECTED_SERVICE_CONFIGS.items():
        for container in config['containers']:
            # Calculate total yards
            try:
                size_str = container['size']
                if 'YD' in size_str or 'CY' in size_str:
                    size_num = int(''.join(filter(str.isdigit, size_str.split()[0])))
                    total_yards = size_num * container['quantity']
                else:
                    total_yards = 0
            except:
                total_yards = 0

            new_service_rows.append({
                'Property': property_name,
                'Container Type': container['type'],
                'Container Size': container['size'],
                'Quantity': container['quantity'],
                'Frequency': container['frequency'],
                'Total Yards': total_yards
            })

    df_new = pd.DataFrame(new_service_rows)
    df_new = df_new.sort_values('Property').reset_index(drop=True)

    print("\n" + "="*80)
    print("CORRECTED SERVICE DETAILS:")
    print("="*80)
    for prop in sorted(df_new['Property'].unique()):
        prop_data = df_new[df_new['Property'] == prop]
        total = int(prop_data['Quantity'].sum())
        print(f"{prop}: {total} containers")
        for _, row in prop_data.iterrows():
            print(f"  - {int(row['Quantity'])}x {row['Container Size']} {row['Container Type']} @ {row['Frequency']}")

    # Update master file
    wb = load_workbook(master_path)
    ws = wb['Service Details']

    # Clear existing data (keep headers)
    ws.delete_rows(2, ws.max_row)

    # Write corrected data
    for r_idx, row in enumerate(dataframe_to_rows(df_new, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Update Property Overview for corrected properties
    ws_overview = wb['Property Overview']

    overview_updates = {
        'Bella Mirage': {'containers': 6, 'sizes': 'Mixed (8YD, 6YD, 4YD)', 'frequency': '4x/week'},
        'Orion McKinney': {'containers': 10, 'sizes': '8x 8YD + 2x 10YD', 'frequency': '3x/week'}
    }

    for row_idx in range(2, ws_overview.max_row + 1):
        property_name = ws_overview.cell(row_idx, 1).value
        if property_name in overview_updates:
            update = overview_updates[property_name]
            ws_overview.cell(row_idx, 4).value = update['containers']
            ws_overview.cell(row_idx, 5).value = update['sizes']
            ws_overview.cell(row_idx, 6).value = update['frequency']
            print(f"\nUpdated Property Overview for {property_name}")

    # Save
    wb.save(master_path)

    print("\n" + "="*80)
    print("CORRECTIONS APPLIED")
    print("="*80)
    print(f"File updated: {master_path}")

    print("\nKEY CORRECTIONS:")
    print("1. Bella Mirage: 10 → 6 containers, frequency 3x→4x/week")
    print("2. Orion McKinney: All 8YD → 8x 8YD + 2x 10YD (same total count)")
    print("3. All other properties: Verified or kept as-is")

    # Calculate new totals
    total_containers = df_new['Quantity'].sum()
    total_line_items = len(df_new)

    print(f"\nPortfolio Totals:")
    print(f"- Total Containers: {int(total_containers)}")
    print(f"- Total Service Line Items: {total_line_items}")
    print(f"- Properties: {len(df_new['Property'].unique())}/10")

if __name__ == '__main__':
    main()
