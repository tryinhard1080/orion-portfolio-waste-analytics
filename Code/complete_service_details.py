"""
Complete Service Details Sheet

Extract detailed service information from validated workbooks and Property Overview
to complete the Service Details sheet in the master file.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

def main():
    print("="*70)
    print("COMPLETING SERVICE DETAILS")
    print("="*70)

    master_path = 'Portfolio_Reports/MASTER_Portfolio_Complete_Data_UPDATED.xlsx'

    # Read current data
    df_overview = pd.read_excel(master_path, sheet_name='Property Overview')
    df_service_current = pd.read_excel(master_path, sheet_name='Service Details')

    print("\nCurrent Service Details has", len(df_service_current), "entries")
    print("Properties with service details:", df_service_current['Property'].unique().tolist())

    # Build complete service details from Property Overview
    service_details = []

    # Properties already in Service Details (keep them)
    existing_properties = df_service_current['Property'].unique()
    print("\nKeeping existing entries for:", existing_properties.tolist())

    # Add missing properties from Property Overview
    missing_properties = [
        'Orion Prosper',
        'McCord Park FL',
        'Orion McKinney',
        'The Club at Millenia',
        'Bella Mirage',
        'Orion Prosper Lakes'
    ]

    print("\nAdding service details for:")
    for prop in missing_properties:
        print(f"  - {prop}")

    # Extract from Property Overview
    for idx, row in df_overview.iterrows():
        property_name = row['Property Name']

        # Skip portfolio total row
        if 'PORTFOLIO' in str(property_name).upper() or pd.isna(property_name):
            continue

        # Skip if already in service details
        if property_name in existing_properties:
            continue

        # Only add if in missing list
        if property_name not in missing_properties:
            continue

        service_type = row['Service Type']
        container_count = row['Container Count']
        container_size = row['Container Size']
        frequency = row['Service Frequency']

        # Determine container type based on service type
        if 'Compactor' in str(service_type):
            container_type = 'Compactor'
        elif 'Dumpster' in str(service_type):
            container_type = 'Dumpster'
        elif 'Mixed' in str(service_type):
            container_type = 'Mixed'
        else:
            container_type = 'Unknown'

        # Parse container size
        if pd.notna(container_size):
            size_str = str(container_size)
        else:
            size_str = 'Unknown'

        # Parse frequency
        if pd.notna(frequency):
            freq_str = str(frequency)
        else:
            freq_str = 'Unknown'

        # Parse container count
        if pd.notna(container_count):
            qty = int(container_count) if container_count > 0 else 1
        else:
            qty = 1

        # Calculate total yards (approximate)
        total_yards = 0
        if container_type in ['Dumpster', 'Compactor']:
            try:
                # Extract numeric size
                size_num = int(''.join(filter(str.isdigit, str(size_str).split()[0])))
                total_yards = size_num * qty
            except:
                total_yards = 0

        service_details.append({
            'Property': property_name,
            'Container Type': container_type,
            'Container Size': size_str,
            'Quantity': qty,
            'Frequency': freq_str,
            'Total Yards': total_yards
        })

    # Create DataFrame with new entries
    df_new = pd.DataFrame(service_details)

    # Combine with existing
    df_complete = pd.concat([df_service_current, df_new], ignore_index=True)

    # Sort by property name
    df_complete = df_complete.sort_values('Property').reset_index(drop=True)

    print(f"\nComplete Service Details: {len(df_complete)} entries")
    print("\nAll properties:")
    for prop in sorted(df_complete['Property'].unique()):
        count = len(df_complete[df_complete['Property'] == prop])
        print(f"  - {prop}: {count} entry(ies)")

    # Update the workbook
    wb = load_workbook(master_path)
    ws = wb['Service Details']

    # Clear existing data (keep headers)
    ws.delete_rows(2, ws.max_row)

    # Write new data
    for r_idx, row in enumerate(dataframe_to_rows(df_complete, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Save
    wb.save(master_path)
    print(f"\nService Details updated successfully!")
    print(f"File: {master_path}")

    # Show summary
    print("\n" + "="*70)
    print("UPDATED SERVICE DETAILS SUMMARY")
    print("="*70)
    print(df_complete.to_string(index=False))


if __name__ == '__main__':
    main()
