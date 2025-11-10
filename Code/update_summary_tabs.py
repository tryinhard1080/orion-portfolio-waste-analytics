"""
Update summary tabs in master file with complete service details and YPD
"""

import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
import shutil

# Paths
BASE_DIR = Path(__file__).parent.parent
MASTER_FILE = BASE_DIR / "Portfolio_Reports" / "MASTER_Portfolio_Complete_Data.xlsx"
BACKUP_FILE = BASE_DIR / "Portfolio_Reports" / f"MASTER_Portfolio_Complete_Data_BACKUP_SUMMARY_UPDATE_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# Property data
PROPERTY_DATA = {
    'Orion Prosper': {
        'state': 'Texas',
        'units': 312,
        'service_type': 'Front Loader',
        'containers': 2,
        'container_size': '10 YD',
        'frequency': '6x/week',
        'ypd': 1.67,
        'monthly_yards': 519.60
    },
    'Orion Prosper Lakes': {
        'state': 'Texas',
        'units': 308,
        'service_type': 'Compactor',
        'containers': 1,
        'container_size': '40 CY',
        'frequency': '2.4x/week (on-call)',
        'ypd': 1.34,
        'monthly_yards': 412.00
    },
    'McCord Park FL': {
        'state': 'Florida',
        'units': 416,
        'service_type': 'Front Loader (FL)',
        'containers': 15,
        'container_size': '8 YD',
        'frequency': '3x/week',
        'ypd': 3.46,
        'monthly_yards': 1437.56
    },
    'Orion McKinney': {
        'state': 'Texas',
        'units': 453,
        'service_type': 'Front Loader (FL)',
        'containers': 10,
        'container_size': '8 YD',
        'frequency': '3x/week',
        'ypd': 2.41,
        'monthly_yards': 1091.16
    },
    'The Club at Millenia': {
        'state': 'Florida',
        'units': 560,
        'service_type': 'Dumpster',
        'containers': 6,
        'container_size': '8 YD',
        'frequency': '4x/week',
        'ypd': 1.30,
        'monthly_yards': 727.44
    },
    'Bella Mirage': {
        'state': 'Arizona',
        'units': 715,
        'service_type': 'Front End Loader (FEL)',
        'containers': 6,
        'container_size': '8 YD',
        'frequency': '3x/week',
        'ypd': 0.87,
        'monthly_yards': 623.52
    },
    'Mandarina': {
        'state': 'Arizona',
        'units': 180,
        'service_type': 'Compactor',
        'containers': 2,
        'container_size': '6 YD',
        'frequency': '3x/week',
        'ypd': 0.87,
        'monthly_yards': 155.88
    },
    'Pavilions at Arrowhead': {
        'state': 'Arizona',
        'units': 248,
        'service_type': 'Dumpster',
        'containers': 4,
        'container_size': '4 YD',
        'frequency': '2x/week',
        'ypd': 0.56,
        'monthly_yards': 138.56
    },
    'Springs at Alta Mesa': {
        'state': 'Arizona',
        'units': 200,
        'service_type': 'Mixed',
        'containers': 16,
        'container_size': 'Mixed',
        'frequency': '3x/week',
        'ypd': 3.00,
        'monthly_yards': 600.63
    },
    'Tempe Vista': {
        'state': 'Arizona',
        'units': 186,
        'service_type': 'Mixed',
        'containers': 8,
        'container_size': 'Mixed',
        'frequency': '3x/week',
        'ypd': 2.65,
        'monthly_yards': 493.62
    }
}

def create_backup():
    """Create backup before updating"""
    print('=' * 80)
    print('CREATING BACKUP')
    print('=' * 80)
    print()
    
    if MASTER_FILE.exists():
        shutil.copy2(MASTER_FILE, BACKUP_FILE)
        print(f'✅ Backup created: {BACKUP_FILE.name}')
        print()
        return True
    else:
        print(f'❌ Master file not found: {MASTER_FILE}')
        return False

def update_property_overview():
    """Update Property Overview tab"""
    print('Updating: Property Overview')
    print('-' * 80)
    
    wb = openpyxl.load_workbook(MASTER_FILE)
    
    if 'Property Overview' not in wb.sheetnames:
        print('  ⚠️ Sheet not found')
        wb.close()
        return False
    
    ws = wb['Property Overview']
    
    # Find or add columns
    headers = []
    for col in range(1, 20):
        header = ws.cell(1, col).value
        if header:
            headers.append(header)
        else:
            break
    
    # Add missing columns
    new_columns = ['Container Count', 'Container Size', 'Service Frequency', 'Monthly Yards', 'YPD']
    
    col_map = {}
    next_col = len(headers) + 1
    
    for new_col in new_columns:
        if new_col in headers:
            col_map[new_col] = headers.index(new_col) + 1
        else:
            ws.cell(1, next_col).value = new_col
            col_map[new_col] = next_col
            next_col += 1
    
    # Update data rows
    row = 2
    rows_updated = 0
    
    while ws.cell(row, 1).value:
        prop_name = ws.cell(row, 1).value
        
        if prop_name in PROPERTY_DATA:
            data = PROPERTY_DATA[prop_name]
            
            ws.cell(row, col_map['Container Count']).value = data['containers']
            ws.cell(row, col_map['Container Size']).value = data['container_size']
            ws.cell(row, col_map['Service Frequency']).value = data['frequency']
            ws.cell(row, col_map['Monthly Yards']).value = round(data['monthly_yards'], 2)
            ws.cell(row, col_map['YPD']).value = round(data['ypd'], 2)
            
            rows_updated += 1
        
        row += 1
    
    wb.save(MASTER_FILE)
    wb.close()
    
    print(f'  ✅ Updated {rows_updated} properties')
    print()
    
    return True

def update_yards_per_door_tab():
    """Update Yards Per Door tab"""
    print('Updating: Yards Per Door')
    print('-' * 80)
    
    wb = openpyxl.load_workbook(MASTER_FILE)
    
    if 'Yards Per Door' not in wb.sheetnames:
        print('  ⚠️ Sheet not found')
        wb.close()
        return False
    
    ws = wb['Yards Per Door']
    
    # Find or add columns
    headers = []
    for col in range(1, 20):
        header = ws.cell(1, col).value
        if header:
            headers.append(header)
        else:
            break
    
    # Ensure we have the right columns
    required_columns = ['Property', 'Units', 'Containers', 'Container Size', 'Monthly Yards', 'YPD', 'Performance']
    
    col_map = {}
    next_col = len(headers) + 1
    
    for req_col in required_columns:
        if req_col in headers:
            col_map[req_col] = headers.index(req_col) + 1
        else:
            ws.cell(1, next_col).value = req_col
            col_map[req_col] = next_col
            next_col += 1
    
    # Update data rows
    row = 2
    rows_updated = 0
    
    while ws.cell(row, 1).value:
        prop_name = ws.cell(row, 1).value
        
        if prop_name in PROPERTY_DATA:
            data = PROPERTY_DATA[prop_name]
            
            ws.cell(row, col_map['Units']).value = data['units']
            ws.cell(row, col_map['Containers']).value = data['containers']
            ws.cell(row, col_map['Container Size']).value = data['container_size']
            ws.cell(row, col_map['Monthly Yards']).value = round(data['monthly_yards'], 2)
            ws.cell(row, col_map['YPD']).value = round(data['ypd'], 2)
            
            # Add performance rating
            ypd = data['ypd']
            if ypd <= 2.0:
                performance = 'Excellent'
            elif ypd <= 2.25:
                performance = 'Good'
            else:
                performance = 'High'
            
            ws.cell(row, col_map['Performance']).value = performance
            
            rows_updated += 1
        
        row += 1
    
    wb.save(MASTER_FILE)
    wb.close()
    
    print(f'  ✅ Updated {rows_updated} properties')
    print()
    
    return True

def main():
    """Main function"""
    
    print('=' * 80)
    print('UPDATING SUMMARY TABS WITH SERVICE DETAILS')
    print('=' * 80)
    print()
    
    # Create backup
    if not create_backup():
        print('❌ Cannot proceed without backup')
        return
    
    # Update tabs
    print('=' * 80)
    print('UPDATING TABS')
    print('=' * 80)
    print()
    
    update_property_overview()
    update_yards_per_door_tab()
    
    # Summary
    print('=' * 80)
    print('UPDATE COMPLETE')
    print('=' * 80)
    print()
    print('Summary tabs updated with:')
    print('  ✅ Container counts')
    print('  ✅ Container sizes')
    print('  ✅ Service frequencies')
    print('  ✅ Monthly yards')
    print('  ✅ YPD calculations')
    print('  ✅ Performance ratings')
    print()

if __name__ == "__main__":
    main()

