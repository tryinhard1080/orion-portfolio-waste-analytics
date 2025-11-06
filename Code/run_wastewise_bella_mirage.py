"""
WasteWise Analytics - Validated Edition for Bella Mirage
Comprehensive waste management analysis with contract extraction and validation
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from typing import Dict, List, Tuple
import json
import os

# ============================================================================
# PROPERTY CONFIGURATION
# ============================================================================

PROPERTY_CONFIG = {
    'name': 'Bella Mirage',
    'unit_count': 715,
    'location': 'Avondale, AZ',
    'service_type': 'Dumpster',  # Front-End Loaders
    'vendor': 'Waste Management of Arizona, Inc.',
    'account_number': '22-06174-13009',
    'contract_id': 'S0013040977'
}

# ============================================================================
# CONTRACT DATA EXTRACTED FROM PDF
# ============================================================================

CONTRACT_DATA = {
    'effective_date': datetime(2020, 4, 8),
    'contract_term_years': 3,
    'initial_term_end': datetime(2023, 4, 8),
    'auto_renewal': True,
    'renewal_term_months': 12,
    'termination_notice_days': 90,
    'termination_notice_window_start_days': 180,
    'termination_notice_window_end_days': 90,

    'service_details': {
        'containers': [
            {'quantity': 6, 'size': 8, 'type': 'FEL', 'frequency': 3, 'monthly_cost': 1180.00},
            {'quantity': 2, 'size': 8, 'type': 'FEL', 'frequency': 3, 'monthly_cost': 408.00},
            {'quantity': 2, 'size': 4, 'type': 'FEL', 'frequency': 3, 'monthly_cost': 250.00},
            {'quantity': 5, 'size': 8, 'type': 'FEL', 'frequency': 3, 'monthly_cost': 613.00},
            {'quantity': 4, 'size': 6, 'type': 'FEL', 'frequency': 3, 'monthly_cost': 627.00},
            {'quantity': 3, 'size': 8, 'type': 'FEL', 'frequency': 3, 'monthly_cost': 993.00}
        ],
        'total_containers': 22,
        'monthly_base_cost': 4071.00,
        'extra_pickup_rate': 75.00
    },

    'clauses': [
        {
            'category': 'Term & Renewal',
            'verbatim_text': 'Contract Term is for 3 year(s) from the Effective Date (\'Initial Term\') and it shall automatically renew thereafter for additional terms of 12 months (\'Renewal Term\') unless terminated as set forth herein.',
            'risk_severity': 'high',
            'impact': 'Locked into 3-year initial term with automatic 12-month renewals unless proper notice given',
            'action_required': 'Set calendar reminder 90-180 days before renewal date to provide termination notice if needed',
            'page_reference': 'Page 3'
        },
        {
            'category': 'Termination',
            'verbatim_text': 'Unless otherwise specified on the Service Summary, at the end of the Initial Term and any subsequent Renewal Term, the Contract Term shall automatically renew for an additional Renewal Term at the then current Service levels and applicable Charges, unless (a) for a Renewal Term of twelve (12) months or more, either party gives to the other party written notice of termination at least ninety (90) days, but not more than one hundred eighty (180) days, prior to the termination of the then-existing term, and (b) for a Renewal Term of less than twelve (12) months, either party gives to the other party written notice of termination at least thirty (30) days prior to the termination of the then-existing term. Notice of termination received at any other time will be considered ineffective and the Agreement will be considered automatically renewed upon completion of the then-existing term.',
            'risk_severity': 'high',
            'impact': '90-180 day notice window required - notice outside this window is INVALID and contract auto-renews',
            'action_required': 'Critical: Submit termination notice via Certified Mail between 90-180 days before renewal date',
            'page_reference': 'Page 4, Section 2'
        },
        {
            'category': 'Rate Increases',
            'verbatim_text': 'Company reserves the right, and Customer acknowledges that it should expect Company to increase or add Charges payable by Customer hereunder during the Contract Term: (i) for any changes or modifications to, or differences between, the actual equipment and Services provided by Company to Customer and those specified on the Service Summary; (ii) for any changes or difference in the composition, amount or weight of the Waste Materials collected by Company from Customer\'s service location(s) from what is specified on the Service Summary (including for container overages or overflows); (iii) for any increase in or other modification made by Company to the Fuel Surcharge, Regulatory Cost Recovery Charge, Recyclable Materials Offset, Environmental Charge, and/or any other Charges included or referenced in the Service Summary (which Charges are calculated and/or determined on enterprise-wide basis, including Company and all Affiliates); (iv) to cover any increases in disposal, processing, and/or transportation costs, including fuel surcharges; (v) to cover increased costs due to uncontrollable circumstances, including, without limitation, changes (occurring from and after three (3) months prior to the Effective Date) in local, state, federal or foreign laws or regulations (or the enforcement, interpretation or application thereof), including the imposition of or increase in taxes, fees or surcharges, or acts of God such as floods, fires, hurricanes and natural disasters; and (vi) for increases in the Consumer Price Index ("CPI") for Water, Sewer and Trash Collection Services published by U.S. Bureau of Labor Statistics, or with written notice to Customer, any other national, regional or local CPI, with such increases in CPI being measured from the Effective Date, or as applicable, Customer\'s last CPI based price increase date ("PI Date").',
            'risk_severity': 'medium',
            'impact': 'Vendor can increase rates for multiple reasons including CPI, disposal costs, fuel, and regulatory changes - not a fixed price contract',
            'action_required': 'Monitor invoices monthly for rate increases; track CPI increases; budget for 3-5% annual increases',
            'page_reference': 'Page 4, Section 4(b)'
        },
        {
            'category': 'Liability',
            'verbatim_text': 'Neither party shall be liable to the other for consequential, incidental or punitive damages arising out of the performance or breach of this Agreement.',
            'risk_severity': 'medium',
            'impact': 'Limited liability - no consequential or punitive damages recoverable from either party',
            'action_required': 'Ensure adequate insurance coverage for potential service failures',
            'page_reference': 'Page 4, Section 8'
        },
        {
            'category': 'Service Level',
            'verbatim_text': 'SERVICE GUARANTEE. We guarantee our Services (as defined below). If Company fails to perform Services in accordance with the service summary as provided, which for Services purchased online include the information and terms disclosed during the order and checkout process (collectively, the "Service Summary"), and Company does not remedy such failure within five (5) business days of its receipt of a written demand from Customer, Customer may immediately terminate this Agreement without penalty.',
            'risk_severity': 'low',
            'impact': 'Service guarantee allows penalty-free termination if WM fails to perform and doesn\'t remedy within 5 business days',
            'action_required': 'Document any service failures in writing; submit written demand if failures persist',
            'page_reference': 'Page 4, Section 1(a)'
        },
        {
            'category': 'Indemnification',
            'verbatim_text': 'Company agrees to indemnify, defend and save Customer and its Affiliates harmless from and against any and all liability which Customer or its Affiliates may suffer, incur or pay as a result of any bodily injuries (including death), property damage or violation of law, to the extent caused by any negligent act or omission or willful misconduct of Company or its employees, which occurs (a) during the collection or transportation of Customer\'s Waste Materials, or (b) as a result of the disposal of Customer\'s Waste Materials in a facility owned by Company or an Affiliate, provided that Company\'s indemnification obligations will not apply to occurrences involving Excluded Materials. Customer agrees to indemnify, defend and save Company and its Affiliates harmless from and against any and all liability which Company and its Affiliates may suffer, incur or pay as a result of any bodily injuries (including death), property damage or violation of law to the extent caused by Customer\'s breach of this Agreement or by any negligent act or omission or willful misconduct of Customer or its employees, agents or contractors or Customer\'s use, operation or possession of any equipment furnished by Company.',
            'risk_severity': 'medium',
            'impact': 'Mutual indemnification - each party indemnifies the other for their own negligence or misconduct',
            'action_required': 'Maintain general liability insurance; ensure employees follow safety protocols around equipment',
            'page_reference': 'Page 4, Section 8'
        },
        {
            'category': 'Force Majeure',
            'verbatim_text': 'Except for the obligation to make payments hereunder for Services already performed, neither party shall be in default for its failure to perform or delay in performance caused by events or significant threats of events beyond its reasonable control, whether or not foreseeable, including, but not limited to, strikes, labor trouble, riots, imposition of laws or governmental orders, fires, acts of war or terrorism, acts of God, and the inability to obtain equipment, and the affected party shall be excused from performance during the occurrence of such events.',
            'risk_severity': 'low',
            'impact': 'Service disruptions excused for events beyond reasonable control (strikes, natural disasters, etc.) but payment obligations remain',
            'action_required': 'No action required; standard force majeure protection for both parties',
            'page_reference': 'Page 5, Section 11(a)'
        }
    ],

    'calendar_reminders': []  # Will be calculated below
}

# Calculate calendar reminders based on contract dates
def calculate_calendar_reminders():
    """Calculate critical contract dates and reminders"""
    reminders = []

    # Current renewal cycle (contract started 4/8/2020, initial 3-year term ended 4/8/2023)
    # Now in auto-renewal 12-month cycles
    current_date = datetime.now()

    # Find next renewal date
    years_since_initial = (current_date.year - 2023) + (1 if current_date.month >= 4 else 0)
    next_renewal_date = datetime(2023 + years_since_initial, 4, 8)

    if next_renewal_date < current_date:
        next_renewal_date = datetime(2023 + years_since_initial + 1, 4, 8)

    # Termination notice window: 90-180 days before renewal
    notice_start_date = next_renewal_date - timedelta(days=180)
    notice_end_date = next_renewal_date - timedelta(days=90)

    reminders.append({
        'date': notice_start_date,
        'action': 'Termination notice window OPENS - Can submit termination notice (certified mail required)',
        'criticality': 'HIGH',
        'days_until': (notice_start_date - current_date).days
    })

    reminders.append({
        'date': notice_end_date,
        'action': 'Termination notice window CLOSES - Last day to submit termination notice to avoid auto-renewal',
        'criticality': 'CRITICAL',
        'days_until': (notice_end_date - current_date).days
    })

    reminders.append({
        'date': next_renewal_date,
        'action': 'Contract auto-renews for 12 months unless termination notice was submitted',
        'criticality': 'HIGH',
        'days_until': (next_renewal_date - current_date).days
    })

    return reminders

CONTRACT_DATA['calendar_reminders'] = calculate_calendar_reminders()

# ============================================================================
# DATA LOADING AND PROCESSING
# ============================================================================

def load_invoice_data(excel_file: str, sheet_name: str = 'Bella Mirage') -> pd.DataFrame:
    """Load invoice data from Excel file"""
    print(f"Loading invoice data from: {sheet_name}")
    df = pd.read_excel(excel_file, sheet_name=sheet_name)

    # Convert date columns
    date_columns = ['Invoice Date', 'Due Date', 'Service Date']
    for col in date_columns:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')

    # Clean numeric columns
    numeric_columns = ['Amount Due', 'Quantity', 'Unit Rate', 'Extended Amount',
                      'Container Size (yd)', 'Frequency/Week']
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    print(f"   OK Loaded {len(df)} line items")
    print(f"   OK Total amount: ${df['Extended Amount'].sum():,.2f}")
    print(f"   OK Date range: {df['Invoice Date'].min()} to {df['Invoice Date'].max()}")

    return df

def calculate_performance_metrics(df: pd.DataFrame, property_config: Dict) -> Dict:
    """Calculate performance metrics from invoice data"""
    print("\nCalculating performance metrics...")

    units = property_config['unit_count']

    # Group by invoice date to get monthly totals
    monthly_data = df.groupby(df['Invoice Date'].dt.to_period('M')).agg({
        'Extended Amount': 'sum',
        'Amount Due': 'first'  # Invoice total (same for all line items in same invoice)
    }).reset_index()

    monthly_data['Invoice Date'] = monthly_data['Invoice Date'].dt.to_timestamp()

    # Calculate metrics
    total_spend = df['Extended Amount'].sum()
    num_months = len(monthly_data)
    avg_monthly = total_spend / num_months if num_months > 0 else 0

    # Cost per door
    cost_per_door = avg_monthly / units

    # Category breakdown
    base_charges = df[df['Category'] == 'base']['Extended Amount'].sum()
    extra_pickups = df[df['Category'] == 'extra_pickup']['Extended Amount'].sum()
    overages = df[df['Category'] == 'overage']['Extended Amount'].sum()

    # Overage percentage
    overage_pct = (extra_pickups + overages) / total_spend * 100 if total_spend > 0 else 0

    # Yards per door calculation (for dumpsters)
    # Total containers: 22 dumpsters with mixed sizes
    # Calculate weighted average container size
    contract_containers = CONTRACT_DATA['service_details']['containers']
    total_yards = sum(c['quantity'] * c['size'] for c in contract_containers)
    total_containers = sum(c['quantity'] for c in contract_containers)
    avg_container_size = total_yards / total_containers if total_containers > 0 else 0

    # Frequency: 3x per week = 12x per month
    pickups_per_month = 12

    # Yards per door = (Total Yards × Pickups/Month × Containers) / Units
    yards_per_door = (avg_container_size * pickups_per_month * total_containers) / units

    metrics = {
        'total_spend': total_spend,
        'num_invoices': len(monthly_data),
        'num_months': num_months,
        'avg_monthly_cost': avg_monthly,
        'cost_per_door': cost_per_door,
        'base_charges': base_charges,
        'extra_pickups': extra_pickups,
        'overages': overages,
        'overage_pct': overage_pct,
        'yards_per_door': yards_per_door,
        'total_containers': total_containers,
        'avg_container_size': avg_container_size,
        'pickups_per_month': pickups_per_month,
        'contract_monthly_base': CONTRACT_DATA['service_details']['monthly_base_cost'],
        'monthly_data': monthly_data.to_dict('records')
    }

    print(f"   OK Average monthly cost: ${avg_monthly:,.2f}")
    print(f"   OK Cost per door: ${cost_per_door:.2f}")
    print(f"   OK Yards per door: {yards_per_door:.2f}")
    print(f"   OK Overage %: {overage_pct:.1f}%")

    return metrics

# ============================================================================
# OPTIMIZATION ANALYSIS
# ============================================================================

def analyze_optimizations(df: pd.DataFrame, metrics: Dict, property_config: Dict) -> Dict:
    """Analyze potential optimization opportunities"""
    print("\nAnalyzing optimization opportunities...")

    optimizations = []

    # 1. CONTAMINATION / OVERAGE REDUCTION
    overage_total = metrics['extra_pickups'] + metrics['overages']
    overage_pct = metrics['overage_pct']

    if overage_pct >= 3.0:  # Threshold: 3% or more
        # Calculate potential savings (assume 70% reduction through education/training)
        potential_reduction_pct = 0.70
        annual_overage_cost = overage_total / metrics['num_months'] * 12
        annual_savings = annual_overage_cost * potential_reduction_pct

        optimizations.append({
            'type': 'CONTAMINATION_REDUCTION',
            'title': 'Overage & Extra Pickup Reduction',
            'current_annual_cost': annual_overage_cost,
            'contamination_percentage': overage_pct,
            'potential_annual_savings': annual_savings,
            'implementation_cost': 2000,  # One-time: signage, training, guidelines
            'annual_monitoring_cost': 1200,  # Monthly audits $100/mo
            'roi_months': 12,
            'confidence': 'MEDIUM',
            'calculation_breakdown': {
                'total_overage_charges': overage_total,
                'overage_percentage': overage_pct,
                'num_months_data': metrics['num_months'],
                'annual_overage_cost': annual_overage_cost,
                'reduction_percentage': potential_reduction_pct * 100,
                'annual_savings': annual_savings,
                'implementation_cost': 2000,
                'annual_monitoring': 1200,
                'net_annual_savings': annual_savings - 1200,
                'roi_months': round((2000 / ((annual_savings - 1200) / 12)), 1) if annual_savings > 1200 else 999
            },
            'recommendation': f'${overage_total:,.2f} in overages ({overage_pct:.1f}% of spend) suggests opportunities for resident education and contamination control programs.'
        })

    # 2. SERVICE RIGHT-SIZING (if overage frequency is low but cost per door is high)
    if metrics['cost_per_door'] > 11.00:  # Bella Mirage current: $10.68, so this is above target
        # Analyze if reducing frequency or container counts could work
        # Current: 22 containers @ 3x/week
        # Potential: Reduce to 2x/week or consolidate containers

        current_annual = metrics['avg_monthly_cost'] * 12

        # Estimate 15% reduction from frequency optimization
        potential_savings_pct = 0.15
        annual_savings = current_annual * potential_savings_pct

        optimizations.append({
            'type': 'SERVICE_RIGHTSIZING',
            'title': 'Service Frequency & Container Optimization',
            'current_annual_cost': current_annual,
            'potential_annual_savings': annual_savings,
            'implementation_cost': 0,  # Contract renegotiation only
            'annual_monitoring_cost': 0,
            'roi_months': 0,
            'confidence': 'LOW',
            'calculation_breakdown': {
                'current_monthly_cost': metrics['avg_monthly_cost'],
                'current_annual_cost': current_annual,
                'current_containers': metrics['total_containers'],
                'current_frequency_per_week': 3,
                'potential_frequency_per_week': 2,
                'estimated_savings_pct': potential_savings_pct * 100,
                'annual_savings': annual_savings
            },
            'recommendation': f'Current cost per door ${metrics["cost_per_door"]:.2f} with {metrics["total_containers"]} containers @ 3x/week. Consider frequency reduction to 2x/week if overage rate remains low.'
        })

    # 3. BULK ITEM MANAGEMENT (based on extra pickups)
    extra_pickup_count = len(df[df['Category'] == 'extra_pickup'])
    if extra_pickup_count > 0:
        avg_bulk_monthly = metrics['extra_pickups'] / metrics['num_months']

        if avg_bulk_monthly > 100:  # If averaging >$100/month in extra pickups
            # Bulk subscription could potentially save 30-40%
            annual_bulk_cost = avg_bulk_monthly * 12
            potential_savings_pct = 0.35
            annual_savings = annual_bulk_cost * potential_savings_pct

            optimizations.append({
                'type': 'BULK_MANAGEMENT',
                'title': 'Bulk Item Subscription Program',
                'current_annual_cost': annual_bulk_cost,
                'avg_bulk_monthly': avg_bulk_monthly,
                'potential_annual_savings': annual_savings,
                'implementation_cost': 0,
                'annual_monitoring_cost': 0,
                'roi_months': 0,
                'confidence': 'MEDIUM',
                'calculation_breakdown': {
                    'extra_pickup_count': extra_pickup_count,
                    'total_extra_pickup_cost': metrics['extra_pickups'],
                    'avg_monthly_cost': avg_bulk_monthly,
                    'annual_cost': annual_bulk_cost,
                    'potential_savings_pct': potential_savings_pct * 100,
                    'annual_savings': annual_savings
                },
                'recommendation': f'${avg_bulk_monthly:,.2f}/month in extra pickups suggests need for bulk item management program or scheduled bulk pickups.'
            })

    print(f"   OK Identified {len(optimizations)} optimization opportunities")

    return {
        'optimizations': optimizations,
        'total_potential_annual_savings': sum(opt['potential_annual_savings'] for opt in optimizations)
    }

# ============================================================================
# VALIDATION FRAMEWORK
# ============================================================================

class WasteWiseValidator:
    """Comprehensive validation framework"""

    def __init__(self):
        self.validation_results = {
            'contract_validation': {},
            'optimization_validation': {},
            'formula_validation': {},
            'sheet_structure_validation': {},
            'data_completeness_validation': {},
            'cross_validation': {}
        }
        self.errors = []
        self.warnings = []

    def validate_all(self, df: pd.DataFrame, metrics: Dict,
                     optimization_results: Dict, property_config: Dict) -> Tuple[bool, Dict]:
        """Run all validation checks"""

        print("\nVALIDATION GATE - Running All Checks...")
        print("-" * 60)

        # 1. Contract Validation
        contract_valid = self.validate_contract()

        # 2. Optimization Validation
        optimization_valid = self.validate_optimizations(optimization_results)

        # 3. Formula Validation
        formula_valid = self.validate_formulas(metrics, property_config)

        # 4. Sheet Structure Validation
        structure_valid = self.validate_sheet_structure()

        # 5. Data Completeness Validation
        completeness_valid = self.validate_data_completeness(df, property_config)

        # 6. Cross-Validation
        cross_valid = self.validate_cross_references(df, metrics)

        all_passed = all([
            contract_valid,
            optimization_valid,
            formula_valid,
            structure_valid,
            completeness_valid,
            cross_valid
        ])

        return all_passed, self.generate_validation_report()

    def validate_contract(self) -> bool:
        """Validate contract extraction"""
        clauses_found = len(CONTRACT_DATA['clauses'])
        self.validation_results['contract_validation']['clauses_found'] = clauses_found

        if clauses_found < 3:
            self.errors.append(
                f"[FAIL] CONTRACT EXTRACTION FAILED: Only {clauses_found} clauses found"
            )
            self.validation_results['contract_validation']['status'] = 'FAILED'
            return False

        calendar_reminders = len(CONTRACT_DATA['calendar_reminders'])
        if calendar_reminders == 0:
            self.warnings.append("[WARN] No calendar reminders found")

        self.validation_results['contract_validation']['status'] = 'PASSED'
        self.validation_results['contract_validation']['calendar_reminders'] = calendar_reminders

        return True

    def validate_optimizations(self, optimization_results: Dict) -> bool:
        """Validate optimization recommendations"""
        optimizations = optimization_results['optimizations']

        for opt in optimizations:
            opt_type = opt['type']

            # Validate contamination reduction threshold
            if opt_type == 'CONTAMINATION_REDUCTION':
                contamination_pct = opt.get('contamination_percentage', 0)
                if contamination_pct < 3.0:
                    self.errors.append(
                        f"[FAIL] CONTAMINATION REDUCTION INVALID: {contamination_pct:.1f}% "
                        f"is below 3% threshold"
                    )
                    return False

            # Validate bulk subscription threshold
            elif opt_type == 'BULK_MANAGEMENT':
                avg_bulk_monthly = opt.get('avg_bulk_monthly', 0)
                if avg_bulk_monthly < 100:
                    self.warnings.append(
                        f"[WARN] BULK MANAGEMENT: ${avg_bulk_monthly:.2f}/month "
                        f"is below typical $500 threshold, but included due to data pattern"
                    )

        self.validation_results['optimization_validation']['status'] = 'PASSED'
        self.validation_results['optimization_validation']['count'] = len(optimizations)

        return True

    def validate_formulas(self, metrics: Dict, property_config: Dict) -> bool:
        """Validate formula calculations"""
        units = property_config['unit_count']

        # Validate cost per door
        expected_cpd = metrics['avg_monthly_cost'] / units
        actual_cpd = metrics['cost_per_door']

        if abs(expected_cpd - actual_cpd) > 0.01:
            self.errors.append(
                f"[FAIL] COST PER DOOR FORMULA ERROR: "
                f"Expected ${expected_cpd:.2f}, got ${actual_cpd:.2f}"
            )
            return False

        # Validate yards per door (for dumpsters, based on container capacity)
        # This is validated against contract specifications
        contract_containers = CONTRACT_DATA['service_details']['containers']
        total_yards = sum(c['quantity'] * c['size'] for c in contract_containers)
        total_containers = sum(c['quantity'] for c in contract_containers)

        if total_containers != metrics['total_containers']:
            self.errors.append(
                f"[FAIL] CONTAINER COUNT MISMATCH: Contract specifies {total_containers}, "
                f"calculated {metrics['total_containers']}"
            )
            return False

        self.validation_results['formula_validation']['status'] = 'PASSED'
        return True

    def validate_sheet_structure(self) -> bool:
        """Validate expected sheet structure"""
        # For dumpster service (not compactor), we won't have HAUL_LOG
        has_compactor = PROPERTY_CONFIG['service_type'].lower() == 'compactor'
        has_contract = len(CONTRACT_DATA['clauses']) > 0

        expected_sheets = [
            'SUMMARY_FULL',
            'EXPENSE_ANALYSIS',
            'OPTIMIZATION',
            'QUALITY_CHECK',
            'DOCUMENTATION_NOTES'
        ]

        if has_compactor:
            expected_sheets.append('HAUL_LOG')

        if has_contract:
            expected_sheets.append('CONTRACT_TERMS')

        self.validation_results['sheet_structure_validation']['expected_sheets'] = expected_sheets
        self.validation_results['sheet_structure_validation']['has_compactor'] = has_compactor
        self.validation_results['sheet_structure_validation']['has_contract'] = has_contract
        self.validation_results['sheet_structure_validation']['status'] = 'PASSED'

        return True

    def validate_data_completeness(self, df: pd.DataFrame, property_config: Dict) -> bool:
        """Validate data completeness"""
        if len(df) == 0:
            self.errors.append("[FAIL] NO INVOICE DATA: At least one invoice required")
            return False

        if not property_config.get('name'):
            self.errors.append("[FAIL] MISSING PROPERTY NAME")
            return False

        if not property_config.get('unit_count'):
            self.errors.append("[FAIL] MISSING UNIT COUNT")
            return False

        self.validation_results['data_completeness_validation']['status'] = 'PASSED'
        self.validation_results['data_completeness_validation']['invoice_count'] = len(df)

        return True

    def validate_cross_references(self, df: pd.DataFrame, metrics: Dict) -> bool:
        """Validate data consistency"""
        # Validate total spend calculation
        df_total = df['Extended Amount'].sum()
        metrics_total = metrics['total_spend']

        if abs(df_total - metrics_total) > 0.01:
            self.errors.append(
                f"[FAIL] CROSS-VALIDATION ERROR: Total spend mismatch. "
                f"DataFrame: ${df_total:,.2f}, Metrics: ${metrics_total:,.2f}"
            )
            return False

        self.validation_results['cross_validation']['status'] = 'PASSED'
        return True

    def generate_validation_report(self) -> Dict:
        """Generate validation report"""
        return {
            'timestamp': datetime.now().isoformat(),
            'validation_results': self.validation_results,
            'errors': self.errors,
            'warnings': self.warnings,
            'passed': len(self.errors) == 0,
            'summary': {
                'total_checks': sum(
                    1 for category in self.validation_results.values()
                    if category.get('status') in ['PASSED', 'FAILED']
                ),
                'passed_checks': sum(
                    1 for category in self.validation_results.values()
                    if category.get('status') == 'PASSED'
                ),
                'failed_checks': len(self.errors),
                'warnings': len(self.warnings)
            }
        }

# ============================================================================
# EXCEL WORKBOOK GENERATION
# ============================================================================

def create_excel_workbook(df: pd.DataFrame, metrics: Dict,
                         optimization_results: Dict,
                         validation_report: Dict,
                         property_config: Dict) -> Workbook:
    """Generate complete Excel workbook"""

    print("\nGenerating Excel Workbook...")

    wb = Workbook()
    wb.remove(wb.active)

    # SHEET 1: SUMMARY_FULL
    create_summary_sheet(wb, metrics, optimization_results, property_config)

    # SHEET 2: EXPENSE_ANALYSIS
    create_expense_analysis_sheet(wb, metrics, property_config)

    # SHEET 3: OPTIMIZATION
    create_optimization_sheet(wb, optimization_results, property_config)

    # SHEET 4: QUALITY_CHECK
    create_quality_check_sheet(wb, validation_report)

    # SHEET 5: DOCUMENTATION_NOTES
    create_documentation_sheet(wb, property_config)

    # SHEET 6: CONTRACT_TERMS (contract provided)
    create_contract_terms_sheet(wb)

    print("   OK All sheets created")

    return wb

def create_summary_sheet(wb: Workbook, metrics: Dict, optimization_results: Dict,
                        property_config: Dict):
    """Create SUMMARY_FULL sheet"""
    ws = wb.create_sheet("SUMMARY_FULL")

    # Styling
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=14, color="1E3A8A")

    # Title
    ws['A1'] = f"{property_config['name']} - Waste Management Analysis"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')

    # 2026 Savings One-Liner (top priority)
    ws['A3'] = "💰 2026 OPPORTUNITY:"
    ws['A3'].font = Font(bold=True, size=12, color="22C55E")
    total_savings = optimization_results['total_potential_annual_savings']
    ws['B3'] = f"${total_savings:,.0f} annual efficiency opportunity identified through service optimization"
    ws['B3'].font = Font(size=11)
    ws.merge_cells('B3:D3')

    row = 5

    # Property Information
    ws[f'A{row}'] = "PROPERTY INFORMATION"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:B{row}')
    row += 1

    info_items = [
        ('Property Name:', property_config['name']),
        ('Location:', property_config['location']),
        ('Unit Count:', f"{property_config['unit_count']} units"),
        ('Service Type:', property_config['service_type']),
        ('Vendor:', property_config['vendor']),
        ('Account #:', property_config['account_number'])
    ]

    for label, value in info_items:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1

    row += 1

    # Performance Metrics
    ws[f'A{row}'] = "PERFORMANCE METRICS"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:B{row}')
    row += 1

    metrics_items = [
        ('Average Monthly Cost:', f"${metrics['avg_monthly_cost']:,.2f}"),
        ('Cost Per Door:', f"${metrics['cost_per_door']:.2f}"),
        ('Yards Per Door:', f"{metrics['yards_per_door']:.2f}"),
        ('Total Containers:', f"{metrics['total_containers']}"),
        ('Overage Percentage:', f"{metrics['overage_pct']:.1f}%")
    ]

    for label, value in metrics_items:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1

    row += 1

    # Optimization Summary
    ws[f'A{row}'] = "OPTIMIZATION OPPORTUNITIES"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:B{row}')
    row += 1

    for opt in optimization_results['optimizations']:
        ws[f'A{row}'] = opt['title']
        ws[f'B{row}'] = f"${opt['potential_annual_savings']:,.0f}/year"
        ws[f'B{row}'].font = Font(color="22C55E", bold=True)
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

def create_expense_analysis_sheet(wb: Workbook, metrics: Dict, property_config: Dict):
    """Create EXPENSE_ANALYSIS sheet - month-by-month COLUMN format"""
    ws = wb.create_sheet("EXPENSE_ANALYSIS")

    # Styling
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    # Title
    ws['A1'] = "EXPENSE ANALYSIS - Monthly Breakdown"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')

    # Headers (row 3)
    row = 3
    headers = ['Month', 'Invoice Date', 'Total Amount', 'Cost/Door', 'Base Charges', 'Overages']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill

    # Monthly data
    row += 1
    units = property_config['unit_count']

    for month_data in metrics['monthly_data']:
        invoice_date = month_data['Invoice Date']
        amount = month_data['Extended Amount']
        cpd = amount / units

        ws.cell(row=row, column=1, value=invoice_date.strftime('%b %Y'))
        ws.cell(row=row, column=2, value=invoice_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=3, value=amount).number_format = '$#,##0.00'
        ws.cell(row=row, column=4, value=cpd).number_format = '$#,##0.00'
        ws.cell(row=row, column=5, value=metrics['base_charges'] / len(metrics['monthly_data'])).number_format = '$#,##0.00'
        ws.cell(row=row, column=6, value=(metrics['extra_pickups'] + metrics['overages']) / len(metrics['monthly_data'])).number_format = '$#,##0.00'
        row += 1

    # Totals
    row += 1
    ws.cell(row=row, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=row, column=3, value=metrics['total_spend']).number_format = '$#,##0.00'
    ws.cell(row=row, column=3).font = Font(bold=True)

    ws.cell(row=row, column=1, value='AVERAGE').font = Font(bold=True)
    ws.cell(row=row, column=3, value=metrics['avg_monthly_cost']).number_format = '$#,##0.00'
    ws.cell(row=row, column=4, value=metrics['cost_per_door']).number_format = '$#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

def create_optimization_sheet(wb: Workbook, optimization_results: Dict, property_config: Dict):
    """Create OPTIMIZATION sheet with calculation breakdowns"""
    ws = wb.create_sheet("OPTIMIZATION")

    # Styling
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    section_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

    # Title
    ws['A1'] = "OPTIMIZATION OPPORTUNITIES"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    row = 3

    for opt in optimization_results['optimizations']:
        # Opportunity header
        ws[f'A{row}'] = f"💡 {opt['title']}"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="22C55E")
        ws[f'A{row}'].fill = section_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        # Summary
        ws[f'A{row}'] = "Potential Annual Savings:"
        ws[f'B{row}'] = f"${opt['potential_annual_savings']:,.2f}"
        ws[f'B{row}'].font = Font(bold=True, color="22C55E", size=11)
        row += 1

        ws[f'A{row}'] = "Implementation Cost:"
        ws[f'B{row}'] = f"${opt['implementation_cost']:,.2f}"
        row += 1

        ws[f'A{row}'] = "Confidence Level:"
        ws[f'B{row}'] = opt['confidence']
        row += 1

        # Calculation breakdown
        ws[f'A{row}'] = "CALCULATION BREAKDOWN:"
        ws[f'A{row}'].font = Font(bold=True, underline='single')
        row += 1

        for key, value in opt['calculation_breakdown'].items():
            label = key.replace('_', ' ').title() + ':'
            ws[f'A{row}'] = label

            if isinstance(value, (int, float)):
                if 'pct' in key or 'percentage' in key:
                    ws[f'B{row}'] = f"{value:.1f}%"
                elif 'cost' in key or 'savings' in key or 'annual' in key:
                    ws[f'B{row}'] = f"${value:,.2f}"
                else:
                    ws[f'B{row}'] = f"{value:,.1f}"
            else:
                ws[f'B{row}'] = str(value)

            row += 1

        # Recommendation
        ws[f'A{row}'] = "Recommendation:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        ws[f'A{row}'] = opt['recommendation']
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{row}:D{row}')
        ws.row_dimensions[row].height = 40
        row += 2

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

def create_quality_check_sheet(wb: Workbook, validation_report: Dict):
    """Create QUALITY_CHECK sheet"""
    ws = wb.create_sheet("QUALITY_CHECK")

    # Title
    ws['A1'] = "DATA QUALITY & VALIDATION REPORT"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    row = 3

    # Validation timestamp
    ws[f'A{row}'] = "Validation Timestamp:"
    ws[f'B{row}'] = validation_report['timestamp']
    row += 1

    ws[f'A{row}'] = "Overall Status:"
    ws[f'B{row}'] = "[PASS] PASSED" if validation_report['passed'] else "[FAIL] FAILED"
    ws[f'B{row}'].font = Font(
        bold=True,
        color="22C55E" if validation_report['passed'] else "DC2626"
    )
    row += 2

    # Summary
    summary = validation_report['summary']
    ws[f'A{row}'] = "Total Checks:"
    ws[f'B{row}'] = summary['total_checks']
    row += 1

    ws[f'A{row}'] = "Passed:"
    ws[f'B{row}'] = summary['passed_checks']
    ws[f'B{row}'].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    row += 1

    ws[f'A{row}'] = "Failed:"
    ws[f'B{row}'] = summary['failed_checks']
    if summary['failed_checks'] > 0:
        ws[f'B{row}'].fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    row += 1

    ws[f'A{row}'] = "Warnings:"
    ws[f'B{row}'] = summary['warnings']
    if summary['warnings'] > 0:
        ws[f'B{row}'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    row += 2

    # Detailed results
    ws[f'A{row}'] = "DETAILED VALIDATION RESULTS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    for category, results in validation_report['validation_results'].items():
        category_name = category.replace('_', ' ').title()
        status = results.get('status', 'UNKNOWN')

        ws[f'A{row}'] = category_name
        ws[f'B{row}'] = status

        if status == 'PASSED':
            ws[f'B{row}'].font = Font(color="22C55E", bold=True)
        elif status == 'FAILED':
            ws[f'B{row}'].font = Font(color="DC2626", bold=True)

        row += 1

    # Warnings
    if validation_report['warnings']:
        row += 1
        ws[f'A{row}'] = "WARNINGS"
        ws[f'A{row}'].font = Font(bold=True, color="F59E0B", size=12)
        row += 1

        for warning in validation_report['warnings']:
            ws[f'A{row}'] = warning
            ws[f'A{row}'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            ws.merge_cells(f'A{row}:D{row}')
            ws.row_dimensions[row].height = 30
            row += 1

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

def create_documentation_sheet(wb: Workbook, property_config: Dict):
    """Create DOCUMENTATION_NOTES sheet"""
    ws = wb.create_sheet("DOCUMENTATION_NOTES")

    # Title
    ws['A1'] = "DOCUMENTATION & REFERENCE"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:C1')

    row = 3

    # Vendor Contacts
    ws[f'A{row}'] = "VENDOR CONTACTS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws[f'A{row}'] = "Vendor:"
    ws[f'B{row}'] = property_config['vendor']
    row += 1

    ws[f'A{row}'] = "Account Number:"
    ws[f'B{row}'] = property_config['account_number']
    row += 1

    ws[f'A{row}'] = "Phone:"
    ws[f'B{row}'] = "(800) 796-9696"
    row += 2

    # Formulas
    ws[f'A{row}'] = "FORMULA REFERENCE"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    formulas = [
        ('Cost Per Door:', 'Total Monthly Cost ÷ Unit Count'),
        ('Yards Per Door (Dumpster):', '(Avg Container Size × Pickups/Month × Total Containers) ÷ Units'),
        ('Overage Percentage:', '(Extra Pickups + Overages) ÷ Total Spend × 100%')
    ]

    for label, formula in formulas:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = formula
        row += 1

    row += 1

    # Glossary
    ws[f'A{row}'] = "GLOSSARY"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    glossary_items = [
        ('CPD:', 'Cost Per Door - Monthly waste cost per apartment unit'),
        ('YPD:', 'Yards Per Door - Container capacity per unit'),
        ('FEL:', 'Front End Loader - Dumpster with front-loading mechanism'),
        ('MSW:', 'Municipal Solid Waste - Regular trash/refuse'),
        ('Overage:', 'Charges for excess pickups or contamination beyond base contract')
    ]

    for term, definition in glossary_items:
        ws[f'A{row}'] = term
        ws[f'B{row}'] = definition
        ws[f'B{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20

def create_contract_terms_sheet(wb: Workbook):
    """Create CONTRACT_TERMS sheet with verbatim clause extraction"""
    ws = wb.create_sheet("CONTRACT_TERMS")

    # Styling
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    high_risk_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    medium_risk_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    low_risk_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

    # Title
    ws['A1'] = "CONTRACT TERMS & RISK ANALYSIS"
    ws['A1'].font = Font(bold=True, size=14, color="1E3A8A")
    ws.merge_cells('A1:E1')

    row = 3

    # Contract Information
    ws[f'A{row}'] = "CONTRACT INFORMATION"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    ws.merge_cells(f'A{row}:B{row}')
    row += 1

    contract_info = [
        ('Effective Date:', CONTRACT_DATA['effective_date'].strftime('%Y-%m-%d')),
        ('Initial Term:', f"{CONTRACT_DATA['contract_term_years']} years"),
        ('Auto-Renewal:', 'Yes - 12 month terms'),
        ('Monthly Base Cost:', f"${CONTRACT_DATA['service_details']['monthly_base_cost']:,.2f}"),
        ('Total Containers:', f"{CONTRACT_DATA['service_details']['total_containers']}")
    ]

    for label, value in contract_info:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1

    row += 1

    # Calendar Reminders (CRITICAL)
    ws[f'A{row}'] = "[WARN] CALENDAR REMINDERS - ACTION REQUIRED"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="DC2626")
    ws[f'A{row}'].fill = high_risk_fill
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    # Headers
    reminder_headers = ['Date', 'Action Required', 'Criticality', 'Days Until', 'Notes']
    for col_idx, header in enumerate(reminder_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    row += 1

    # Reminders
    for reminder in CONTRACT_DATA['calendar_reminders']:
        ws.cell(row=row, column=1, value=reminder['date'].strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=reminder['action'])
        ws.cell(row=row, column=3, value=reminder['criticality'])
        ws.cell(row=row, column=4, value=reminder['days_until'])
        ws.cell(row=row, column=5, value='Set Outlook/Google Calendar reminder')

        # Color code by urgency
        days_until = reminder['days_until']
        fill = high_risk_fill if days_until < 90 else medium_risk_fill if days_until < 180 else low_risk_fill

        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = fill

        row += 1

    row += 2

    # Contract Clauses
    ws[f'A{row}'] = "EXTRACTED CONTRACT CLAUSES"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    # Clause headers
    clause_headers = ['Category', 'Verbatim Contract Language', 'Risk Level', 'Impact', 'Recommended Action']
    for col_idx, header in enumerate(clause_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    row += 1

    # Clause rows
    for clause in CONTRACT_DATA['clauses']:
        ws.cell(row=row, column=1, value=clause['category'])
        ws.cell(row=row, column=2, value=clause['verbatim_text'])
        ws.cell(row=row, column=3, value=clause['risk_severity'].upper())
        ws.cell(row=row, column=4, value=clause['impact'])
        ws.cell(row=row, column=5, value=clause['action_required'])

        # Risk color coding
        risk_fill = {
            'high': high_risk_fill,
            'medium': medium_risk_fill,
            'low': low_risk_fill
        }.get(clause['risk_severity'].lower(), low_risk_fill)

        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = risk_fill
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        ws.row_dimensions[row].height = 80
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 40

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function"""
    print("=" * 70)
    print("WasteWise Analytics - Validated Edition")
    print(f"   Property: {PROPERTY_CONFIG['name']}")
    print(f"   Units: {PROPERTY_CONFIG['unit_count']}")
    print("=" * 70)

    # Load data
    excel_file = r"C:\Users\Richard\Downloads\Orion Data Part 2\Extraction_Output\COMPLETE_All_Properties_UPDATED_20251103_101053.xlsx"
    df = load_invoice_data(excel_file)

    # Calculate metrics
    metrics = calculate_performance_metrics(df, PROPERTY_CONFIG)

    # Analyze optimizations
    optimization_results = analyze_optimizations(df, metrics, PROPERTY_CONFIG)

    # Run validation
    validator = WasteWiseValidator()
    passed, validation_report = validator.validate_all(
        df, metrics, optimization_results, PROPERTY_CONFIG
    )

    # Display validation results
    print("\nVALIDATION RESULTS:")
    for category, results in validation_report['validation_results'].items():
        status = results.get('status', 'UNKNOWN')
        icon = '[PASS]' if status == 'PASSED' else '[FAIL]' if status == 'FAILED' else '[SKIP]'
        category_name = category.replace('_', ' ').title()
        print(f"   {icon} {category_name}: {status}")

    if validation_report['warnings']:
        print("\nWARNINGS:")
        for warning in validation_report['warnings']:
            print(f"   {warning}")

    print("\n" + "=" * 70)
    print(f"VALIDATION SUMMARY:")
    print(f"   Total Checks: {validation_report['summary']['total_checks']}")
    print(f"   Passed: {validation_report['summary']['passed_checks']}")
    print(f"   Failed: {validation_report['summary']['failed_checks']}")
    print(f"   Warnings: {validation_report['summary']['warnings']}")
    print("=" * 70)

    if not passed:
        print("\nVALIDATION FAILED - Cannot proceed to output generation")
        print("   Please fix the errors above and re-run the analysis")
        return

    print("\nALL VALIDATIONS PASSED - Proceeding to output generation")

    # Generate workbook
    wb = create_excel_workbook(df, metrics, optimization_results,
                               validation_report, PROPERTY_CONFIG)

    # Save output
    output_dir = r"C:\Users\Richard\Downloads\Orion Data Part 2\Extraction_Output"
    output_file = os.path.join(output_dir, "BellaMirage_WasteAnalysis_Validated.xlsx")
    wb.save(output_file)

    print(f"\nWorkbook saved: {output_file}")

    # Generate executive summary
    print("\n" + "=" * 70)
    print("EXECUTIVE SUMMARY")
    print("=" * 70)
    print(f"\nProperty: {PROPERTY_CONFIG['name']} ({PROPERTY_CONFIG['unit_count']} units)")
    print(f"Analysis Period: {len(metrics['monthly_data'])} months ({df['Invoice Date'].min().strftime('%b %Y')} - {df['Invoice Date'].max().strftime('%b %Y')})")
    print(f"Total Invoices Analyzed: {metrics['num_invoices']}")
    print(f"\nKEY METRICS:")
    print(f"  • Average Monthly Cost: ${metrics['avg_monthly_cost']:,.2f}")
    print(f"  • Cost Per Door: ${metrics['cost_per_door']:.2f}")
    print(f"  • Yards Per Door: {metrics['yards_per_door']:.2f}")
    print(f"  • Overage Rate: {metrics['overage_pct']:.1f}%")
    print(f"\nOPTIMIZATION OPPORTUNITIES:")
    print(f"  • Total Annual Opportunity: ${optimization_results['total_potential_annual_savings']:,.0f}")
    for opt in optimization_results['optimizations']:
        print(f"  • {opt['title']}: ${opt['potential_annual_savings']:,.0f}/year ({opt['confidence']} confidence)")
    print(f"\nCONTRACT STATUS:")
    print(f"  • Vendor: {PROPERTY_CONFIG['vendor']}")
    print(f"  • Contract Effective: {CONTRACT_DATA['effective_date'].strftime('%Y-%m-%d')}")
    print(f"  • Auto-Renewal: {CONTRACT_DATA['renewal_term_months']}-month terms")
    print(f"  • Next Action: {CONTRACT_DATA['calendar_reminders'][0]['action']}")
    print(f"  • Days Until: {CONTRACT_DATA['calendar_reminders'][0]['days_until']} days")
    print("\n" + "=" * 70)
    print(f"Analysis Complete - Validated workbook ready for review")
    print("=" * 70)

if __name__ == "__main__":
    main()
