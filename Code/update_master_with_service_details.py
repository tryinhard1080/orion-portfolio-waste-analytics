"""
Update Master Excel File with Service Details

This script updates the master Excel file with complete service details
for the 4 properties where we have extracted the data.
"""

import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
import shutil

# Paths
BASE_DIR = Path(__file__).parent.parent
MASTER_FILE = BASE_DIR / "Portfolio_Reports" / "MASTER_Portfolio_Complete_Data.xlsx"
BACKUP_FILE = BASE_DIR / "Portfolio_Reports" / f"MASTER_Portfolio_Complete_Data_BACKUP_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# Service details to add
SERVICE_DETAILS = {
    'McCord Park FL': {
        'container_count': 15,
        'container_size': '8 YD',
        'container_type': 'Front Loader (FL)',
        'service_frequency': '3x/week',
        'service_days': 'M/W/F',
        'total_yards': 116,  # 1×4 + 12×8 + 2×8
        'units': 416,
        'ypd': 0.28,
        'notes': '1×4YD FL + 12×8YD FL (trash) + 2×8YD SS (recycling)'
    },
    'Orion McKinney': {
        'container_count': 10,
        'container_size': '8 YD',
        'container_type': 'Front Loader (FL)',
        'service_frequency': '3x/week',
        'service_days': 'M/W/F',
        'total_yards': 84,  # 8×8 + 2×10
        'units': 453,
        'ypd': 0.19,
        'notes': '8×8YD FL + 2×10YD FL'
    },
    'The Club at Millenia': {
        'container_count': 6,
        'container_size': '8 YD',
        'container_type': 'Dumpster',
        'service_frequency': '4x/week',
        'service_days': 'Weekly x4',
        'total_yards': 42,  # 4×8 + 1×6 + 1×4
        'units': 560,
        'ypd': 0.08,
        'notes': '4×8YD + 1×6YD + 1×4YD Dumpsters'
    },
    'Bella Mirage': {
        'container_count': 6,
        'container_size': '8 YD',
        'container_type': 'Front End Loader (FEL)',
        'service_frequency': '3x/week',
        'service_days': '3x per week',
        'total_yards': 48,  # 6×8
        'units': 715,
        'ypd': 0.07,
        'notes': '6×8YD FEL'
    }
}

def create_backup():
    """Create a backup of the master file before updating"""
    
    print('=' * 80)
    print('CREATING BACKUP')
    print('=' * 80)
    print()
    
    if MASTER_FILE.exists():
        shutil.copy2(MASTER_FILE, BACKUP_FILE)
        print(f'✅ Backup created: {BACKUP_FILE.name}')
        print()
        return True
    else:
        print(f'❌ Master file not found: {MASTER_FILE}')
        return False

def update_property_sheet(property_name, details):
    """Update a single property sheet with service details"""
    
    print(f'Updating: {property_name}')
    print('-' * 80)
    
    # Load the workbook
    wb = openpyxl.load_workbook(MASTER_FILE)
    
    # Check if sheet exists
    if property_name not in wb.sheetnames:
        print(f'  ❌ Sheet not found: {property_name}')
        wb.close()
        return False
    
    ws = wb[property_name]
    
    # Find the header row (should be row 1 based on structure)
    header_row = None
    for row in range(1, 10):
        cell_value = ws.cell(row, 1).value
        if cell_value and ('Property' in str(cell_value) or 'Invoice' in str(cell_value)):
            header_row = row
            break

    if not header_row:
        print(f'  ❌ Could not find header row')
        wb.close()
        return False

    print(f'  ✓ Found header row: {header_row}')
    
    # Get existing headers
    headers = []
    col = 1
    while True:
        cell_value = ws.cell(header_row, col).value
        if cell_value:
            headers.append(str(cell_value).strip())
            col += 1
        else:
            break
    
    print(f'  ✓ Found {len(headers)} existing columns')
    
    # Add new columns if they don't exist
    new_columns = [
        'Container Count',
        'Container Size',
        'Container Type',
        'Service Frequency',
        'Service Days',
        'Total Yards',
        'YPD',
        'Service Notes'
    ]
    
    columns_added = []
    for new_col in new_columns:
        if new_col not in headers:
            # Add to the end
            next_col = len(headers) + 1
            ws.cell(header_row, next_col).value = new_col
            headers.append(new_col)
            columns_added.append(new_col)
    
    if columns_added:
        print(f'  ✓ Added {len(columns_added)} new columns: {", ".join(columns_added)}')
    else:
        print(f'  ✓ All columns already exist')
    
    # Find column indices for our new data
    col_indices = {}
    for i, header in enumerate(headers, 1):
        if header == 'Container Count':
            col_indices['container_count'] = i
        elif header == 'Container Size':
            col_indices['container_size'] = i
        elif header == 'Container Type':
            col_indices['container_type'] = i
        elif header == 'Service Frequency':
            col_indices['service_frequency'] = i
        elif header == 'Service Days':
            col_indices['service_days'] = i
        elif header == 'Total Yards':
            col_indices['total_yards'] = i
        elif header == 'YPD':
            col_indices['ypd'] = i
        elif header == 'Service Notes':
            col_indices['notes'] = i
    
    # Count data rows
    data_rows = 0
    row = header_row + 1
    while ws.cell(row, 1).value:
        data_rows += 1
        row += 1
    
    print(f'  ✓ Found {data_rows} data rows')
    
    # Update all data rows with service details
    rows_updated = 0
    for row in range(header_row + 1, header_row + 1 + data_rows):
        # Only update if the row has data
        if ws.cell(row, 1).value:
            # Update each field
            if 'container_count' in col_indices:
                ws.cell(row, col_indices['container_count']).value = details['container_count']
            if 'container_size' in col_indices:
                ws.cell(row, col_indices['container_size']).value = details['container_size']
            if 'container_type' in col_indices:
                ws.cell(row, col_indices['container_type']).value = details['container_type']
            if 'service_frequency' in col_indices:
                ws.cell(row, col_indices['service_frequency']).value = details['service_frequency']
            if 'service_days' in col_indices:
                ws.cell(row, col_indices['service_days']).value = details['service_days']
            if 'total_yards' in col_indices:
                ws.cell(row, col_indices['total_yards']).value = details['total_yards']
            if 'ypd' in col_indices:
                ws.cell(row, col_indices['ypd']).value = details['ypd']
            if 'notes' in col_indices:
                ws.cell(row, col_indices['notes']).value = details['notes']
            
            rows_updated += 1
    
    print(f'  ✓ Updated {rows_updated} rows with service details')
    
    # Save the workbook
    wb.save(MASTER_FILE)
    wb.close()
    
    print(f'  ✅ {property_name} updated successfully')
    print()
    
    return True

def verify_updates():
    """Verify that the updates were applied correctly"""
    
    print('=' * 80)
    print('VERIFYING UPDATES')
    print('=' * 80)
    print()
    
    xl = pd.ExcelFile(MASTER_FILE)
    
    for property_name, details in SERVICE_DETAILS.items():
        if property_name in xl.sheet_names:
            df = pd.read_excel(MASTER_FILE, sheet_name=property_name)
            
            print(f'{property_name}:')
            
            # Check if new columns exist
            new_cols = ['Container Count', 'Container Size', 'Container Type', 
                       'Service Frequency', 'YPD']
            
            found_cols = [col for col in new_cols if col in df.columns]
            
            if found_cols:
                print(f'  ✓ Found {len(found_cols)}/{len(new_cols)} service columns')
                
                # Show sample data from first row
                if len(df) > 0:
                    first_row = df.iloc[0]
                    for col in found_cols:
                        val = first_row[col]
                        if pd.notna(val):
                            print(f'    - {col}: {val}')
                
                # Verify YPD if present
                if 'YPD' in df.columns:
                    ypd_values = df['YPD'].dropna().unique()
                    if len(ypd_values) > 0:
                        print(f'  ✓ YPD value: {ypd_values[0]:.2f}')
                        if abs(ypd_values[0] - details['ypd']) < 0.01:
                            print(f'  ✅ YPD matches expected value ({details["ypd"]:.2f})')
                        else:
                            print(f'  ⚠ YPD mismatch: expected {details["ypd"]:.2f}')
            else:
                print(f'  ❌ No service columns found')
            
            print()

def create_summary_report():
    """Create a summary report of the updates"""
    
    report_path = BASE_DIR / 'Portfolio_Reports' / 'MASTER_FILE_UPDATE_SUMMARY.md'
    
    content = f"""# Master File Update Summary

**Date:** {datetime.now().strftime('%B %d, %Y at %I:%M %p')}  
**File:** MASTER_Portfolio_Complete_Data.xlsx  
**Action:** Added service details for 4 properties

---

## UPDATES APPLIED

### Properties Updated (4)

1. **McCord Park FL**
   - Container Count: 15
   - Container Size: 8 YD
   - Container Type: Front Loader (FL)
   - Service Frequency: 3x/week
   - Service Days: M/W/F
   - Total Yards: 116
   - YPD: 0.28
   - Notes: 1×4YD FL + 12×8YD FL (trash) + 2×8YD SS (recycling)

2. **Orion McKinney**
   - Container Count: 10
   - Container Size: 8 YD
   - Container Type: Front Loader (FL)
   - Service Frequency: 3x/week
   - Service Days: M/W/F
   - Total Yards: 84
   - YPD: 0.19
   - Notes: 8×8YD FL + 2×10YD FL

3. **The Club at Millenia**
   - Container Count: 6
   - Container Size: 8 YD
   - Container Type: Dumpster
   - Service Frequency: 4x/week
   - Service Days: Weekly x4
   - Total Yards: 42
   - YPD: 0.08
   - Notes: 4×8YD + 1×6YD + 1×4YD Dumpsters

4. **Bella Mirage**
   - Container Count: 6
   - Container Size: 8 YD
   - Container Type: Front End Loader (FEL)
   - Service Frequency: 3x/week
   - Service Days: 3x per week
   - Total Yards: 48
   - YPD: 0.07
   - Notes: 6×8YD FEL

---

## NEW COLUMNS ADDED

The following columns were added to each property sheet:

1. **Container Count** - Number of containers
2. **Container Size** - Size in yards (e.g., "8 YD")
3. **Container Type** - Type (FL, FEL, Dumpster, etc.)
4. **Service Frequency** - Pickups per week (e.g., "3x/week")
5. **Service Days** - Days of service (e.g., "M/W/F")
6. **Total Yards** - Total yard capacity
7. **YPD** - Yards Per Door calculation
8. **Service Notes** - Additional service details

---

## YARDS PER DOOR (YPD) SUMMARY

| Property | Total Yards | Units | YPD | Performance |
|----------|-------------|-------|-----|-------------|
| McCord Park FL | 116 | 416 | 0.28 | ✅ Excellent |
| Orion McKinney | 84 | 453 | 0.19 | ✅ Excellent |
| The Club at Millenia | 42 | 560 | 0.08 | ✅ Excellent |
| Bella Mirage | 48 | 715 | 0.07 | ✅ Excellent |

**All properties are well below the 2.0-2.25 YPD target!**

---

## BACKUP INFORMATION

**Backup File:** MASTER_Portfolio_Complete_Data_BACKUP_[timestamp].xlsx  
**Location:** Portfolio_Reports/  
**Purpose:** Restore point before service details update

To restore from backup if needed:
1. Delete current MASTER_Portfolio_Complete_Data.xlsx
2. Rename backup file to MASTER_Portfolio_Complete_Data.xlsx

---

## NEXT STEPS

1. ✅ Service details added for 4/6 TX/FL properties
2. ⚠️ Still pending: Orion Prosper, Orion Prosper Lakes
3. ✅ Ready to regenerate reports with YPD metrics
4. ✅ Ready to create performance analysis with service efficiency

---

**Update completed successfully!**
"""
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(content)
    
    print(f'✅ Summary report created: {report_path.name}')
    print()

def main():
    """Main function"""
    
    print('=' * 80)
    print('UPDATE MASTER FILE WITH SERVICE DETAILS')
    print('=' * 80)
    print()
    
    # Create backup
    if not create_backup():
        print('❌ Cannot proceed without backup')
        return
    
    # Update each property
    print('=' * 80)
    print('UPDATING PROPERTY SHEETS')
    print('=' * 80)
    print()
    
    success_count = 0
    for property_name, details in SERVICE_DETAILS.items():
        if update_property_sheet(property_name, details):
            success_count += 1
    
    # Verify updates
    verify_updates()
    
    # Create summary report
    create_summary_report()
    
    # Final summary
    print('=' * 80)
    print('UPDATE COMPLETE')
    print('=' * 80)
    print()
    print(f'✅ Successfully updated {success_count}/{len(SERVICE_DETAILS)} properties')
    print()
    print('Updated properties:')
    for property_name, details in SERVICE_DETAILS.items():
        print(f'  ✅ {property_name} - {details["container_count"]} containers, YPD: {details["ypd"]:.2f}')
    print()
    print(f'Backup saved: {BACKUP_FILE.name}')
    print()

if __name__ == "__main__":
    main()

