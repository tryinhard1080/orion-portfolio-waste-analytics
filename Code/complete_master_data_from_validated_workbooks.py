"""
Complete Master Data File from Validated Workbooks

Extracts service details, contract information, and vendor data from all 10 property
validated workbooks and updates the master Portfolio data file.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
import re

# Property mappings
PROPERTIES = {
    'Orion Prosper': 'Properties/Orion_Prosper/Orion_Prosper_WasteAnalysis_Validated.xlsx',
    'McCord Park FL': 'Properties/McCord_Park_FL/McCord_Park_FL_WasteAnalysis_Validated.xlsx',
    'Orion McKinney': 'Properties/Orion_McKinney/Orion_McKinney_WasteAnalysis_Validated.xlsx',
    'The Club at Millenia': 'Properties/The_Club_at_Millenia/The_Club_at_Millenia_WasteAnalysis_Validated.xlsx',
    'Bella Mirage': 'Properties/Bella_Mirage/Bella_Mirage_WasteAnalysis_Validated.xlsx',
    'Orion Prosper Lakes': 'Properties/Orion_Prosper_Lakes/Orion_Prosper_Lakes_WasteAnalysis_Validated.xlsx',
    'Mandarina': 'Properties/Mandarina/Mandarina_Complete_WasteAnalysis_Validated.xlsx',
    'Pavilions at Arrowhead': 'Properties/Pavilions_at_Arrowhead/Pavilions_at_Arrowhead_WasteAnalysis_Validated.xlsx',
    'Springs at Alta Mesa': 'Properties/Springs_at_Alta_Mesa/Springs_at_Alta_Mesa_WasteAnalysis_Validated.xlsx',
    'Tempe Vista': 'Properties/Tempe_Vista/Tempe_Vista_WasteAnalysis_Validated.xlsx'
}

def extract_workbook_data(property_name, workbook_path):
    """
    Extract service details and contract information from a validated workbook
    """
    print(f"\n{'='*70}")
    print(f"Extracting: {property_name}")
    print(f"{'='*70}")

    if not Path(workbook_path).exists():
        print(f"  [WARNING] File not found: {workbook_path}")
        return None

    data = {
        'property_name': property_name,
        'vendor': None,
        'service_type': None,
        'containers': [],
        'contract': {}
    }

    try:
        # Read SUMMARY_FULL sheet
        df_summary = pd.read_excel(workbook_path, sheet_name='SUMMARY_FULL', header=None)

        # Extract vendor (usually in first few rows)
        for idx, row in df_summary.iterrows():
            row_str = ' '.join([str(cell) for cell in row if pd.notna(cell)])

            # Look for vendor names
            if 'republic' in row_str.lower():
                data['vendor'] = 'Republic Services'
                break
            elif 'waste management' in row_str.lower() or 'wm' in row_str.lower():
                data['vendor'] = 'Waste Management'
                break
            elif 'waste connections' in row_str.lower():
                data['vendor'] = 'Waste Connections'
                break
            elif 'community waste' in row_str.lower():
                data['vendor'] = 'Community Waste Disposal'
                break
            elif 'frontier' in row_str.lower():
                data['vendor'] = 'Frontier Waste Solutions'
                break
            elif 'ally waste' in row_str.lower():
                data['vendor'] = 'Ally Waste (Waste Consolidators Inc)'
                break
            elif 'city of' in row_str.lower():
                data['vendor'] = f"City of {property_name.split()[-1]}"
                break

        # Determine service type from summary
        service_keywords = {
            'compactor': ['compactor', 'tons', 'tonnage', '30 yard', '40 yard', '34 yard'],
            'dumpster': ['dumpster', 'front load', 'fel', '2 yard', '4 yard', '6 yard', '8 yard'],
            'mixed': ['mixed', 'dumpster + bulk', 'city service']
        }

        summary_text = df_summary.to_string().lower()
        for service_type, keywords in service_keywords.items():
            if any(keyword in summary_text for keyword in keywords):
                data['service_type'] = service_type.capitalize()
                if service_type == 'mixed':
                    data['service_type'] = 'Mixed'
                break

        print(f"  Vendor: {data['vendor']}")
        print(f"  Service Type: {data['service_type']}")

    except Exception as e:
        print(f"  [ERROR] Could not read SUMMARY_FULL: {e}")

    # Try to extract contract information from CONTRACT_TERMS sheet
    try:
        if 'CONTRACT_TERMS' in pd.ExcelFile(workbook_path).sheet_names:
            df_contract = pd.read_excel(workbook_path, sheet_name='CONTRACT_TERMS', header=None)

            contract_text = df_contract.to_string()

            # Look for contract dates
            date_patterns = [
                r'effective[:\s]+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
                r'start[:\s]+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
                r'expiration[:\s]+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
                r'(\d{1,2}[/-]\d{1,2}[/-]\d{4})'
            ]

            dates_found = []
            for pattern in date_patterns:
                matches = re.findall(pattern, contract_text, re.IGNORECASE)
                dates_found.extend(matches)

            if dates_found:
                data['contract']['dates_mentioned'] = dates_found[:3]  # First 3 dates found

            # Look for term length
            term_match = re.search(r'(\d+)\s*year', contract_text, re.IGNORECASE)
            if term_match:
                data['contract']['term_years'] = int(term_match.group(1))

            # Look for notice period
            notice_match = re.search(r'(\d+)\s*day[s]?\s*notice', contract_text, re.IGNORECASE)
            if notice_match:
                data['contract']['notice_days'] = int(notice_match.group(1))

            print(f"  Contract Info Extracted: {data['contract']}")

    except Exception as e:
        print(f"  [INFO] No CONTRACT_TERMS sheet or extraction failed: {e}")

    # Extract service details from EXPENSE_ANALYSIS or HAUL_LOG
    try:
        sheet_to_read = None
        excel_file = pd.ExcelFile(workbook_path)

        if 'EXPENSE_ANALYSIS' in excel_file.sheet_names:
            sheet_to_read = 'EXPENSE_ANALYSIS'
        elif 'HAUL_LOG' in excel_file.sheet_names:
            sheet_to_read = 'HAUL_LOG'

        if sheet_to_read:
            df_expense = pd.read_excel(workbook_path, sheet_name=sheet_to_read)

            # Look for container information
            if 'Container' in df_expense.columns or 'Container Size' in df_expense.columns:
                container_col = 'Container Size' if 'Container Size' in df_expense.columns else 'Container'
                containers = df_expense[container_col].dropna().unique()
                data['containers'] = [str(c) for c in containers if str(c) != 'nan']
                print(f"  Containers Found: {data['containers']}")

            # Look for frequency
            if 'Frequency' in df_expense.columns:
                frequencies = df_expense['Frequency'].dropna().unique()
                data['frequency'] = [str(f) for f in frequencies if str(f) != 'nan']
                print(f"  Frequencies Found: {data['frequency']}")

    except Exception as e:
        print(f"  [INFO] Could not extract detailed service info: {e}")

    return data


def update_master_file(master_path, extracted_data):
    """
    Update the master Excel file with extracted data
    """
    print(f"\n{'='*70}")
    print("UPDATING MASTER FILE")
    print(f"{'='*70}")

    wb = load_workbook(master_path)

    # Update Property Overview
    ws_overview = wb['Property Overview']

    for property_name, data in extracted_data.items():
        if not data:
            continue

        # Find the property row
        for row_idx in range(2, ws_overview.max_row + 1):
            cell_value = ws_overview.cell(row_idx, 1).value
            if cell_value and property_name in str(cell_value):
                # Update service type if found
                if data.get('service_type'):
                    ws_overview.cell(row_idx, 4).value = data['service_type']  # Column D
                    print(f"  Updated {property_name} Service Type: {data['service_type']}")
                break

    # Update Contract Terms
    ws_contract = wb['Contract Terms']

    # Define contract information we extracted from PDFs
    contract_updates = {
        'The Club at Millenia': {
            'vendor': 'Waste Connections of Florida',
            'start_date': '05/25/2021',
            'term': '3 years',
            'end_date': '05/25/2024',
            'auto_renew': 'Yes (per T&C)',
            'notice_period': '90 days (review needed)',
            'status': 'Expired - Needs Review'
        },
        'Bella Mirage': {
            'vendor': 'Waste Management',
            'start_date': '04/08/2020',
            'term': '3 years',
            'end_date': '04/08/2023',
            'auto_renew': 'Yes - 12 month terms',
            'notice_period': '90 days (min), 180 days (max)',
            'status': 'Auto-Renewed - Currently Active'
        },
        'Orion Prosper Lakes': {
            'vendor': 'Republic Services',
            'start_date': '01/01/2025',
            'term': 'TBD - Review Contract',
            'end_date': 'TBD',
            'auto_renew': 'TBD - Review Contract',
            'notice_period': 'TBD - Review Contract',
            'status': 'Active - Recent Start'
        },
        'Orion Prosper': {
            'vendor': 'Republic Services',
            'start_date': 'Unknown',
            'term': 'Unknown',
            'end_date': 'Unknown',
            'auto_renew': 'Unknown',
            'notice_period': 'Unknown',
            'status': 'Active - Contract Needed'
        },
        'McCord Park FL': {
            'vendor': 'Community Waste Disposal, LP',
            'start_date': 'Unknown',
            'term': 'Unknown',
            'end_date': 'Unknown',
            'auto_renew': 'Unknown',
            'notice_period': 'Unknown',
            'status': 'Active - Contract Needed'
        },
        'Orion McKinney': {
            'vendor': 'Frontier Waste Solutions',
            'start_date': 'TBD - Review Contract',
            'term': 'TBD',
            'end_date': 'TBD',
            'auto_renew': 'TBD',
            'notice_period': 'TBD - Review Contract',
            'status': 'Active - Contract Review Needed'
        },
        'Mandarina': {
            'vendor': 'Waste Management + Ally Waste',
            'start_date': 'Unknown',
            'term': 'Unknown',
            'end_date': 'Unknown',
            'auto_renew': 'Unknown',
            'notice_period': 'Unknown',
            'status': 'Active - Contract Needed'
        },
        'Pavilions at Arrowhead': {
            'vendor': 'City of Glendale + Waste Consolidators Inc (Bulk)',
            'start_date': 'TBD - Review Bulk Contract',
            'term': 'TBD',
            'end_date': 'TBD',
            'auto_renew': 'TBD',
            'notice_period': 'TBD - Review Contract',
            'status': 'Active - Bulk Contract Review Needed'
        },
        'Springs at Alta Mesa': {
            'vendor': 'City of Mesa + Waste Consolidators Inc (Bulk)',
            'start_date': 'TBD - Review Bulk Contract',
            'term': 'TBD',
            'end_date': 'TBD',
            'auto_renew': 'TBD',
            'notice_period': 'TBD - Review Contract',
            'status': 'Active - Bulk Contract Review Needed'
        },
        'Tempe Vista': {
            'vendor': 'Waste Management + Waste Consolidators Inc (Bulk)',
            'start_date': 'TBD - Review Contracts',
            'term': 'TBD',
            'end_date': 'TBD',
            'auto_renew': 'TBD',
            'notice_period': 'TBD - Review Contract',
            'status': 'Active - Contract Review Needed'
        }
    }

    # Update contract rows
    for row_idx in range(2, ws_contract.max_row + 1):
        property_cell = ws_contract.cell(row_idx, 1).value
        if not property_cell:
            continue

        property_name = str(property_cell).strip()

        # Clean up property name for matching
        if 'Orion Prosper Lakes' in property_name or 'Little Elm' in property_name:
            property_key = 'Orion Prosper Lakes'
        else:
            property_key = property_name

        if property_key in contract_updates:
            update = contract_updates[property_key]

            # Column B: Vendor
            ws_contract.cell(row_idx, 2).value = update['vendor']
            # Column C: Contract Start
            ws_contract.cell(row_idx, 3).value = update['start_date']
            # Column D: Contract Term
            ws_contract.cell(row_idx, 4).value = update['term']
            # Column E: Contract End
            ws_contract.cell(row_idx, 5).value = update['end_date']
            # Column F: Auto Renewal
            ws_contract.cell(row_idx, 6).value = update['auto_renew']
            # Column G: Notice Period
            ws_contract.cell(row_idx, 7).value = update['notice_period']
            # Column J: Status
            ws_contract.cell(row_idx, 10).value = update['status']

            print(f"  Updated contract for: {property_name}")

    # Save updated workbook
    output_path = master_path.replace('.xlsx', '_UPDATED.xlsx')
    wb.save(output_path)
    print(f"\n✅ Master file updated: {output_path}")

    return output_path


def main():
    """
    Main execution
    """
    print("="*70)
    print("MASTER DATA COMPLETION FROM VALIDATED WORKBOOKS")
    print("="*70)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    master_path = 'Portfolio_Reports/MASTER_Portfolio_Complete_Data.xlsx'

    # Extract data from all properties
    extracted_data = {}

    for property_name, workbook_path in PROPERTIES.items():
        data = extract_workbook_data(property_name, workbook_path)
        if data:
            extracted_data[property_name] = data

    # Update master file
    updated_path = update_master_file(master_path, extracted_data)

    # Generate summary report
    print(f"\n{'='*70}")
    print("EXTRACTION SUMMARY")
    print(f"{'='*70}")

    for property_name, data in extracted_data.items():
        print(f"\n{property_name}:")
        print(f"  Vendor: {data.get('vendor', 'Not Found')}")
        print(f"  Service Type: {data.get('service_type', 'Not Found')}")
        if data.get('containers'):
            print(f"  Containers: {', '.join(data['containers'])}")
        if data.get('contract'):
            print(f"  Contract Info: {data['contract']}")

    print(f"\n{'='*70}")
    print(f"✅ COMPLETION SUCCESSFUL")
    print(f"{'='*70}")
    print(f"Finished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"\nUpdated file: {updated_path}")
    print(f"Original file: {master_path} (preserved)")


if __name__ == '__main__':
    main()
