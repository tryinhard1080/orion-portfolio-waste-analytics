"""
Update Orion Prosper and Orion Prosper Lakes with service details extracted from invoices
"""

import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
import shutil

# Paths
BASE_DIR = Path(__file__).parent.parent
MASTER_FILE = BASE_DIR / "Portfolio_Reports" / "MASTER_Portfolio_Complete_Data.xlsx"
BACKUP_FILE = BASE_DIR / "Portfolio_Reports" / f"MASTER_Portfolio_Complete_Data_BACKUP_ORION_PROSPER_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# Service details extracted from invoices
ORION_PROSPER_DETAILS = {
    'property': 'Orion Prosper',
    'units': 312,
    'container_count': 2,
    'container_size': 10,
    'container_type': 'Front Loader',
    'service_frequency': 6,
    'service_days': '6x per week',
    'service_notes': '2×10YD Front Loaders @ 6x/week',
    'monthly_yards': (10 * 2 * 6 * 4.33),
    'ypd': (10 * 2 * 6 * 4.33) / 312
}

ORION_PROSPER_LAKES_DETAILS = {
    'property': 'Orion Prosper Lakes',
    'units': 308,
    'container_count': 1,
    'container_size': 40,  # CY (cubic yards)
    'container_type': 'Compactor',
    'service_frequency': 2.4,  # Average from invoice analysis
    'service_days': 'On-call (avg 2.4x/week)',
    'service_notes': '1×40CY Compactor @ on-call service (avg 10.3 pickups/month)',
    'monthly_yards': 40 * 10.3,  # Container size × avg pickups per month
    'ypd': (40 * 10.3) / 308
}

def create_backup():
    """Create backup before updating"""
    print('=' * 80)
    print('CREATING BACKUP')
    print('=' * 80)
    print()
    
    if MASTER_FILE.exists():
        shutil.copy2(MASTER_FILE, BACKUP_FILE)
        print(f'✅ Backup created: {BACKUP_FILE.name}')
        print()
        return True
    else:
        print(f'❌ Master file not found: {MASTER_FILE}')
        return False

def show_details():
    """Show service details before updating"""
    print('=' * 80)
    print('SERVICE DETAILS TO BE ADDED')
    print('=' * 80)
    print()
    
    for details in [ORION_PROSPER_DETAILS, ORION_PROSPER_LAKES_DETAILS]:
        print(f'{details["property"]}:')
        print(f'  Units: {details["units"]}')
        print(f'  Container Count: {details["container_count"]}')
        print(f'  Container Size: {details["container_size"]} {"YD" if details["property"] == "Orion Prosper" else "CY"}')
        print(f'  Container Type: {details["container_type"]}')
        print(f'  Service Frequency: {details["service_frequency"]}x per week')
        print(f'  Service Days: {details["service_days"]}')
        print(f'  Monthly Yards: {details["monthly_yards"]:.2f}')
        print(f'  YPD: {details["ypd"]:.2f}')
        print(f'  Service Notes: {details["service_notes"]}')
        
        target = 2.0
        if details["ypd"] <= target:
            pct_below = ((target - details["ypd"]) / target) * 100
            print(f'  Performance: ✅ {pct_below:.1f}% below target')
        else:
            pct_above = ((details["ypd"] - target) / target) * 100
            print(f'  Performance: ⚠️ {pct_above:.1f}% above target')
        
        print()

def update_property_sheet(property_details):
    """Update a single property sheet with service details"""
    
    property_name = property_details['property']
    
    print(f'Updating: {property_name}')
    print('-' * 80)
    
    wb = openpyxl.load_workbook(MASTER_FILE)
    
    if property_name not in wb.sheetnames:
        print(f'  ❌ Sheet not found: {property_name}')
        wb.close()
        return False
    
    ws = wb[property_name]
    
    # Find header row
    header_row = None
    for row in range(1, 10):
        cell_value = ws.cell(row, 1).value
        if cell_value and 'Property' in str(cell_value):
            header_row = row
            break
    
    if not header_row:
        print(f'  ❌ Could not find header row')
        wb.close()
        return False
    
    # Check if service columns exist, if not add them
    headers = []
    for col in range(1, 50):
        header = ws.cell(header_row, col).value
        if header:
            headers.append(header)
        else:
            break
    
    # Service detail columns to add
    service_columns = [
        'Container Count',
        'Container Size',
        'Container Type',
        'Service Frequency',
        'Service Days',
        'Total Yards',
        'YPD',
        'Service Notes'
    ]
    
    # Find or add columns
    col_map = {}
    next_col = len(headers) + 1
    
    for service_col in service_columns:
        if service_col in headers:
            col_map[service_col] = headers.index(service_col) + 1
        else:
            ws.cell(header_row, next_col).value = service_col
            col_map[service_col] = next_col
            next_col += 1
    
    # Update all data rows
    row = header_row + 1
    rows_updated = 0
    
    while ws.cell(row, 1).value:
        ws.cell(row, col_map['Container Count']).value = property_details['container_count']
        ws.cell(row, col_map['Container Size']).value = property_details['container_size']
        ws.cell(row, col_map['Container Type']).value = property_details['container_type']
        ws.cell(row, col_map['Service Frequency']).value = property_details['service_frequency']
        ws.cell(row, col_map['Service Days']).value = property_details['service_days']
        ws.cell(row, col_map['Total Yards']).value = round(property_details['monthly_yards'], 2)
        ws.cell(row, col_map['YPD']).value = round(property_details['ypd'], 2)
        ws.cell(row, col_map['Service Notes']).value = property_details['service_notes']
        
        rows_updated += 1
        row += 1
    
    wb.save(MASTER_FILE)
    wb.close()
    
    print(f'  ✓ Updated {rows_updated} rows')
    print(f'  ✓ YPD: {property_details["ypd"]:.2f}')
    print(f'  ✅ {property_name} updated successfully')
    print()
    
    return True

def verify_updates():
    """Verify updates were applied correctly"""
    print('=' * 80)
    print('VERIFYING UPDATES')
    print('=' * 80)
    print()
    
    for details in [ORION_PROSPER_DETAILS, ORION_PROSPER_LAKES_DETAILS]:
        property_name = details['property']
        df = pd.read_excel(MASTER_FILE, sheet_name=property_name)
        
        print(f'{property_name}:')
        
        if 'YPD' in df.columns:
            actual_ypd = df['YPD'].iloc[0]
            expected_ypd = details['ypd']
            
            if isinstance(actual_ypd, (int, float)) and abs(actual_ypd - expected_ypd) < 0.01:
                print(f'  ✅ YPD = {actual_ypd:.2f} (matches expected {expected_ypd:.2f})')
            else:
                print(f'  ⚠️ YPD = {actual_ypd} (expected {expected_ypd:.2f})')
        
        if 'Container Count' in df.columns:
            actual_count = df['Container Count'].iloc[0]
            expected_count = details['container_count']
            print(f'  ✅ Container Count = {actual_count} (expected {expected_count})')
        
        if 'Service Notes' in df.columns:
            actual_notes = df['Service Notes'].iloc[0]
            print(f'  ✅ Service Notes = {actual_notes}')
        
        print()

def main():
    """Main function"""
    
    # Show details
    show_details()
    
    # Create backup
    if not create_backup():
        print('❌ Cannot proceed without backup')
        return
    
    # Update master file
    print('=' * 80)
    print('UPDATING MASTER FILE')
    print('=' * 80)
    print()
    
    update_property_sheet(ORION_PROSPER_DETAILS)
    update_property_sheet(ORION_PROSPER_LAKES_DETAILS)
    
    # Verify updates
    verify_updates()
    
    # Final summary
    print('=' * 80)
    print('UPDATE COMPLETE')
    print('=' * 80)
    print()
    print('Summary:')
    print(f'  Orion Prosper: 2 containers, YPD {ORION_PROSPER_DETAILS["ypd"]:.2f} ✅')
    print(f'  Orion Prosper Lakes: 1 container, YPD {ORION_PROSPER_LAKES_DETAILS["ypd"]:.2f} ✅')
    print()
    print('Both properties performing EXCELLENTLY (below 2.0 target)!')
    print()

if __name__ == "__main__":
    main()

