"""
Expense Report Generator - Creates comprehensive Excel workbooks for property expense analysis

This script generates budget-friendly Excel workbooks from extracted monthly expense data.
Each workbook includes multiple tabs with detailed expense tracking, calculations, and visualizations
for property management teams to use in budget planning and variance analysis.

Author: Claude Code
Date: November 13, 2025
"""

import pandas as pd
import json
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference


class ExpenseReportGenerator:
    """Generate comprehensive Excel workbooks from extracted expense data"""

    def __init__(self, property_config_path, vendor_mapping_path):
        """
        Initialize generator with configuration

        Args:
            property_config_path: Path to property_config.json
            vendor_mapping_path: Path to vendor_name_mapping.json
        """
        with open(property_config_path, 'r') as f:
            config_data = json.load(f)
            self.properties = config_data['properties']

        with open(vendor_mapping_path, 'r') as f:
            self.vendor_mapping = json.load(f)

    def generate_report(self, property_name, expense_csv_path, validation_json_path, output_path):
        """
        Generate comprehensive expense workbook for a property

        Args:
            property_name: Name of the property
            expense_csv_path: Path to extracted expense CSV
            validation_json_path: Path to validation JSON
            output_path: Path for output Excel file
        """
        print(f"\n{'='*70}")
        print(f"GENERATING EXPENSE REPORT: {property_name}")
        print(f"{'='*70}")

        # Load data
        expense_df = pd.read_csv(expense_csv_path)
        with open(validation_json_path, 'r') as f:
            validation = json.load(f)

        prop_config = self.properties[property_name]

        print(f"Loaded {len(expense_df)} months of expense data")

        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Generate tabs
        self._create_summary_tab(wb, property_name, expense_df, prop_config, validation)
        self._create_expense_detail_tab(wb, property_name, expense_df, prop_config)
        self._create_budget_projection_tab(wb, property_name, expense_df, prop_config)
        self._create_service_details_tab(wb, property_name, prop_config)
        self._create_validation_tab(wb, property_name, validation)

        # Save workbook
        wb.save(output_path)
        print(f"[OK] Workbook saved: {output_path}")
        print(f"     Tabs created: {len(wb.sheetnames)}")
        print(f"     File size: {Path(output_path).stat().st_size / 1024:.1f} KB")

        return output_path

    def _create_summary_tab(self, wb, property_name, expense_df, prop_config, validation):
        """Create Executive Summary tab"""
        ws = wb.create_sheet("Executive Summary", 0)

        # Header styling
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=14)

        # Title
        ws['A1'] = f"{property_name} - Expense Analysis"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws['A2'].font = Font(italic=True, size=10)
        ws.merge_cells('A2:D2')

        # Property Information section
        row = 4
        ws[f'A{row}'] = "PROPERTY INFORMATION"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')

        row += 1
        info_data = [
            ["Property Name:", property_name],
            ["Units:", prop_config['units']],
            ["State:", prop_config['state']],
            ["Service Type:", prop_config['service_type']],
            ["Data Pattern:", prop_config['data_pattern']]
        ]

        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1

        # Financial Summary section
        row += 1
        ws[f'A{row}'] = "FINANCIAL SUMMARY"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:D{row}')

        row += 1
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        ws[f'C{row}'] = "Per Unit"
        ws[f'D{row}'] = "Notes"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        row += 1
        total_spend = expense_df['Amount'].sum()
        months_covered = len(expense_df)
        avg_monthly = total_spend / months_covered
        avg_cpd = expense_df['Cost_Per_Door'].mean()

        summary_data = [
            ["Total Spend", f"${total_spend:,.2f}", f"${total_spend/prop_config['units']:,.2f}", f"{months_covered} months"],
            ["Average Monthly", f"${avg_monthly:,.2f}", f"${avg_cpd:,.2f}", "Per unit per month"],
            ["Highest Month", f"${expense_df['Amount'].max():,.2f}", f"${expense_df['Cost_Per_Door'].max():,.2f}", expense_df.loc[expense_df['Amount'].idxmax(), 'Month']],
            ["Lowest Month", f"${expense_df['Amount'].min():,.2f}", f"${expense_df['Cost_Per_Door'].min():,.2f}", expense_df.loc[expense_df['Amount'].idxmin(), 'Month']],
        ]

        for metric, value, per_unit, notes in summary_data:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'C{row}'] = per_unit
            ws[f'D{row}'] = notes
            row += 1

        # Anomalies section
        row += 1
        ws[f'A{row}'] = "ANOMALIES DETECTED"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:D{row}')

        row += 1
        anomalies = expense_df[expense_df['Notes'].notna() & (expense_df['Notes'] != '')]
        if len(anomalies) > 0:
            ws[f'A{row}'] = f"{len(anomalies)} months flagged for review"
            ws[f'A{row}'].font = Font(color="C00000", bold=True)
            row += 1
            for _, anom in anomalies.iterrows():
                ws[f'A{row}'] = anom['Month']
                ws[f'B{row}'] = f"${anom['Amount']:,.2f}"
                ws[f'C{row}'] = anom['Notes']
                ws.merge_cells(f'C{row}:D{row}')
                row += 1
        else:
            ws[f'A{row}'] = "No anomalies detected"
            ws[f'A{row}'].font = Font(color="00B050")

        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30

    def _create_expense_detail_tab(self, wb, property_name, expense_df, prop_config):
        """Create detailed expense tracking tab"""
        ws = wb.create_sheet("Monthly Expense Detail")

        # Header
        headers = ['Month', 'Invoice Number(s)', 'Invoice Date', 'Vendor', 'Amount',
                   'Cost Per Door', 'YTD Total', 'YTD Avg CPD', 'Notes/Flags']

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Data rows
        for row_num, (_, row_data) in enumerate(expense_df.iterrows(), 2):
            ws.cell(row=row_num, column=1, value=row_data['Month'])
            ws.cell(row=row_num, column=2, value=row_data['Invoice Number'])
            ws.cell(row=row_num, column=3, value=row_data['Invoice Date'])
            ws.cell(row=row_num, column=4, value=row_data['Vendor'])

            # Amount - formatted as currency
            amount_cell = ws.cell(row=row_num, column=5, value=row_data['Amount'])
            amount_cell.number_format = '$#,##0.00'

            # Cost Per Door - formatted as currency
            cpd_cell = ws.cell(row=row_num, column=6, value=row_data['Cost_Per_Door'])
            cpd_cell.number_format = '$#,##0.00'

            # YTD Total - formatted as currency
            ytd_cell = ws.cell(row=row_num, column=7, value=row_data['YTD_Total'])
            ytd_cell.number_format = '$#,##0.00'

            # YTD Avg CPD - formatted as currency
            ytd_avg_cell = ws.cell(row=row_num, column=8, value=row_data['YTD_Avg_CPD'])
            ytd_avg_cell.number_format = '$#,##0.00'

            # Notes
            notes = row_data['Notes'] if pd.notna(row_data['Notes']) else ''
            notes_cell = ws.cell(row=row_num, column=9, value=notes)

            # Highlight rows with anomalies
            if notes and ('Cost increased' in notes or 'Cost decreased' in notes or 'Overage' in notes):
                for col in range(1, 10):
                    ws.cell(row=row_num, column=col).fill = PatternFill(
                        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                    )

        # Add totals row
        total_row = len(expense_df) + 2
        ws.cell(row=total_row, column=4, value="TOTAL:").font = Font(bold=True)

        # Sum formula for Amount
        ws.cell(row=total_row, column=5, value=f"=SUM(E2:E{total_row-1})")
        ws.cell(row=total_row, column=5).font = Font(bold=True)
        ws.cell(row=total_row, column=5).number_format = '$#,##0.00'

        # Average formula for Cost Per Door
        ws.cell(row=total_row, column=6, value=f"=AVERAGE(F2:F{total_row-1})")
        ws.cell(row=total_row, column=6).font = Font(bold=True)
        ws.cell(row=total_row, column=6).number_format = '$#,##0.00'

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 40

        # Freeze panes (header row)
        ws.freeze_panes = 'A2'

    def _create_budget_projection_tab(self, wb, property_name, expense_df, prop_config):
        """Create budget projection tab for forward planning"""
        ws = wb.create_sheet("Budget Projection")

        # Title
        ws['A1'] = f"{property_name} - Budget Projection"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')

        ws['A2'] = "Based on actual expense data for budget planning"
        ws['A2'].font = Font(italic=True, size=10)
        ws.merge_cells('A2:F2')

        # Historical averages
        row = 4
        ws[f'A{row}'] = "HISTORICAL AVERAGES"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws[f'A{row}'].font = Font(color="FFFFFF", bold=True)
        ws.merge_cells(f'A{row}:F{row}')

        row += 1
        avg_monthly = expense_df['Amount'].mean()
        avg_cpd = expense_df['Cost_Per_Door'].mean()
        months_data = len(expense_df)

        ws[f'A{row}'] = "Average Monthly Expense:"
        ws[f'B{row}'] = avg_monthly
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Average Cost Per Door:"
        ws[f'B{row}'] = avg_cpd
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Data Period:"
        ws[f'B{row}'] = f"{months_data} months ({expense_df['Month'].min()} to {expense_df['Month'].max()})"
        ws[f'A{row}'].font = Font(bold=True)

        # Projection table
        row += 2
        ws[f'A{row}'] = "BUDGET PROJECTION (Next 12 Months)"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws[f'A{row}'].font = Font(color="FFFFFF", bold=True)
        ws.merge_cells(f'A{row}:F{row}')

        row += 1
        projection_headers = ['Period', 'Projected Monthly', 'Cost Per Door', 'Quarterly Total', 'Annual Total', 'Notes']
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        for col_num, header in enumerate(projection_headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = header_fill

        # Quarterly projections
        row += 1
        quarters = ['Q1', 'Q2', 'Q3', 'Q4']
        for q_num, quarter in enumerate(quarters, 1):
            ws.cell(row=row, column=1, value=quarter)

            # Projected monthly
            monthly_cell = ws.cell(row=row, column=2, value=avg_monthly)
            monthly_cell.number_format = '$#,##0.00'

            # CPD
            cpd_cell = ws.cell(row=row, column=3, value=avg_cpd)
            cpd_cell.number_format = '$#,##0.00'

            # Quarterly total (3 months)
            quarterly_cell = ws.cell(row=row, column=4, value=avg_monthly * 3)
            quarterly_cell.number_format = '$#,##0.00'

            # Annual running total
            annual_cell = ws.cell(row=row, column=5, value=avg_monthly * 3 * q_num)
            annual_cell.number_format = '$#,##0.00'

            # Notes
            ws.cell(row=row, column=6, value="Based on historical average")

            row += 1

        # Annual total
        ws.cell(row=row, column=1, value="ANNUAL TOTAL:").font = Font(bold=True)
        annual_total_cell = ws.cell(row=row, column=5, value=avg_monthly * 12)
        annual_total_cell.number_format = '$#,##0.00'
        annual_total_cell.font = Font(bold=True)
        annual_total_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        # Variance scenarios
        row += 2
        ws[f'A{row}'] = "VARIANCE SCENARIOS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws[f'A{row}'].font = Font(color="FFFFFF", bold=True)
        ws.merge_cells(f'A{row}:F{row}')

        row += 1
        ws[f'A{row}'] = "Scenario"
        ws[f'B{row}'] = "Monthly"
        ws[f'C{row}'] = "Annual"
        ws[f'D{row}'] = "Variance"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = header_fill

        row += 1
        scenarios = [
            ("Best Case (-10%)", avg_monthly * 0.9, avg_monthly * 0.9 * 12, -0.10),
            ("Expected (Baseline)", avg_monthly, avg_monthly * 12, 0.00),
            ("Conservative (+10%)", avg_monthly * 1.1, avg_monthly * 1.1 * 12, 0.10),
            ("Worst Case (+20%)", avg_monthly * 1.2, avg_monthly * 1.2 * 12, 0.20),
        ]

        for scenario_name, monthly, annual, variance in scenarios:
            ws.cell(row=row, column=1, value=scenario_name)

            monthly_cell = ws.cell(row=row, column=2, value=monthly)
            monthly_cell.number_format = '$#,##0.00'

            annual_cell = ws.cell(row=row, column=3, value=annual)
            annual_cell.number_format = '$#,##0.00'

            variance_cell = ws.cell(row=row, column=4, value=variance)
            variance_cell.number_format = '0.0%'

            # Color code
            if variance < 0:
                ws.cell(row=row, column=4).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif variance > 0.1:
                ws.cell(row=row, column=4).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 30

    def _create_service_details_tab(self, wb, property_name, prop_config):
        """Create service details reference tab"""
        ws = wb.create_sheet("Service Details")

        # Title
        ws['A1'] = f"{property_name} - Service Configuration"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')

        row = 3
        details = [
            ("Property Name", property_name),
            ("Units", prop_config['units']),
            ("State", prop_config['state']),
            ("City", prop_config.get('city', 'N/A')),
            ("Property Type", prop_config['property_type']),
            ("Service Type", prop_config['service_type']),
            ("", ""),
            ("Data Configuration", ""),
            ("Data Pattern", prop_config['data_pattern']),
            ("Amount Field", prop_config['amount_field']),
            ("", ""),
            ("Data Coverage", ""),
            ("Date Range Start", prop_config['date_range']['start']),
            ("Date Range End", prop_config['date_range']['end']),
            ("Months Covered", prop_config['date_range']['months_covered']),
            ("Total Records", prop_config['total_records']),
            ("", ""),
            ("Vendors", ""),
            ("Unique Vendors", prop_config['unique_vendors']),
        ]

        for label, value in details:
            if label == "":
                row += 1
                continue

            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value

            if label in ["Data Configuration", "Data Coverage", "Vendors"]:
                ws[f'A{row}'].font = Font(bold=True, size=12)
                ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

            row += 1

        # Notes section
        if 'notes' in prop_config and prop_config['notes']:
            row += 1
            ws[f'A{row}'] = "NOTES"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            row += 1
            ws[f'A{row}'] = prop_config['notes']
            ws.merge_cells(f'A{row}:B{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40

    def _create_validation_tab(self, wb, property_name, validation):
        """Create validation results tab"""
        ws = wb.create_sheet("Validation")

        # Title
        ws['A1'] = "Data Validation Report"
        ws['A1'].font = Font(bold=True, size=14)

        ws['A2'] = f"Property: {property_name}"
        ws['A3'] = f"Extraction Date: {validation.get('extraction_date', 'N/A')}"
        ws['A4'] = f"Overall Status: {validation['status']}"

        # Status color
        if validation['status'] == 'PASSED':
            ws['A4'].font = Font(color="00B050", bold=True)
        else:
            ws['A4'].font = Font(color="C00000", bold=True)

        # Validation checks
        row = 6
        ws[f'A{row}'] = "VALIDATION CHECKS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws[f'A{row}'].font = Font(color="FFFFFF", bold=True)
        ws.merge_cells(f'A{row}:D{row}')

        row += 1
        ws[f'A{row}'] = "Check"
        ws[f'B{row}'] = "Expected"
        ws[f'C{row}'] = "Actual"
        ws[f'D{row}'] = "Status"

        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        row += 1
        for check_name, check_data in validation['checks'].items():
            ws[f'A{row}'] = check_name.replace('_', ' ').title()

            if 'expected' in check_data:
                ws[f'B{row}'] = check_data['expected']
            if 'extracted' in check_data:
                ws[f'C{row}'] = check_data['extracted']

            # Format currency values
            if 'spend' in check_name:
                if 'expected' in check_data:
                    ws[f'B{row}'].number_format = '$#,##0.00'
                if 'extracted' in check_data:
                    ws[f'C{row}'].number_format = '$#,##0.00'

            status = "PASS" if check_data.get('passed', False) else "FAIL"
            ws[f'D{row}'] = status

            if status == "PASS":
                ws[f'D{row}'].font = Font(color="00B050", bold=True)
            else:
                ws[f'D{row}'].font = Font(color="C00000", bold=True)

            row += 1

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10


def main():
    """Generate expense reports for pilot properties"""
    import sys

    base_dir = Path(r'C:\Users\Richard\Downloads\Orion Data Part 2')
    config_file = base_dir / 'Code' / 'property_config.json'
    vendor_mapping = base_dir / 'Code' / 'vendor_name_mapping.json'

    # Get property name from command line
    if len(sys.argv) < 2:
        print("Usage: python generate_expense_reports.py <property_name>")
        print("\nPilot properties:")
        print("  - Springs at Alta Mesa")
        print("  - Orion Prosper")
        print("  - The Club at Millenia")
        return

    property_name = sys.argv[1]

    # Paths
    property_folder = base_dir / 'Properties' / property_name.replace(' ', '_')
    expense_csv = property_folder / f"{property_name.replace(' ', '_')}_expense_data.csv"
    validation_json = property_folder / f"{property_name.replace(' ', '_')}_validation.json"
    output_excel = property_folder / f"{property_name.replace(' ', '_')}_Expense_Report.xlsx"

    # Validate inputs
    if not expense_csv.exists():
        print(f"[FAIL] Expense CSV not found: {expense_csv}")
        return

    if not validation_json.exists():
        print(f"[FAIL] Validation JSON not found: {validation_json}")
        return

    # Generate report
    generator = ExpenseReportGenerator(config_file, vendor_mapping)
    output_path = generator.generate_report(
        property_name=property_name,
        expense_csv_path=expense_csv,
        validation_json_path=validation_json,
        output_path=output_excel
    )

    print(f"\n{'='*70}")
    print("REPORT GENERATION COMPLETE")
    print(f"{'='*70}")
    print(f"Property: {property_name}")
    print(f"Output: {output_path}")
    print(f"\nWorkbook includes:")
    print("  1. Executive Summary - Key metrics and anomalies")
    print("  2. Monthly Expense Detail - Invoice-level tracking")
    print("  3. Budget Projection - Forward planning scenarios")
    print("  4. Service Details - Property configuration")
    print("  5. Validation - Data quality checks")


if __name__ == '__main__':
    main()
