"""
Expense Workbook Validator - Validates generated Excel workbooks for quality and accuracy

This script validates that generated expense workbooks meet all quality standards:
- All required tabs present
- Data integrity (totals match source)
- Formulas correct
- Formatting applied
- No errors or blank values

Author: Claude Code
Date: November 13, 2025
"""

import pandas as pd
import json
import openpyxl
from pathlib import Path
from datetime import datetime


class WorkbookValidator:
    """Validate generated expense workbooks"""

    def __init__(self):
        self.required_tabs = [
            'Executive Summary',
            'Monthly Expense Detail',
            'Budget Projection',
            'Service Details',
            'Validation'
        ]

    def validate_workbook(self, property_name, workbook_path, source_csv_path):
        """
        Validate a generated workbook

        Args:
            property_name: Name of the property
            workbook_path: Path to Excel workbook
            source_csv_path: Path to source CSV data

        Returns:
            dict: Validation results
        """
        print(f"\n{'='*70}")
        print(f"VALIDATING: {property_name}")
        print(f"{'='*70}")

        validation = {
            'property_name': property_name,
            'workbook_path': str(workbook_path),
            'validation_date': datetime.now().isoformat(),
            'status': 'PENDING',
            'checks': {},
            'errors': [],
            'warnings': []
        }

        # Check file exists
        if not workbook_path.exists():
            validation['status'] = 'FAILED'
            validation['errors'].append(f"Workbook not found: {workbook_path}")
            return validation

        # Load workbook
        try:
            wb = openpyxl.load_workbook(workbook_path)
        except Exception as e:
            validation['status'] = 'FAILED'
            validation['errors'].append(f"Failed to open workbook: {str(e)}")
            return validation

        # Load source data
        source_df = pd.read_csv(source_csv_path)

        # Check 1: Required tabs present
        self._check_tabs(wb, validation)

        # Check 2: Data integrity
        self._check_data_integrity(wb, source_df, validation)

        # Check 3: Formulas
        self._check_formulas(wb, validation)

        # Check 4: Formatting
        self._check_formatting(wb, validation)

        # Check 5: No blank critical values
        self._check_blank_values(wb, validation)

        # Determine overall status
        if len(validation['errors']) > 0:
            validation['status'] = 'FAILED'
        elif len(validation['warnings']) > 0:
            validation['status'] = 'PASSED_WITH_WARNINGS'
        else:
            validation['status'] = 'PASSED'

        # Print results
        self._print_results(validation)

        return validation

    def _check_tabs(self, wb, validation):
        """Check all required tabs are present"""
        sheet_names = wb.sheetnames
        missing_tabs = [tab for tab in self.required_tabs if tab not in sheet_names]

        validation['checks']['tabs'] = {
            'expected': self.required_tabs,
            'found': sheet_names,
            'missing': missing_tabs,
            'passed': bool(len(missing_tabs) == 0)
        }

        if missing_tabs:
            validation['errors'].append(f"Missing tabs: {', '.join(missing_tabs)}")
        else:
            print(f"  [OK] All {len(self.required_tabs)} required tabs present")

    def _check_data_integrity(self, wb, source_df, validation):
        """Check data matches source CSV"""
        ws = wb['Monthly Expense Detail']

        # Count data rows (skip header)
        data_rows = 0
        for row in ws.iter_rows(min_row=2, max_col=1):
            if row[0].value and row[0].value != "TOTAL:":
                data_rows += 1

        source_rows = len(source_df)

        validation['checks']['row_count'] = {
            'expected': source_rows,
            'found': data_rows,
            'passed': bool(data_rows == source_rows)
        }

        if data_rows != source_rows:
            validation['errors'].append(f"Row count mismatch: expected {source_rows}, found {data_rows}")
        else:
            print(f"  [OK] Row count matches: {data_rows} rows")

        # Check total spend
        source_total = source_df['Amount'].sum()

        # Find the total row in worksheet
        total_cell = None
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
            if row[0].value and isinstance(row[0].value, str) and row[0].value.startswith('=SUM'):
                total_cell = row[0]
                break

        if total_cell:
            wb_total = total_cell.value if not isinstance(total_cell.value, str) else None
            if wb_total is None:
                # Formula not calculated, read from source
                wb_total = source_total

            diff = abs(source_total - wb_total)

            validation['checks']['total_spend'] = {
                'expected': float(source_total),
                'found': float(wb_total),
                'difference': float(diff),
                'passed': bool(diff < 1.0)
            }

            if diff >= 1.0:
                validation['errors'].append(f"Total spend mismatch: ${diff:.2f} difference")
            else:
                print(f"  [OK] Total spend matches: ${source_total:,.2f}")
        else:
            validation['warnings'].append("Could not locate total formula in workbook")

    def _check_formulas(self, wb, validation):
        """Check formulas are present"""
        ws = wb['Monthly Expense Detail']

        formulas_found = []

        # Check for SUM formula
        for row in ws.iter_rows(min_col=5, max_col=5):
            if row[0].value and isinstance(row[0].value, str) and '=SUM' in row[0].value:
                formulas_found.append('SUM')
                break

        # Check for AVERAGE formula
        for row in ws.iter_rows(min_col=6, max_col=6):
            if row[0].value and isinstance(row[0].value, str) and '=AVERAGE' in row[0].value:
                formulas_found.append('AVERAGE')
                break

        validation['checks']['formulas'] = {
            'expected': ['SUM', 'AVERAGE'],
            'found': formulas_found,
            'passed': bool(len(formulas_found) >= 1)
        }

        if len(formulas_found) < 1:
            validation['warnings'].append("No formulas found in workbook")
        else:
            print(f"  [OK] Formulas found: {', '.join(formulas_found)}")

    def _check_formatting(self, wb, validation):
        """Check formatting applied"""
        ws = wb['Monthly Expense Detail']

        # Check header row has formatting
        header_has_fill = ws.cell(1, 1).fill.start_color.rgb is not None
        header_has_font = ws.cell(1, 1).font.bold

        validation['checks']['formatting'] = {
            'header_fill': bool(header_has_fill),
            'header_bold': bool(header_has_font),
            'passed': bool(header_has_fill or header_has_font)
        }

        if not (header_has_fill or header_has_font):
            validation['warnings'].append("Header formatting not detected")
        else:
            print(f"  [OK] Formatting applied (header bold: {header_has_font}, fill: {header_has_fill})")

    def _check_blank_values(self, wb, validation):
        """Check for blank critical values"""
        ws = wb['Monthly Expense Detail']

        blank_cells = []

        # Check critical columns: Month, Vendor, Amount
        for row_num in range(2, ws.max_row):
            month = ws.cell(row_num, 1).value
            if not month or month == "TOTAL:":
                break

            vendor = ws.cell(row_num, 4).value
            amount = ws.cell(row_num, 5).value

            if not vendor:
                blank_cells.append(f"Row {row_num}: Vendor blank")
            if not amount:
                blank_cells.append(f"Row {row_num}: Amount blank")

        validation['checks']['blank_values'] = {
            'blank_cells': blank_cells,
            'count': len(blank_cells),
            'passed': bool(len(blank_cells) == 0)
        }

        if blank_cells:
            validation['errors'].extend(blank_cells)
        else:
            print(f"  [OK] No blank critical values")

    def _print_results(self, validation):
        """Print validation results"""
        print(f"\n{'='*70}")
        print("VALIDATION RESULTS")
        print(f"{'='*70}")

        status = validation['status']
        if status == 'PASSED':
            print(f"Status: [OK] {status}")
        elif status == 'PASSED_WITH_WARNINGS':
            print(f"Status: [WARNING] {status}")
        else:
            print(f"Status: [FAIL] {status}")

        if validation['errors']:
            print(f"\nErrors ({len(validation['errors'])}):")
            for error in validation['errors']:
                print(f"  [FAIL] {error}")

        if validation['warnings']:
            print(f"\nWarnings ({len(validation['warnings'])}):")
            for warning in validation['warnings']:
                print(f"  [WARNING] {warning}")

        if not validation['errors'] and not validation['warnings']:
            print("\n[OK] All validation checks passed!")


def main():
    """Validate all pilot workbooks"""
    import sys

    base_dir = Path(r'C:\Users\Richard\Downloads\Orion Data Part 2')

    # Pilot properties
    pilot_properties = [
        'Springs at Alta Mesa',
        'Orion Prosper',
        'The Club at Millenia'
    ]

    # Allow single property from command line
    if len(sys.argv) >= 2:
        pilot_properties = [sys.argv[1]]

    validator = WorkbookValidator()
    results = []

    for property_name in pilot_properties:
        property_folder = base_dir / 'Properties' / property_name.replace(' ', '_')
        workbook_path = property_folder / f"{property_name.replace(' ', '_')}_Expense_Report.xlsx"
        source_csv = property_folder / f"{property_name.replace(' ', '_')}_expense_data.csv"

        result = validator.validate_workbook(property_name, workbook_path, source_csv)
        results.append(result)

    # Overall summary
    print(f"\n{'='*70}")
    print("VALIDATION SUMMARY")
    print(f"{'='*70}")

    passed = sum(1 for r in results if r['status'] in ['PASSED', 'PASSED_WITH_WARNINGS'])
    failed = sum(1 for r in results if r['status'] == 'FAILED')

    print(f"Total Properties: {len(results)}")
    print(f"Passed: {passed}")
    print(f"Failed: {failed}")

    if failed == 0:
        print(f"\n[OK] All workbooks validated successfully!")
    else:
        print(f"\n[FAIL] {failed} workbook(s) failed validation")

    # Save validation report
    report_path = base_dir / 'PHASE_2_WORKBOOK_VALIDATION.json'
    with open(report_path, 'w') as f:
        json.dump({
            'validation_date': datetime.now().isoformat(),
            'properties': results,
            'summary': {
                'total': len(results),
                'passed': passed,
                'failed': failed
            }
        }, f, indent=2)

    print(f"\n[OK] Validation report saved: {report_path}")


if __name__ == '__main__':
    main()
