"""
Update Property_Reference_Sheet with data from MASTER_Portfolio_Complete_Data
and regulatory compliance research findings
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

# File paths
MASTER_FILE = "Portfolio_Reports/MASTER_Portfolio_Complete_Data.xlsx"
REFERENCE_FILE = "Portfolio_Reports/Property_Reference_Sheet.xlsx"
OUTPUT_FILE = "Portfolio_Reports/Property_Reference_Sheet.xlsx"

# Regulatory data from our research
REGULATORY_DATA = {
    'Orion Prosper': {
        'city': 'Prosper',
        'state': 'TX',
        'county': 'Collin County',
        'regulatory_status': 'VOLUNTARY',
        'mandatory_recycling': 'NO'
    },
    'Orion Prosper Lakes': {
        'city': 'Prosper',
        'state': 'TX',
        'county': 'Collin County',
        'regulatory_status': 'VOLUNTARY',
        'mandatory_recycling': 'NO'
    },
    'Orion McKinney': {
        'city': 'McKinney',
        'state': 'TX',
        'county': 'Collin County',
        'regulatory_status': 'COMMERCIAL',
        'mandatory_recycling': 'NO'
    },
    'McCord Park FL': {
        'city': 'Little Elm',
        'state': 'TX',
        'county': 'Denton County',
        'regulatory_status': 'MANDATORY',
        'mandatory_recycling': 'YES'
    },
    'The Club at Millenia': {
        'city': 'Orlando',
        'state': 'FL',
        'county': 'Orange County',
        'regulatory_status': 'MANDATORY',
        'mandatory_recycling': 'YES'
    },
    'Bella Mirage': {
        'city': 'Phoenix',
        'state': 'AZ',
        'county': 'Maricopa County',
        'regulatory_status': 'VOLUNTARY',
        'mandatory_recycling': 'NO'
    },
    'Mandarina': {
        'city': 'Phoenix',
        'state': 'AZ',
        'county': 'Maricopa County',
        'regulatory_status': 'VOLUNTARY',
        'mandatory_recycling': 'NO'
    },
    'Pavilions at Arrowhead': {
        'city': 'Glendale',
        'state': 'AZ',
        'county': 'Maricopa County',
        'regulatory_status': 'VOLUNTARY',
        'mandatory_recycling': 'NO'
    },
    'Springs at Alta Mesa': {
        'city': 'Mesa',
        'state': 'AZ',
        'county': 'Maricopa County',
        'regulatory_status': 'VOLUNTARY',
        'mandatory_recycling': 'NO'
    },
    'Tempe Vista': {
        'city': 'Tempe',
        'state': 'AZ',
        'county': 'Maricopa County',
        'regulatory_status': 'VOLUNTARY',
        'mandatory_recycling': 'NO'
    }
}

print("=" * 80)
print("UPDATING PROPERTY REFERENCE SHEET")
print("=" * 80)
print(f"\nReading data from master file...")

# Read master data
df_master = pd.read_excel(MASTER_FILE, sheet_name='Property Overview')

# Read current reference sheet
df_ref = pd.read_excel(REFERENCE_FILE)

print(f"  Master file properties: {len(df_master) - 1}")  # -1 for PORTFOLIO TOTAL row
print(f"  Reference sheet properties: {len(df_ref)}")

# Remove PORTFOLIO TOTAL row from master
df_master = df_master[df_master['Property Name'] != 'PORTFOLIO TOTAL'].copy()

print(f"\nUpdating reference sheet with master data...")

# Create updated dataframe
updated_data = []

for idx, row in df_ref.iterrows():
    prop_name = row['Property Name']

    # Get data from master file
    master_row = df_master[df_master['Property Name'] == prop_name]

    if len(master_row) > 0:
        master_row = master_row.iloc[0]
        reg_data = REGULATORY_DATA.get(prop_name, {})

        # Build updated row
        updated_row = {
            'Property Name': prop_name,
            'Address': row['Address'] if pd.notna(row['Address']) else 'TBD',
            'City': reg_data.get('city', row['City']),
            'State': reg_data.get('state', row['State']),
            'County': reg_data.get('county', 'TBD'),
            'Zip Code': row['Zip Code'] if pd.notna(row['Zip Code']) else 'TBD',
            'Units': int(master_row['Unit Count']) if pd.notna(master_row['Unit Count']) else row['Units'],
            'Property Type': master_row['Property Type'] if pd.notna(master_row['Property Type']) else row['Property Type'],
            'Vendor': row['Vendor'],
            'Service Type': master_row['Service Type'] if pd.notna(master_row['Service Type']) else row['Service Type'],
            'Container Count': int(master_row['Container Count']) if pd.notna(master_row['Container Count']) else 'TBD',
            'Container Size': master_row['Container Size'] if pd.notna(master_row['Container Size']) else 'TBD',
            'Service Frequency': master_row['Service Frequency'] if pd.notna(master_row['Service Frequency']) else 'TBD',
            'Monthly Yards': f"{master_row['Monthly Yards']:.2f}" if pd.notna(master_row['Monthly Yards']) else 'TBD',
            'YPD': f"{master_row['YPD']:.2f}" if pd.notna(master_row['YPD']) else 'TBD',
            'Regulatory Status': reg_data.get('regulatory_status', 'TBD'),
            'Mandatory Recycling': reg_data.get('mandatory_recycling', 'TBD'),
            'PDF Invoices': row['PDF Invoices'] if pd.notna(row['PDF Invoices']) else 0,
            'Excel Files': row['Excel Files'] if pd.notna(row['Excel Files']) else 0,
            'Contracts': row['Contracts'] if pd.notna(row['Contracts']) else 0,
            'Extracted Records': row['Extracted Records'] if pd.notna(row['Extracted Records']) else 0,
            'Status': row['Status'],
            'Notes': row['Notes'] if pd.notna(row['Notes']) else ''
        }

        updated_data.append(updated_row)

        print(f"  Updated: {prop_name}")
        print(f"    City: {updated_row['City']}, Units: {updated_row['Units']}, YPD: {updated_row['YPD']}")
    else:
        print(f"  [WARNING] {prop_name} not found in master file - keeping original data")
        updated_data.append(row.to_dict())

# Create new dataframe
df_updated = pd.DataFrame(updated_data)

# Reorder columns for better presentation
column_order = [
    'Property Name', 'Address', 'City', 'State', 'County', 'Zip Code', 'Units',
    'Property Type', 'Vendor', 'Service Type', 'Container Count', 'Container Size',
    'Service Frequency', 'Monthly Yards', 'YPD', 'Regulatory Status', 'Mandatory Recycling',
    'PDF Invoices', 'Excel Files', 'Contracts', 'Extracted Records', 'Status', 'Notes'
]

df_updated = df_updated[column_order]

print(f"\nWriting updated data to Excel...")

# Create workbook with formatting
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Property Reference"

# Styling
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Title
ws['A1'] = "Property Reference Sheet - Updated with Master Data"
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:W1')

ws['A2'] = f"Last Updated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
ws['A2'].font = Font(size=10, italic=True)
ws.merge_cells('A2:W2')

# Headers
row = 4
for col_idx, column in enumerate(df_updated.columns, start=1):
    cell = ws.cell(row=row, column=col_idx)
    cell.value = column
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER

# Data rows
for idx, data_row in df_updated.iterrows():
    row += 1
    for col_idx, value in enumerate(data_row, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        cell.alignment = LEFT_ALIGN if col_idx <= 3 or col_idx >= 20 else CENTER_ALIGN
        cell.border = THIN_BORDER

        # Color coding for regulatory status
        if df_updated.columns[col_idx-1] == 'Mandatory Recycling':
            if value == 'YES':
                cell.font = Font(color="FF0000", bold=True)
            elif value == 'NO':
                cell.font = Font(color="008000", bold=True)

# Column widths
column_widths = {
    'A': 25,  # Property Name
    'B': 30,  # Address
    'C': 15,  # City
    'D': 8,   # State
    'E': 15,  # County
    'F': 10,  # Zip Code
    'G': 8,   # Units
    'H': 15,  # Property Type
    'I': 25,  # Vendor
    'J': 18,  # Service Type
    'K': 12,  # Container Count
    'L': 15,  # Container Size
    'M': 20,  # Service Frequency
    'N': 12,  # Monthly Yards
    'O': 8,   # YPD
    'P': 18,  # Regulatory Status
    'Q': 18,  # Mandatory Recycling
    'R': 12,  # PDF Invoices
    'S': 12,  # Excel Files
    'T': 10,  # Contracts
    'U': 15,  # Extracted Records
    'V': 12,  # Status
    'W': 30,  # Notes
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Freeze panes
ws.freeze_panes = 'A5'

# Save
wb.save(OUTPUT_FILE)

print(f"\n[OK] Property Reference Sheet updated successfully!")
print(f"     {OUTPUT_FILE}")
print(f"\nSummary of updates:")
print(f"  - Added County column")
print(f"  - Updated Cities (Bella Mirage: Phoenix, Orion Prosper Lakes: Prosper)")
print(f"  - Updated Units (Pavilions: 248, Tempe Vista: 186)")
print(f"  - Added Container Count, Container Size, Service Frequency")
print(f"  - Added Monthly Yards and YPD calculations")
print(f"  - Added Regulatory Status and Mandatory Recycling columns")
print(f"  - Applied color coding (RED=Mandatory, GREEN=No Requirements)")
print("\n" + "=" * 80)
print("PROPERTY REFERENCE SHEET UPDATE COMPLETE")
print("=" * 80)
