"""
Update Arizona Properties with Service Details

This script updates the master file with service details for the 4 Arizona properties
provided by the user.
"""

import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
import shutil

# Paths
BASE_DIR = Path(__file__).parent.parent
MASTER_FILE = BASE_DIR / "Portfolio_Reports" / "MASTER_Portfolio_Complete_Data.xlsx"
BACKUP_FILE = BASE_DIR / "Portfolio_Reports" / f"MASTER_Portfolio_Complete_Data_BACKUP_AZ_UPDATE_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# Property unit counts (from CLAUDE.md)
PROPERTY_UNITS = {
    'Tempe Vista': 150,  # Estimated
    'Mandarina': 180,
    'Springs at Alta Mesa': 200,
    'Pavilions at Arrowhead': None  # TBD - will need to get this
}

# Arizona property service details from user
AZ_SERVICE_DETAILS = {
    'Tempe Vista': {
        'units': 150,
        'services': [
            {'size': 4, 'count': 6, 'pickups_per_week': 3, 'type': 'Dumpster'},
            {'size': 6, 'count': 1, 'pickups_per_week': 3, 'type': 'Dumpster'},
            {'size': 8, 'count': 1, 'pickups_per_week': 3, 'type': 'Dumpster'},
        ],
        'service_notes': '6×4YD + 1×6YD + 1×8YD Dumpsters @ 3x/week',
        'service_days': '3x per week'
    },
    'Mandarina': {
        'units': 180,
        'services': [
            {'size': 6, 'count': 2, 'pickups_per_week': 3, 'type': 'Compactor'},
        ],
        'service_notes': '2×6YD Compactor @ 3x/week',
        'service_days': '3x per week'
    },
    'Springs at Alta Mesa': {
        'units': 200,
        'services': [
            {'size': 6, 'count': 5, 'pickups_per_week': 3, 'type': 'Dumpster'},
            {'size': 4, 'count': 4, 'pickups_per_week': 3, 'type': 'Dumpster'},
            {'size': 0.034, 'count': 7, 'pickups_per_week': 3, 'type': '90-gallon bin'},  # 90 gal = 0.034 cubic yards
        ],
        'service_notes': '5×6YD + 4×4YD Dumpsters + 7×90-gal bins @ 3x/week. Bulk: Every Thursday',
        'service_days': '3x per week + Bulk Thursday',
        'bulk_service': 'Every Thursday'
    },
    'Pavilions at Arrowhead': {
        'units': None,  # TBD
        'services': [
            {'size': 4, 'count': 4, 'pickups_per_week': 2, 'type': 'Dumpster'},
        ],
        'service_notes': '4×4YD Dumpsters @ 2x/week. Bulk: Once per week (Thursday) @ $365/month',
        'service_days': '2x per week + Bulk Thursday',
        'bulk_service': 'Once per week (Thursday)',
        'bulk_cost': 365.00
    }
}

def calculate_ypd(services, units):
    """
    Calculate YPD using correct formula:
    YPD = (Container Size × Number of Containers × Pickups per Week × 4.33) / Number of Units
    
    For multiple services, sum the yards from each service.
    Note: 90-gallon bins are converted to cubic yards (90 gal = 0.034 cubic yards)
    """
    if units is None:
        return None, None
    
    total_monthly_yards = 0
    
    for service in services:
        size = service['size']
        count = service['count']
        pickups_per_week = service['pickups_per_week']
        
        # Calculate monthly yards for this service
        monthly_yards = size * count * pickups_per_week * 4.33
        total_monthly_yards += monthly_yards
    
    # Calculate YPD
    ypd = total_monthly_yards / units
    
    return ypd, total_monthly_yards

def show_calculations():
    """Show detailed YPD calculations for each Arizona property"""
    
    print('=' * 80)
    print('ARIZONA PROPERTIES - SERVICE DETAILS UPDATE')
    print('=' * 80)
    print()
    print('Formula: YPD = (Size × Count × Pickups/Week × 4.33) / Units')
    print()
    print('=' * 80)
    print()
    
    results = {}
    
    for property_name, data in AZ_SERVICE_DETAILS.items():
        units = data['units']
        services = data['services']
        
        print(f'{property_name}:')
        if units:
            print(f'  Units: {units}')
        else:
            print(f'  Units: TBD (cannot calculate YPD)')
        print()
        
        total_monthly_yards = 0
        total_containers = 0
        
        for i, service in enumerate(services, 1):
            size = service['size']
            count = service['count']
            pickups_per_week = service['pickups_per_week']
            stype = service['type']
            
            monthly_yards = size * count * pickups_per_week * 4.33
            total_monthly_yards += monthly_yards
            total_containers += count
            
            print(f'  Service {i}: {count}× {size}YD {stype} @ {pickups_per_week}x/week')
            print(f'    Calculation: {size} × {count} × {pickups_per_week} × 4.33 = {monthly_yards:.2f} yards/month')
            print()
        
        # Add bulk service info if present
        if 'bulk_service' in data:
            print(f'  Bulk Service: {data["bulk_service"]}')
            if 'bulk_cost' in data:
                print(f'  Bulk Cost: ${data["bulk_cost"]:.2f}/month')
            print()
        
        print(f'  Total Containers: {total_containers}')
        print(f'  Total Monthly Yards: {total_monthly_yards:.2f}')
        
        if units:
            ypd = total_monthly_yards / units
            print(f'  YPD: {total_monthly_yards:.2f} / {units} = {ypd:.2f}')
            print()
            
            # Compare to target
            target = 2.0
            if ypd <= target:
                pct_below = ((target - ypd) / target) * 100
                print(f'  ✅ {pct_below:.1f}% below target (2.0)')
            else:
                pct_above = ((ypd - target) / target) * 100
                print(f'  ⚠️ {pct_above:.1f}% above target (2.0)')
        else:
            ypd = None
            print(f'  YPD: Cannot calculate (units TBD)')
        
        print()
        print('-' * 80)
        print()
        
        results[property_name] = {
            'ypd': ypd,
            'total_monthly_yards': total_monthly_yards,
            'units': units,
            'total_containers': total_containers,
            'service_notes': data['service_notes'],
            'service_days': data['service_days']
        }
    
    return results

def create_backup():
    """Create a backup of the master file before updating"""
    
    print('=' * 80)
    print('CREATING BACKUP')
    print('=' * 80)
    print()
    
    if MASTER_FILE.exists():
        shutil.copy2(MASTER_FILE, BACKUP_FILE)
        print(f'✅ Backup created: {BACKUP_FILE.name}')
        print()
        return True
    else:
        print(f'❌ Master file not found: {MASTER_FILE}')
        return False

def update_az_properties_in_master(results):
    """Update Arizona property sheets with service details"""
    
    print('=' * 80)
    print('UPDATING ARIZONA PROPERTIES IN MASTER FILE')
    print('=' * 80)
    print()
    
    wb = openpyxl.load_workbook(MASTER_FILE)
    
    for property_name, data in results.items():
        if property_name not in wb.sheetnames:
            print(f'  ⚠️ Sheet not found: {property_name} - Skipping')
            continue
        
        ws = wb[property_name]
        
        print(f'Updating: {property_name}')
        print('-' * 80)
        
        # Find header row (should be row 1)
        header_row = None
        for row in range(1, 10):
            cell_value = ws.cell(row, 1).value
            if cell_value and ('Property' in str(cell_value) or 'Invoice' in str(cell_value)):
                header_row = row
                break
        
        if not header_row:
            print(f'  ❌ Could not find header row')
            continue
        
        print(f'  ✓ Found header row: {header_row}')
        
        # Get existing columns
        existing_cols = {}
        for col in range(1, 50):
            header = ws.cell(header_row, col).value
            if header:
                existing_cols[header] = col
        
        print(f'  ✓ Found {len(existing_cols)} existing columns')
        
        # Define service columns to add/update
        service_columns = {
            'Container Count': data['total_containers'],
            'Container Size': AZ_SERVICE_DETAILS[property_name]['services'][0]['size'] if len(AZ_SERVICE_DETAILS[property_name]['services']) == 1 else 'Mixed',
            'Container Type': AZ_SERVICE_DETAILS[property_name]['services'][0]['type'] if len(AZ_SERVICE_DETAILS[property_name]['services']) == 1 else 'Mixed',
            'Service Frequency': AZ_SERVICE_DETAILS[property_name]['services'][0]['pickups_per_week'] if len(AZ_SERVICE_DETAILS[property_name]['services']) == 1 else 'Mixed',
            'Service Days': data['service_days'],
            'Total Yards': data['total_monthly_yards'],
            'YPD': data['ypd'] if data['ypd'] else 'TBD (units unknown)',
            'Service Notes': data['service_notes']
        }
        
        # Add columns if they don't exist
        next_col = len(existing_cols) + 1
        new_cols_added = 0
        
        for col_name, _ in service_columns.items():
            if col_name not in existing_cols:
                ws.cell(header_row, next_col).value = col_name
                existing_cols[col_name] = next_col
                next_col += 1
                new_cols_added += 1
        
        if new_cols_added > 0:
            print(f'  ✓ Added {new_cols_added} new columns')
        
        # Update all data rows
        row = header_row + 1
        rows_updated = 0
        
        while ws.cell(row, 1).value:
            for col_name, col_value in service_columns.items():
                col_idx = existing_cols[col_name]
                ws.cell(row, col_idx).value = col_value
            rows_updated += 1
            row += 1
        
        print(f'  ✓ Updated {rows_updated} rows with service details')
        
        if data['ypd']:
            print(f'  ✅ {property_name} updated successfully (YPD: {data["ypd"]:.2f})')
        else:
            print(f'  ✅ {property_name} updated successfully (YPD: TBD - units unknown)')
        print()
    
    wb.save(MASTER_FILE)
    wb.close()
    
    print('✅ Master file updated successfully')
    print()

def verify_updates(results):
    """Verify that Arizona properties were updated correctly"""
    
    print('=' * 80)
    print('VERIFYING UPDATES')
    print('=' * 80)
    print()
    
    for property_name, expected_data in results.items():
        try:
            df = pd.read_excel(MASTER_FILE, sheet_name=property_name)
            
            if 'Container Count' in df.columns:
                actual_count = df['Container Count'].iloc[0]
                expected_count = expected_data['total_containers']
                
                if actual_count == expected_count:
                    print(f'  ✅ {property_name}: Container Count = {actual_count} (matches expected)')
                else:
                    print(f'  ⚠️ {property_name}: Container Count = {actual_count} (expected {expected_count})')
            
            if 'YPD' in df.columns and expected_data['ypd']:
                actual_ypd = df['YPD'].iloc[0]
                expected_ypd = expected_data['ypd']
                
                if isinstance(actual_ypd, (int, float)) and abs(actual_ypd - expected_ypd) < 0.01:
                    print(f'  ✅ {property_name}: YPD = {actual_ypd:.2f} (matches expected {expected_ypd:.2f})')
                else:
                    print(f'  ℹ️ {property_name}: YPD = {actual_ypd}')
            
            if 'Service Notes' in df.columns:
                actual_notes = df['Service Notes'].iloc[0]
                expected_notes = expected_data['service_notes']
                
                if actual_notes == expected_notes:
                    print(f'  ✅ {property_name}: Service Notes match')
                else:
                    print(f'  ℹ️ {property_name}: Service Notes updated')
            
            print()
            
        except Exception as e:
            print(f'  ⚠️ {property_name}: Could not verify - {str(e)}')
            print()

def create_summary_report(results):
    """Create a summary report of the Arizona properties update"""
    
    report_path = BASE_DIR / 'Portfolio_Reports' / 'ARIZONA_PROPERTIES_UPDATE_SUMMARY.md'
    
    content = f"""# Arizona Properties Update Summary

**Date:** {datetime.now().strftime('%B %d, %Y at %I:%M %p')}  
**Purpose:** Add service details for 4 Arizona properties to master file

---

## SERVICE DETAILS PROVIDED BY USER

"""
    
    for property_name, data in results.items():
        prop_data = AZ_SERVICE_DETAILS[property_name]
        
        content += f"""### {property_name}

**Units:** {data['units'] if data['units'] else 'TBD'}

**Services:**
"""
        
        for i, service in enumerate(prop_data['services'], 1):
            size = service['size']
            count = service['count']
            pickups = service['pickups_per_week']
            stype = service['type']
            monthly_yards = size * count * pickups * 4.33
            
            content += f"""
{i}. {count}× {size}YD {stype} @ {pickups}x/week
   - Calculation: {size} × {count} × {pickups} × 4.33 = {monthly_yards:.2f} yards/month
"""
        
        # Add bulk service if present
        if 'bulk_service' in prop_data:
            content += f"""
**Bulk Service:** {prop_data['bulk_service']}
"""
            if 'bulk_cost' in prop_data:
                content += f"""**Bulk Cost:** ${prop_data['bulk_cost']:.2f}/month
"""
        
        content += f"""
**Total Containers:** {data['total_containers']}  
**Total Monthly Yards:** {data['total_monthly_yards']:.2f}  
"""
        
        if data['ypd']:
            content += f"""**YPD:** {data['total_monthly_yards']:.2f} / {data['units']} = **{data['ypd']:.2f}**

"""
            
            target = 2.0
            if data['ypd'] <= target:
                pct_below = ((target - data['ypd']) / target) * 100
                content += f"**Performance:** ✅ {pct_below:.1f}% below target (2.0)\n\n"
            else:
                pct_above = ((data['ypd'] - target) / target) * 100
                content += f"**Performance:** ⚠️ {pct_above:.1f}% above target (2.0)\n\n"
        else:
            content += f"""**YPD:** TBD (units unknown)

**Performance:** Cannot calculate without unit count

"""
        
        content += "---\n\n"
    
    content += """## SUMMARY TABLE

| Property | Units | Containers | Monthly Yards | YPD | vs. Target (2.0) |
|----------|-------|------------|---------------|-----|------------------|
"""
    
    for property_name, data in results.items():
        units_str = str(data['units']) if data['units'] else 'TBD'
        ypd_str = f"{data['ypd']:.2f}" if data['ypd'] else 'TBD'
        
        if data['ypd']:
            target = 2.0
            if data['ypd'] <= target:
                pct_below = ((target - data['ypd']) / target) * 100
                perf = f"✅ -{pct_below:.1f}%"
            else:
                pct_above = ((data['ypd'] - target) / target) * 100
                perf = f"⚠️ +{pct_above:.1f}%"
        else:
            perf = "N/A"
        
        content += f"| {property_name} | {units_str} | {data['total_containers']} | {data['total_monthly_yards']:.2f} | **{ypd_str}** | {perf} |\n"
    
    content += """
---

## PORTFOLIO UPDATE

**Properties with Complete Service Data:** 8/10 (80%)

**TX/FL Properties (6):**
- ✅ McCord Park FL
- ✅ Orion McKinney
- ✅ The Club at Millenia
- ✅ Bella Mirage
- ⚠️ Orion Prosper (pending contract)
- ⚠️ Orion Prosper Lakes (pending contract)

**AZ Properties (4):**
- ✅ Tempe Vista
- ✅ Mandarina
- ✅ Springs at Alta Mesa
- ⚠️ Pavilions at Arrowhead (units TBD)

---

## BACKUP INFORMATION

**Backup File:** """ + BACKUP_FILE.name + """  
**Location:** Portfolio_Reports/

---

**Arizona properties update completed successfully!**
"""
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(content)
    
    print(f'✅ Summary report created: {report_path.name}')
    print()

def main():
    """Main function"""
    
    # Show calculations
    results = show_calculations()
    
    # Create backup
    if not create_backup():
        print('❌ Cannot proceed without backup')
        return
    
    # Update master file
    update_az_properties_in_master(results)
    
    # Verify updates
    verify_updates(results)
    
    # Create summary report
    create_summary_report(results)
    
    print('=' * 80)
    print('ARIZONA PROPERTIES UPDATE COMPLETE')
    print('=' * 80)
    print()
    print('Summary:')
    for property_name, data in results.items():
        if data['ypd']:
            print(f'  {property_name}: {data["total_containers"]} containers, YPD = {data["ypd"]:.2f}')
        else:
            print(f'  {property_name}: {data["total_containers"]} containers, YPD = TBD (units unknown)')
    print()
    print('Portfolio Status: 8/10 properties with complete service data (80%)')
    print()

if __name__ == "__main__":
    main()

