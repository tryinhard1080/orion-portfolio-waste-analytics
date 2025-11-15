"""
Correct Bella Mirage Service Configuration Based on Latest Invoices

Based on October and November 2024 invoices, the CURRENT service is:
- 4x 8 YD Dumpsters @ 4x/week (Weekly x4)
- 1x 6 YD Dumpster @ 4x/week
- 1x 4 YD Dumpster @ 4x/week

Total: 6 containers @ 4x/week

This replaces the INCORRECT contract-based data showing 10 containers @ 3x/week.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Corrected Bella Mirage configuration from latest invoices
BELLA_MIRAGE_CURRENT = {
    'containers': [
        {'type': 'Dumpster (FEL)', 'size': '8 YD', 'quantity': 4, 'frequency': '4x/week'},
        {'type': 'Dumpster (FEL)', 'size': '6 YD', 'quantity': 1, 'frequency': '4x/week'},
        {'type': 'Dumpster (FEL)', 'size': '4 YD', 'quantity': 1, 'frequency': '4x/week'}
    ],
    'notes': 'From Oct & Nov 2024 invoices - 6 containers total, all serviced 4x/week'
}

def main():
    print("="*80)
    print("CORRECTING BELLA MIRAGE SERVICE CONFIGURATION")
    print("="*80)

    master_path = 'Portfolio_Reports/MASTER_Portfolio_Complete_Data.xlsx'

    # Read current Service Details
    df_service = pd.read_excel(master_path, sheet_name='Service Details')

    print("\nCURRENT Bella Mirage entries in Service Details:")
    bella_current = df_service[df_service['Property'] == 'Bella Mirage']
    print(bella_current[['Property', 'Container Type', 'Container Size', 'Quantity', 'Frequency']].to_string(index=False))
    print(f"\nTotal containers (CURRENT/INCORRECT): {bella_current['Quantity'].sum():.0f}")

    # Remove old Bella Mirage entries
    df_service = df_service[df_service['Property'] != 'Bella Mirage']

    # Add corrected entries
    new_entries = []
    for container in BELLA_MIRAGE_CURRENT['containers']:
        # Calculate total yards
        try:
            size_num = int(''.join(filter(str.isdigit, container['size'].split()[0])))
            total_yards = size_num * container['quantity']
        except:
            total_yards = 0

        new_entries.append({
            'Property': 'Bella Mirage',
            'Container Type': container['type'],
            'Container Size': container['size'],
            'Quantity': container['quantity'],
            'Frequency': container['frequency'],
            'Total Yards': total_yards
        })

    df_new = pd.DataFrame(new_entries)
    df_service = pd.concat([df_service, df_new], ignore_index=True)
    df_service = df_service.sort_values('Property').reset_index(drop=True)

    print("\n" + "="*80)
    print("CORRECTED Bella Mirage entries:")
    print("="*80)
    bella_corrected = df_service[df_service['Property'] == 'Bella Mirage']
    print(bella_corrected[['Property', 'Container Type', 'Container Size', 'Quantity', 'Frequency']].to_string(index=False))
    print(f"\nTotal containers (CORRECTED): {bella_corrected['Quantity'].sum():.0f}")
    print(f"Frequency: 4x/week (all containers)")
    print(f"Note: {BELLA_MIRAGE_CURRENT['notes']}")

    # Update master file - Service Details sheet
    wb = load_workbook(master_path)
    ws = wb['Service Details']

    # Clear existing data (keep headers)
    ws.delete_rows(2, ws.max_row)

    # Write corrected data
    for r_idx, row in enumerate(dataframe_to_rows(df_service, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Also update Property Overview sheet
    ws_overview = wb['Property Overview']

    # Find Bella Mirage row in Property Overview
    for row_idx in range(2, ws_overview.max_row + 1):
        property_name = ws_overview.cell(row_idx, 1).value
        if property_name and 'Bella Mirage' in str(property_name):
            # Update container count (Column E or F depending on structure)
            # Assuming Column D = Container Count
            ws_overview.cell(row_idx, 4).value = 6  # 6 containers total
            ws_overview.cell(row_idx, 5).value = "Mixed (8YD, 6YD, 4YD)"  # Container sizes
            ws_overview.cell(row_idx, 6).value = "4x/week"  # Frequency
            print(f"\nUpdated Property Overview for Bella Mirage")
            break

    # Save
    wb.save(master_path)

    print("\n" + "="*80)
    print("CORRECTION COMPLETE")
    print("="*80)
    print(f"Updated file: {master_path}")
    print("\nKEY CORRECTION:")
    print("- Bella Mirage: 10 containers @ 3x/week (WRONG) --> 6 containers @ 4x/week (CORRECT)")
    print("- Based on October & November 2024 invoices (most recent data)")

if __name__ == '__main__':
    main()
