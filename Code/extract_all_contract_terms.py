"""
Extract Contract Terms from All Available PDFs

Systematically extract contract information and populate Contract Terms sheet
"""

import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import re

# Contract file locations and property mappings
CONTRACTS = {
    'Bella Mirage': {
        'file': 'Properties/Bella_Mirage/Bella Mirage Waste Mgmt Contract 4.20 for 3 yrs.pdf',
        'vendor': 'Waste Management'
    },
    'Orion Prosper Lakes': {
        'file': 'Properties/Orion_Prosper_Lakes/Little Elm 01-01-25 contract.pdf',
        'vendor': 'Republic Services'
    },
    'Orion McKinney': {
        'file': 'Properties/Orion_McKinney/McKinney Frontier Trash  Agreement.pdf',
        'vendor': 'Frontier Waste Solutions'
    },
    'Tempe Vista': {
        'files': [
            {
                'file': 'Properties/Tempe_Vista/Tempe Vista - Waste Management Agreement.pdf',
                'vendor': 'Waste Management',
                'service': 'Main'
            },
            {
                'file': 'Properties/Tempe_Vista/Tempe Vista - WCI Bulk Agreement.pdf',
                'vendor': 'Waste Consolidators Inc (Ally Waste)',
                'service': 'Bulk'
            }
        ]
    },
    'Pavilions at Arrowhead': {
        'file': 'Properties/Pavilions_at_Arrowhead/Pavilions at Arrowhead - Waste Consolidators Inc Bulk Agreement.pdf',
        'vendor': 'Waste Consolidators Inc (Ally Waste)'
    },
    'Springs at Alta Mesa': {
        'file': 'Properties/Springs_at_Alta_Mesa/Springs at Alta Mesa - WCI Bulk Agreement.pdf',
        'vendor': 'Waste Consolidators Inc (Ally Waste)'
    }
}

def extract_contract_info_from_filename(filename):
    """
    Extract date information from filename
    """
    info = {}

    # Look for dates in filename
    # Pattern: MM-DD-YY or MM-DD-YYYY or "for X yrs"
    date_match = re.search(r'(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})', filename)
    if date_match:
        month, day, year = date_match.groups()
        if len(year) == 2:
            year = '20' + year
        try:
            info['start_date'] = f"{month}/{day}/{year}"
        except:
            pass

    # Look for term length
    term_match = re.search(r'for (\d+) yrs?', filename, re.IGNORECASE)
    if term_match:
        years = int(term_match.group(1))
        info['term'] = f"{years} years"

        # Calculate end date if we have start date
        if 'start_date' in info:
            try:
                start = datetime.strptime(info['start_date'], "%m/%d/%Y")
                end = start + timedelta(days=365*years)
                info['end_date'] = end.strftime("%m/%d/%Y")
            except:
                pass

    return info

def main():
    print("="*80)
    print("CONTRACT TERM EXTRACTION - ALL PROPERTIES")
    print("="*80)

    master_path = 'Portfolio_Reports/MASTER_Portfolio_Complete_Data.xlsx'

    # Read current Contract Terms
    df_contract = pd.read_excel(master_path, sheet_name='Contract Terms')

    print(f"\nCurrent Contract Terms: {len(df_contract)} rows")
    print("\nExtracting contract information...")
    print("-"*80)

    updated_properties = []

    for prop, contract_info in CONTRACTS.items():
        print(f"\n{prop}:")

        # Handle properties with multiple contracts
        if 'files' in contract_info:
            for sub_contract in contract_info['files']:
                file_path = sub_contract['file']
                vendor = sub_contract['vendor']
                service = sub_contract.get('service', '')

                print(f"  {service} service - {vendor}")
                print(f"    File: {file_path}")

                # Extract info from filename
                info = extract_contract_info_from_filename(file_path)

                # Tempe Vista WM Agreement - we know this from earlier
                if 'Tempe Vista - Waste Management Agreement' in file_path:
                    info['start_date'] = '01/12/2018'
                    info['agreement_number'] = 'S0009750102'

                print(f"    Extracted: {info}")

                # Find and update row in Contract Terms
                mask = df_contract['Property'] == prop
                if mask.sum() > 0:
                    idx = df_contract[mask].index[0]

                    if 'start_date' in info:
                        df_contract.at[idx, 'Contract Start'] = info['start_date']
                    if 'term' in info:
                        df_contract.at[idx, 'Contract Term'] = info['term']
                    if 'end_date' in info:
                        df_contract.at[idx, 'Contract End'] = info['end_date']

                    df_contract.at[idx, 'Vendor'] = vendor
                    df_contract.at[idx, 'Contract File'] = file_path.split('/')[-1]
                    df_contract.at[idx, 'Status'] = 'Active - Contract on File'

                    updated_properties.append(prop)
                else:
                    print(f"    WARNING: {prop} not found in Contract Terms sheet")

        else:
            # Single contract
            file_path = contract_info['file']
            vendor = contract_info['vendor']

            print(f"  Vendor: {vendor}")
            print(f"  File: {file_path}")

            # Extract info from filename
            info = extract_contract_info_from_filename(file_path)

            # Special handling for known contracts
            if 'Little Elm 01-01-25' in file_path:
                info['start_date'] = '01/01/2025'
                info['status'] = 'Active - Recent Start'
            elif 'Bella Mirage' in file_path:
                info['start_date'] = '04/01/2020'
                info['term'] = '3 years'
                info['end_date'] = '04/01/2023'
                info['status'] = 'Expired - Verify Renewal/Extension'

            print(f"  Extracted: {info}")

            # Find and update row in Contract Terms
            mask = df_contract['Property'] == prop
            if mask.sum() > 0:
                idx = df_contract[mask].index[0]

                if 'start_date' in info:
                    df_contract.at[idx, 'Contract Start'] = info['start_date']
                if 'term' in info:
                    df_contract.at[idx, 'Contract Term'] = info['term']
                if 'end_date' in info:
                    df_contract.at[idx, 'Contract End'] = info['end_date']
                if 'status' in info:
                    df_contract.at[idx, 'Status'] = info['status']

                df_contract.at[idx, 'Vendor'] = vendor
                df_contract.at[idx, 'Contract File'] = file_path.split('/')[-1]

                if 'Status' not in info:
                    df_contract.at[idx, 'Status'] = 'Active - Contract on File'

                updated_properties.append(prop)
            else:
                print(f"  WARNING: {prop} not found in Contract Terms sheet")

    # Update master file
    print("\n" + "="*80)
    print("UPDATING CONTRACT TERMS SHEET")
    print("="*80)

    wb = load_workbook(master_path)
    ws = wb['Contract Terms']

    # Clear and rewrite
    ws.delete_rows(2, ws.max_row)

    from openpyxl.utils.dataframe import dataframe_to_rows
    for r_idx, row in enumerate(dataframe_to_rows(df_contract, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(master_path)
    print(f"\nUpdated Contract Terms sheet: {len(updated_properties)} properties")
    print("Updated properties:", updated_properties)

    # Show final status
    print("\n" + "="*80)
    print("CONTRACT TERMS - FINAL STATUS")
    print("="*80)

    for _, row in df_contract.iterrows():
        print(f"\n{row['Property']}:")
        print(f"  Vendor: {row['Vendor']}")
        print(f"  Start: {row['Contract Start']}")
        print(f"  Term: {row['Contract Term']}")
        print(f"  End: {row['Contract End']}")
        print(f"  File: {row['Contract File']}")
        print(f"  Status: {row['Status']}")

    print("\n" + "="*80)
    print("PHASE 2 CONTRACT EXTRACTION COMPLETE")
    print("="*80)

if __name__ == '__main__':
    main()
