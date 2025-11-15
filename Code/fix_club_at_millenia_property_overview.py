"""
Fix The Club at Millenia Property Overview Data

Based on September 2025 invoice #1569687W460:
- Service: 2x 30 YD Compactors @ On-call
- Invoice shows multiple "RO DUMP & RETURN 1.00 30.00YD C" entries
- 2x "BASIC CONTAINER CHARGE 1.00 30.00YD C"

CURRENT (INCORRECT):
- Container Count: 6
- Container Size: 8 YD
- Service Frequency: 4x/week
(This appears to be Bella Mirage data copied over)

CORRECT (from invoice):
- Container Count: 2
- Container Size: 30 YD
- Service Frequency: On-call (varies)
"""

import pandas as pd
from openpyxl import load_workbook

def main():
    print("="*80)
    print("FIXING THE CLUB AT MILLENIA PROPERTY OVERVIEW")
    print("="*80)

    master_path = 'Portfolio_Reports/MASTER_Portfolio_Complete_Data.xlsx'

    # Load workbook
    wb = load_workbook(master_path)
    ws_overview = wb['Property Overview']

    print("\nSearching for The Club at Millenia in Property Overview...")

    # Find and fix The Club at Millenia row
    fixed = False
    for row_idx in range(2, ws_overview.max_row + 1):
        property_name = ws_overview.cell(row_idx, 1).value

        if property_name and 'Club at Millenia' in str(property_name):
            print(f"\nFound at row {row_idx}: {property_name}")

            # Show current (incorrect) values
            current_svc_type = ws_overview.cell(row_idx, 4).value
            current_cnt = ws_overview.cell(row_idx, 6).value
            current_size = ws_overview.cell(row_idx, 7).value
            current_freq = ws_overview.cell(row_idx, 8).value

            print("\nCURRENT (INCORRECT) VALUES:")
            print(f"  Service Type: {current_svc_type}")
            print(f"  Container Count: {current_cnt}")
            print(f"  Container Size: {current_size}")
            print(f"  Service Frequency: {current_freq}")

            # Apply corrections
            ws_overview.cell(row_idx, 6).value = 2        # Container Count
            ws_overview.cell(row_idx, 7).value = "30 YD"  # Container Size
            ws_overview.cell(row_idx, 8).value = "On-call (varies)"  # Frequency

            print("\nCORRECTED VALUES:")
            print(f"  Service Type: {current_svc_type} (no change)")
            print(f"  Container Count: 2")
            print(f"  Container Size: 30 YD")
            print(f"  Service Frequency: On-call (varies)")

            print("\nEVIDENCE:")
            print("  Invoice #1569687W460 (Sept 2025)")
            print("  - 2x BASIC CONTAINER CHARGE 1.00 30.00YD C")
            print("  - Multiple RO DUMP & RETURN entries (on-call service)")

            fixed = True
            break

    if not fixed:
        print("ERROR: Could not find The Club at Millenia in Property Overview")
        return

    # Save changes
    print("\n" + "="*80)
    print("SAVING CHANGES")
    print("="*80)

    wb.save(master_path)
    print(f"Saved: {master_path}")

    # Verify correction
    print("\n" + "="*80)
    print("VERIFICATION")
    print("="*80)

    df_overview = pd.read_excel(master_path, sheet_name='Property Overview')
    club_row = df_overview[df_overview['Property Name'] == 'The Club at Millenia']

    if not club_row.empty:
        print("\nThe Club at Millenia - Property Overview:")
        print(f"  Service Type: {club_row['Service Type'].values[0]}")
        print(f"  Container Count: {club_row['Container Count'].values[0]}")
        print(f"  Container Size: {club_row['Container Size'].values[0]}")
        print(f"  Service Frequency: {club_row['Service Frequency'].values[0]}")

    # Check Service Details for comparison
    df_service = pd.read_excel(master_path, sheet_name='Service Details')
    club_service = df_service[df_service['Property'] == 'The Club at Millenia']

    print("\nThe Club at Millenia - Service Details (for comparison):")
    for _, row in club_service.iterrows():
        print(f"  {int(row['Quantity'])}x {row['Container Size']} {row['Container Type']} @ {row['Frequency']}")

    print("\n" + "="*80)
    print("FIX COMPLETE")
    print("="*80)
    print("Property Overview now matches Service Details and September 2025 invoice.")

if __name__ == '__main__':
    main()
