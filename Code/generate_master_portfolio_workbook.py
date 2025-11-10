"""
Generate Master Portfolio Summary Workbook
Combines all 10 properties with WasteWise Analytics and Regulatory Compliance data
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime


# Property configuration (same as generate_property_workbooks_with_regulatory.py)
PROPERTIES = {
    'Orion Prosper': {
        'units': 312,
        'location': 'Prosper, TX',
        'state': 'TX',
        'vendor': 'Republic Services',
        'service_type': 'FEL Dumpsters',
        'tab_name': 'Orion Prosper',
        'confidence': 'LOW',
        'recycling_status': 'VERIFICATION REQUIRED',
        'contact': 'Town of Prosper: 945-234-1924'
    },
    'Orion Prosper Lakes': {
        'units': 308,
        'location': 'Prosper, TX',
        'state': 'TX',
        'vendor': 'Republic Services',
        'service_type': 'Compactor',
        'tab_name': 'Orion Prosper Lakes',
        'confidence': 'LOW',
        'recycling_status': 'VERIFICATION REQUIRED',
        'contact': 'Town of Prosper: 945-234-1924'
    },
    'Orion McKinney': {
        'units': 453,
        'location': 'McKinney, TX',
        'state': 'TX',
        'vendor': 'Frontier Waste',
        'service_type': 'FEL Dumpsters',
        'tab_name': 'Orion McKinney',
        'confidence': 'LOW',
        'recycling_status': 'VERIFICATION REQUIRED',
        'contact': 'McKinney Solid Waste: 972-547-7385'
    },
    'McCord Park FL': {
        'units': 416,
        'location': 'Florida',
        'state': 'FL',
        'vendor': 'Community Waste',
        'service_type': 'Dumpster',
        'tab_name': 'McCord Park FL',
        'confidence': 'PENDING',
        'recycling_status': 'RESEARCH NEEDED',
        'contact': 'TBD'
    },
    'The Club at Millenia': {
        'units': 560,
        'location': 'Orlando, FL',
        'state': 'FL',
        'vendor': 'Waste Connections',
        'service_type': 'Compactor',
        'tab_name': 'The Club at Millenia',
        'confidence': 'HIGH',
        'recycling_status': 'MANDATORY (City ordinance - April 2019)',
        'contact': 'Orlando Solid Waste'
    },
    'Bella Mirage': {
        'units': 715,
        'location': 'Phoenix, AZ',
        'state': 'AZ',
        'vendor': 'Waste Management',
        'service_type': 'Dumpster',
        'tab_name': 'Bella Mirage',
        'confidence': 'MEDIUM',
        'recycling_status': 'VOLUNTARY ONLY (State law prohibits mandates)',
        'contact': 'Phoenix Public Works: 602-262-6251'
    },
    'Mandarina': {
        'units': 180,
        'location': 'Phoenix, AZ',
        'state': 'AZ',
        'vendor': 'WM + Ally Waste',
        'service_type': 'Compactor + Bulk',
        'tab_name': 'Mandarina',
        'confidence': 'MEDIUM',
        'recycling_status': 'VOLUNTARY ONLY (State law prohibits mandates)',
        'contact': 'Phoenix Public Works: 602-262-6251'
    },
    'Pavilions at Arrowhead': {
        'units': None,
        'location': 'Glendale, AZ',
        'state': 'AZ',
        'vendor': 'City + Ally Waste',
        'service_type': 'Mixed',
        'tab_name': 'Pavilions at Arrowhead',
        'confidence': 'MEDIUM',
        'recycling_status': 'VOLUNTARY (Program available)',
        'contact': 'Glendale Solid Waste'
    },
    'Springs at Alta Mesa': {
        'units': 200,
        'location': 'Mesa, AZ',
        'state': 'AZ',
        'vendor': 'City + Ally Waste',
        'service_type': 'Dumpster + Bulk',
        'tab_name': 'Springs at Alta Mesa',
        'confidence': 'MEDIUM',
        'recycling_status': 'VOLUNTARY (Multi-unit program available)',
        'contact': 'Mesa Solid Waste'
    },
    'Tempe Vista': {
        'units': 150,
        'location': 'Tempe, AZ',
        'state': 'AZ',
        'vendor': 'WM + Ally Waste',
        'service_type': 'Mixed',
        'tab_name': 'Tempe Vista',
        'confidence': 'MEDIUM',
        'recycling_status': 'VOLUNTARY (Multi-family program available)',
        'contact': 'Tempe Solid Waste'
    }
}


def create_portfolio_overview_sheet(wb, master_file):
    """Create PORTFOLIO_OVERVIEW sheet with all properties"""

    ws = wb.create_sheet("PORTFOLIO_OVERVIEW", 0)

    # Title
    ws['A1'] = "ORION PORTFOLIO - WASTE MANAGEMENT ANALYTICS"
    ws['A1'].font = Font(bold=True, size=14, color="1E3A8A")
    ws.merge_cells('A1:H1')

    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:H2')

    row = 4

    # Portfolio summary metrics
    total_units = sum(p['units'] for p in PROPERTIES.values() if p['units'])

    ws[f'A{row}'] = "PORTFOLIO SUMMARY"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws.merge_cells(f'A{row}:B{row}')
    row += 1

    ws[f'A{row}'] = "Total Properties:"
    ws[f'B{row}'] = len(PROPERTIES)
    row += 1

    ws[f'A{row}'] = "Total Units:"
    ws[f'B{row}'] = total_units
    row += 1

    ws[f'A{row}'] = "States:"
    ws[f'B{row}'] = "TX (3), FL (2), AZ (5)"
    row += 2

    # Load invoice data for each property
    total_spend = 0
    invoice_counts = {}

    for property_name, property_info in PROPERTIES.items():
        try:
            tab_name = property_info.get('tab_name', property_name)
            invoice_data = pd.read_excel(master_file, sheet_name=tab_name)

            # Determine amount column
            if 'Extended Amount' in invoice_data.columns:
                amount_col = 'Extended Amount'
            elif 'Line Item Amount' in invoice_data.columns:
                amount_col = 'Line Item Amount'
            else:
                amount_col = 'Invoice Amount'

            property_total = invoice_data[amount_col].sum()
            total_spend += property_total
            invoice_counts[property_name] = len(invoice_data)
        except Exception as e:
            print(f"[WARNING] Could not load data for {property_name}: {e}")
            invoice_counts[property_name] = 0

    ws[f'A{row}'] = "Total Annual Spend:"
    ws[f'B{row}'] = f"${total_spend:,.2f}"
    ws[f'B{row}'].font = Font(bold=True)
    row += 1

    ws[f'A{row}'] = "Average Cost Per Unit:"
    avg_cost_per_unit = total_spend / total_units if total_units > 0 else 0
    ws[f'B{row}'] = f"${avg_cost_per_unit:,.2f}"
    row += 2

    # Property details header
    ws[f'A{row}'] = "PROPERTY DETAILS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws.merge_cells(f'A{row}:H{row}')
    row += 1

    # Column headers
    headers = ['Property', 'Location', 'Units', 'Vendor', 'Service Type', 'Invoice Count', 'Workbook Path']
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    row += 1

    # Property data
    for property_name, property_info in PROPERTIES.items():
        ws.cell(row=row, column=1, value=property_name)
        ws.cell(row=row, column=2, value=property_info['location'])
        ws.cell(row=row, column=3, value=property_info['units'] if property_info['units'] else 'TBD')
        ws.cell(row=row, column=4, value=property_info['vendor'])
        ws.cell(row=row, column=5, value=property_info['service_type'])
        ws.cell(row=row, column=6, value=invoice_counts.get(property_name, 0))

        # Workbook path
        folder_name = property_name.replace(' ', '_')
        workbook_path = f"Properties/{folder_name}/{folder_name}_WasteAnalysis_Validated.xlsx"
        ws.cell(row=row, column=7, value=workbook_path)

        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 60


def create_regulatory_summary_sheet(wb):
    """Create REGULATORY_COMPLIANCE_SUMMARY sheet"""

    ws = wb.create_sheet("REGULATORY_SUMMARY")

    # Title
    ws['A1'] = "REGULATORY COMPLIANCE SUMMARY - ALL PROPERTIES"
    ws['A1'].font = Font(bold=True, size=14, color="1E3A8A")
    ws.merge_cells('A1:F1')

    ws['A2'] = f"Research Date: {datetime.now().strftime('%Y-%m-%d')}"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:F2')

    row = 4

    # Confidence distribution
    ws[f'A{row}'] = "RESEARCH CONFIDENCE DISTRIBUTION"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws.merge_cells(f'A{row}:B{row}')
    row += 1

    confidence_counts = {}
    for property_info in PROPERTIES.values():
        conf = property_info['confidence']
        confidence_counts[conf] = confidence_counts.get(conf, 0) + 1

    confidence_colors = {
        'HIGH': PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        'MEDIUM': PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        'LOW': PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        'PENDING': PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    }

    for confidence, count in confidence_counts.items():
        ws[f'A{row}'] = confidence
        ws[f'A{row}'].fill = confidence_colors.get(confidence, confidence_colors['PENDING'])
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f"{count} properties"
        row += 1

    row += 1

    # Property regulatory details
    ws[f'A{row}'] = "PROPERTY REGULATORY STATUS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    # Headers
    headers = ['Property', 'Location', 'State', 'Confidence', 'Recycling Status', 'Contact']
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    row += 1

    # Property data
    for property_name, property_info in PROPERTIES.items():
        ws.cell(row=row, column=1, value=property_name)
        ws.cell(row=row, column=2, value=property_info['location'])
        ws.cell(row=row, column=3, value=property_info['state'])

        # Confidence with color
        conf_cell = ws.cell(row=row, column=4, value=property_info['confidence'])
        conf_cell.fill = confidence_colors.get(property_info['confidence'], confidence_colors['PENDING'])
        conf_cell.font = Font(bold=True)

        ws.cell(row=row, column=5, value=property_info['recycling_status'])
        ws.cell(row=row, column=6, value=property_info['contact'])

        row += 1

    row += 2

    # Key findings section
    ws[f'A{row}'] = "KEY REGULATORY FINDINGS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    ws[f'A{row}'] = "Texas Properties (3):"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = "- LOW confidence: Ordinances not publicly accessible"
    row += 1
    ws[f'A{row}'] = "- Manual verification required (contact information provided)"
    row += 2

    ws[f'A{row}'] = "Orlando Property (1):"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = "- HIGH confidence: MANDATORY recycling confirmed (April 2019 ordinance)"
    row += 1
    ws[f'A{row}'] = "- Compliance verification required"
    row += 2

    ws[f'A{row}'] = "Phoenix Properties (2):"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = "- MEDIUM confidence: VOLUNTARY only (state law prohibits mandates)"
    row += 1
    ws[f'A{row}'] = "- No recycling compliance required"
    row += 2

    ws[f'A{row}'] = "Other Arizona Properties (3):"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = "- MEDIUM confidence: Voluntary programs available"
    row += 1
    ws[f'A{row}'] = "- No mandatory requirements identified"

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['F'].width = 35


def create_spend_summary_sheet(wb, master_file):
    """Create SPEND_SUMMARY sheet with spend breakdown"""

    ws = wb.create_sheet("SPEND_SUMMARY")

    # Title
    ws['A1'] = "PORTFOLIO SPEND SUMMARY"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    row = 3

    # Headers
    headers = ['Property', 'Location', 'Invoice Count', 'Total Spend', 'Avg Cost/Unit']
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    row += 1

    # Calculate spend for each property
    total_portfolio_spend = 0

    for property_name, property_info in PROPERTIES.items():
        try:
            tab_name = property_info.get('tab_name', property_name)
            invoice_data = pd.read_excel(master_file, sheet_name=tab_name)

            # Determine amount column
            if 'Extended Amount' in invoice_data.columns:
                amount_col = 'Extended Amount'
            elif 'Line Item Amount' in invoice_data.columns:
                amount_col = 'Line Item Amount'
            else:
                amount_col = 'Invoice Amount'

            property_total = invoice_data[amount_col].sum()
            invoice_count = len(invoice_data)

            ws.cell(row=row, column=1, value=property_name)
            ws.cell(row=row, column=2, value=property_info['location'])
            ws.cell(row=row, column=3, value=invoice_count)
            ws.cell(row=row, column=4, value=f"${property_total:,.2f}")

            if property_info['units']:
                cost_per_unit = property_total / property_info['units']
                ws.cell(row=row, column=5, value=f"${cost_per_unit:,.2f}")
            else:
                ws.cell(row=row, column=5, value="N/A")

            total_portfolio_spend += property_total

        except Exception as e:
            print(f"[WARNING] Could not load spend data for {property_name}: {e}")
            ws.cell(row=row, column=1, value=property_name)
            ws.cell(row=row, column=2, value=property_info['location'])
            ws.cell(row=row, column=3, value=0)
            ws.cell(row=row, column=4, value="$0.00")
            ws.cell(row=row, column=5, value="N/A")

        row += 1

    # Total row
    ws.cell(row=row, column=1, value="PORTFOLIO TOTAL")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=4, value=f"${total_portfolio_spend:,.2f}")
    ws.cell(row=row, column=4).font = Font(bold=True)

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18


def generate_master_workbook():
    """Generate master portfolio summary workbook"""

    print("="*60)
    print("MASTER PORTFOLIO SUMMARY WORKBOOK GENERATION")
    print("="*60)
    print()

    # Check for master file
    master_file = Path("Portfolio_Reports/MASTER_Portfolio_Complete_Data.xlsx")

    if not master_file.exists():
        print(f"[ERROR] Master file not found: {master_file}")
        return False

    print(f"[OK] Loading data from: {master_file}")

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create sheets
    print("Creating PORTFOLIO_OVERVIEW sheet...")
    create_portfolio_overview_sheet(wb, master_file)

    print("Creating REGULATORY_SUMMARY sheet...")
    create_regulatory_summary_sheet(wb)

    print("Creating SPEND_SUMMARY sheet...")
    create_spend_summary_sheet(wb, master_file)

    # Save workbook
    output_path = Path("Portfolio_Reports/MASTER_Portfolio_Summary_with_Regulatory.xlsx")
    wb.save(output_path)
    print(f"[OK] Master workbook saved: {output_path}")

    return True


if __name__ == '__main__':
    import sys

    success = generate_master_workbook()

    print()
    print("="*60)
    if success:
        print("[OK] Master portfolio summary workbook generated successfully")
        print("="*60)
        sys.exit(0)
    else:
        print("[ERROR] Failed to generate master workbook")
        print("="*60)
        sys.exit(1)
