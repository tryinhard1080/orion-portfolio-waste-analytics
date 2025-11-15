"""
Update Tempe Vista Service Configuration from WM Agreement

Based on Waste Management Agreement S0009750102
Effective Date: 1/12/2018
Salesperson: Brittney Sappington

Service Configuration from Agreement:
- 1x 4 YD FEL (Recycling) @ 1x/week
- 3x 3 YD FEL (MSW Commercial) @ 3x/week
- 5x 4 YD FEL (MSW Commercial) @ 3x/week

Total: 9 containers (not 8)
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def main():
    print("="*80)
    print("UPDATING TEMPE VISTA SERVICE CONFIGURATION")
    print("="*80)

    master_path = 'Portfolio_Reports/MASTER_Portfolio_Complete_Data.xlsx'

    # Read current Service Details
    df_service = pd.read_excel(master_path, sheet_name='Service Details')

    print("\nCURRENT Tempe Vista configuration:")
    tempe_current = df_service[df_service['Property'] == 'Tempe Vista']
    print(tempe_current[['Property', 'Container Type', 'Container Size', 'Quantity', 'Frequency']].to_string(index=False))
    print(f"Total containers (CURRENT): {int(tempe_current['Quantity'].sum())}")

    # Remove old Tempe Vista entries
    df_service = df_service[df_service['Property'] != 'Tempe Vista']

    # Add corrected entries from WM Agreement
    tempe_vista_corrected = [
        {'Property': 'Tempe Vista', 'Container Type': 'Dumpster (FEL)', 'Container Size': '4 YD',
         'Quantity': 1, 'Frequency': '1x/week', 'Total Yards': 4},
        {'Property': 'Tempe Vista', 'Container Type': 'Dumpster (FEL)', 'Container Size': '3 YD',
         'Quantity': 3, 'Frequency': '3x/week', 'Total Yards': 9},
        {'Property': 'Tempe Vista', 'Container Type': 'Dumpster (FEL)', 'Container Size': '4 YD',
         'Quantity': 5, 'Frequency': '3x/week', 'Total Yards': 20}
    ]

    df_new = pd.DataFrame(tempe_vista_corrected)
    df_service = pd.concat([df_service, df_new], ignore_index=True)
    df_service = df_service.sort_values('Property').reset_index(drop=True)

    print("\n" + "="*80)
    print("CORRECTED Tempe Vista configuration (from WM Agreement):")
    print("="*80)
    tempe_corrected = df_service[df_service['Property'] == 'Tempe Vista']
    print(tempe_corrected[['Property', 'Container Type', 'Container Size', 'Quantity', 'Frequency']].to_string(index=False))
    print(f"Total containers (CORRECTED): {int(tempe_corrected['Quantity'].sum())}")

    # Update master file
    wb = load_workbook(master_path)
    ws = wb['Service Details']

    # Clear existing data (keep headers)
    ws.delete_rows(2, ws.max_row)

    # Write corrected data
    for r_idx, row in enumerate(dataframe_to_rows(df_service, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Update Property Overview
    ws_overview = wb['Property Overview']
    for row_idx in range(2, ws_overview.max_row + 1):
        property_name = ws_overview.cell(row_idx, 1).value
        if property_name and 'Tempe Vista' in str(property_name):
            ws_overview.cell(row_idx, 4).value = 9  # Container count
            ws_overview.cell(row_idx, 5).value = "Mixed (3YD, 4YD)"  # Container sizes
            ws_overview.cell(row_idx, 6).value = "1x-3x/week"  # Frequency
            print("\nUpdated Property Overview for Tempe Vista")
            break

    # Save
    wb.save(master_path)

    print("\n" + "="*80)
    print("UPDATE COMPLETE")
    print("="*80)
    print(f"File updated: {master_path}")
    print("\nKEY CORRECTION:")
    print("- Tempe Vista: 8 -> 9 containers")
    print("- Configuration: 1x 4YD (1x/wk) + 3x 3YD (3x/wk) + 5x 4YD (3x/wk)")
    print("- Source: WM Agreement S0009750102 (Effective 1/12/2018)")

    # Calculate new portfolio total
    total_containers = df_service['Quantity'].sum()
    print(f"\nNew Portfolio Total Containers: {int(total_containers)}")

if __name__ == '__main__':
    main()
