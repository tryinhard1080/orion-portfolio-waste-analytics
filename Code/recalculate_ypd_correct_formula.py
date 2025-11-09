"""
Recalculate YPD with Correct Formula

This script recalculates YPD using the correct dumpster service formula:
YPD = (Container Size × Number of Containers × Pickups per Week × 4.33) / Number of Units
"""

import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
import shutil

# Paths
BASE_DIR = Path(__file__).parent.parent
MASTER_FILE = BASE_DIR / "Portfolio_Reports" / "MASTER_Portfolio_Complete_Data.xlsx"
BACKUP_FILE = BASE_DIR / "Portfolio_Reports" / f"MASTER_Portfolio_Complete_Data_BACKUP_YPD_FIX_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# Property service details with CORRECT YPD calculations
PROPERTIES = {
    'McCord Park FL': {
        'units': 416,
        'services': [
            {'size': 4, 'count': 1, 'pickups_per_week': 3},   # 1×4YD FL @ 3x/week
            {'size': 8, 'count': 12, 'pickups_per_week': 3},  # 12×8YD FL @ 3x/week
            {'size': 8, 'count': 2, 'pickups_per_week': 2},   # 2×8YD SS @ 2x/week (recycling)
        ]
    },
    'Orion McKinney': {
        'units': 453,
        'services': [
            {'size': 8, 'count': 8, 'pickups_per_week': 3},   # 8×8YD FL @ 3x/week
            {'size': 10, 'count': 2, 'pickups_per_week': 3},  # 2×10YD FL @ 3x/week
        ]
    },
    'The Club at Millenia': {
        'units': 560,
        'services': [
            {'size': 8, 'count': 4, 'pickups_per_week': 4},   # 4×8YD Dumpster @ 4x/week
            {'size': 6, 'count': 1, 'pickups_per_week': 4},   # 1×6YD Dumpster @ 4x/week
            {'size': 4, 'count': 1, 'pickups_per_week': 4},   # 1×4YD Dumpster @ 4x/week
        ]
    },
    'Bella Mirage': {
        'units': 715,
        'services': [
            {'size': 8, 'count': 6, 'pickups_per_week': 3},   # 6×8YD FEL @ 3x/week
        ]
    }
}

def calculate_ypd(services, units):
    """
    Calculate YPD using correct formula:
    YPD = (Container Size × Number of Containers × Pickups per Week × 4.33) / Number of Units
    
    For multiple services, sum the yards from each service.
    """
    total_monthly_yards = 0
    
    for service in services:
        size = service['size']
        count = service['count']
        pickups_per_week = service['pickups_per_week']
        
        # Calculate monthly yards for this service
        monthly_yards = size * count * pickups_per_week * 4.33
        total_monthly_yards += monthly_yards
    
    # Calculate YPD
    ypd = total_monthly_yards / units
    
    return ypd, total_monthly_yards

def show_calculations():
    """Show detailed YPD calculations for each property"""
    
    print('=' * 80)
    print('YPD RECALCULATION - CORRECT FORMULA')
    print('=' * 80)
    print()
    print('Formula: YPD = (Size × Count × Pickups/Week × 4.33) / Units')
    print()
    print('=' * 80)
    print()
    
    results = {}
    
    for property_name, data in PROPERTIES.items():
        units = data['units']
        services = data['services']
        
        print(f'{property_name}:')
        print(f'  Units: {units}')
        print()
        
        total_monthly_yards = 0
        
        for i, service in enumerate(services, 1):
            size = service['size']
            count = service['count']
            pickups_per_week = service['pickups_per_week']
            
            monthly_yards = size * count * pickups_per_week * 4.33
            total_monthly_yards += monthly_yards
            
            print(f'  Service {i}:')
            print(f'    {count}× {size}YD containers @ {pickups_per_week}x/week')
            print(f'    Calculation: {size} × {count} × {pickups_per_week} × 4.33 = {monthly_yards:.2f} yards/month')
            print()
        
        ypd = total_monthly_yards / units
        
        print(f'  Total Monthly Yards: {total_monthly_yards:.2f}')
        print(f'  YPD: {total_monthly_yards:.2f} / {units} = {ypd:.2f}')
        print()
        
        # Compare to target
        target = 2.0
        if ypd <= target:
            pct_below = ((target - ypd) / target) * 100
            print(f'  ✅ {pct_below:.1f}% below target (2.0)')
        else:
            pct_above = ((ypd - target) / target) * 100
            print(f'  ⚠️ {pct_above:.1f}% above target (2.0)')
        
        print()
        print('-' * 80)
        print()
        
        results[property_name] = {
            'ypd': ypd,
            'total_monthly_yards': total_monthly_yards,
            'units': units
        }
    
    return results

def create_backup():
    """Create a backup of the master file before updating"""
    
    print('=' * 80)
    print('CREATING BACKUP')
    print('=' * 80)
    print()
    
    if MASTER_FILE.exists():
        shutil.copy2(MASTER_FILE, BACKUP_FILE)
        print(f'✅ Backup created: {BACKUP_FILE.name}')
        print()
        return True
    else:
        print(f'❌ Master file not found: {MASTER_FILE}')
        return False

def update_ypd_in_master(results):
    """Update YPD values in master file"""
    
    print('=' * 80)
    print('UPDATING YPD IN MASTER FILE')
    print('=' * 80)
    print()
    
    wb = openpyxl.load_workbook(MASTER_FILE)
    
    for property_name, data in results.items():
        if property_name not in wb.sheetnames:
            print(f'  ⚠️ Sheet not found: {property_name}')
            continue
        
        ws = wb[property_name]
        
        # Find YPD column (should be in row 1 header)
        ypd_col = None
        total_yards_col = None
        
        for col in range(1, 50):
            header = ws.cell(1, col).value
            if header == 'YPD':
                ypd_col = col
            elif header == 'Total Yards':
                total_yards_col = col
        
        if not ypd_col:
            print(f'  ⚠️ YPD column not found in {property_name}')
            continue
        
        # Update all data rows
        row = 2
        rows_updated = 0
        while ws.cell(row, 1).value:
            ws.cell(row, ypd_col).value = round(data['ypd'], 2)
            if total_yards_col:
                ws.cell(row, total_yards_col).value = round(data['total_monthly_yards'], 2)
            rows_updated += 1
            row += 1
        
        print(f'  ✅ {property_name}: Updated {rows_updated} rows with YPD = {data["ypd"]:.2f}')
    
    wb.save(MASTER_FILE)
    wb.close()
    
    print()
    print('✅ Master file updated successfully')
    print()

def verify_updates(results):
    """Verify that YPD was updated correctly"""
    
    print('=' * 80)
    print('VERIFYING UPDATES')
    print('=' * 80)
    print()
    
    for property_name, expected_data in results.items():
        df = pd.read_excel(MASTER_FILE, sheet_name=property_name)
        
        if 'YPD' in df.columns:
            actual_ypd = df['YPD'].iloc[0]
            expected_ypd = expected_data['ypd']
            
            if abs(actual_ypd - expected_ypd) < 0.01:
                print(f'  ✅ {property_name}: YPD = {actual_ypd:.2f} (matches expected {expected_ypd:.2f})')
            else:
                print(f'  ❌ {property_name}: YPD = {actual_ypd:.2f} (expected {expected_ypd:.2f})')
        else:
            print(f'  ⚠️ {property_name}: YPD column not found')
    
    print()

def create_summary_report(results):
    """Create a summary report of the YPD recalculation"""
    
    report_path = BASE_DIR / 'Portfolio_Reports' / 'YPD_RECALCULATION_SUMMARY.md'
    
    content = f"""# YPD Recalculation Summary

**Date:** {datetime.now().strftime('%B %d, %Y at %I:%M %p')}  
**Purpose:** Recalculate YPD using correct dumpster service formula

---

## CORRECT FORMULA

**For Dumpster/Front Loader Service:**

```
YPD = (Container Size × Number of Containers × Pickups per Week × 4.33) / Number of Units
```

**Components:**
- **Container Size:** Cubic yards of a single container
- **Number of Containers:** Total count of containers at the property
- **Pickups per Week:** Service frequency (how many times per week containers are emptied)
- **4.33:** Weeks per month multiplier (52 weeks ÷ 12 months)
- **Number of Units:** Total residential units at the property

---

## RECALCULATED YPD VALUES

"""
    
    for property_name, data in results.items():
        prop_data = PROPERTIES[property_name]
        
        content += f"""### {property_name}

**Units:** {data['units']}

**Services:**
"""
        
        for i, service in enumerate(prop_data['services'], 1):
            size = service['size']
            count = service['count']
            pickups = service['pickups_per_week']
            monthly_yards = size * count * pickups * 4.33
            
            content += f"""
{i}. {count}× {size}YD containers @ {pickups}x/week
   - Calculation: {size} × {count} × {pickups} × 4.33 = {monthly_yards:.2f} yards/month
"""
        
        content += f"""
**Total Monthly Yards:** {data['total_monthly_yards']:.2f}  
**YPD:** {data['total_monthly_yards']:.2f} / {data['units']} = **{data['ypd']:.2f}**

"""
        
        target = 2.0
        if data['ypd'] <= target:
            pct_below = ((target - data['ypd']) / target) * 100
            content += f"**Performance:** ✅ {pct_below:.1f}% below target (2.0)\n\n"
        else:
            pct_above = ((data['ypd'] - target) / target) * 100
            content += f"**Performance:** ⚠️ {pct_above:.1f}% above target (2.0)\n\n"
        
        content += "---\n\n"
    
    content += """## SUMMARY TABLE

| Property | Units | Monthly Yards | YPD | vs. Target (2.0) | Performance |
|----------|-------|---------------|-----|------------------|-------------|
"""
    
    for property_name, data in results.items():
        target = 2.0
        if data['ypd'] <= target:
            pct_below = ((target - data['ypd']) / target) * 100
            perf = f"✅ -{pct_below:.1f}%"
        else:
            pct_above = ((data['ypd'] - target) / target) * 100
            perf = f"⚠️ +{pct_above:.1f}%"
        
        content += f"| {property_name} | {data['units']} | {data['total_monthly_yards']:.2f} | **{data['ypd']:.2f}** | {perf} | "
        
        if data['ypd'] <= 2.0:
            content += "Excellent |\n"
        elif data['ypd'] <= 2.25:
            content += "Good |\n"
        else:
            content += "Needs Review |\n"
    
    content += """
---

## BACKUP INFORMATION

**Backup File:** """ + BACKUP_FILE.name + """  
**Location:** Portfolio_Reports/

---

**YPD recalculation completed successfully!**
"""
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(content)
    
    print(f'✅ Summary report created: {report_path.name}')
    print()

def main():
    """Main function"""
    
    # Show calculations
    results = show_calculations()
    
    # Create backup
    if not create_backup():
        print('❌ Cannot proceed without backup')
        return
    
    # Update master file
    update_ypd_in_master(results)
    
    # Verify updates
    verify_updates(results)
    
    # Create summary report
    create_summary_report(results)
    
    print('=' * 80)
    print('YPD RECALCULATION COMPLETE')
    print('=' * 80)
    print()
    print('Summary:')
    for property_name, data in results.items():
        print(f'  {property_name}: YPD = {data["ypd"]:.2f}')
    print()

if __name__ == "__main__":
    main()

