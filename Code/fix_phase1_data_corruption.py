"""
Fix Phase 1 Critical Data Corruption Issues

Based on baseline check findings:
1. Orion McKinney: Service Type='10', Container Count='3x/week' (CORRUPTED)
2. Bella Mirage: Service Type='6', Container Count='4x/week' (CORRUPTED)
3. Tempe Vista: Service Type='9', Container Count='1x-3x/week' (CORRUPTED)
4. Property name standardization: "Orion Prosper Lakes (Little Elm)" -> "Orion Prosper Lakes"

CORRECTIONS TO APPLY:
- Orion McKinney: Service Type='Mixed', Container Count=10, Service Frequency='3x/week'
- Bella Mirage: Service Type='Dumpster', Container Count=6, Service Frequency='4x/week'
- Tempe Vista: Service Type='Mixed', Container Count=9, Service Frequency='1x-3x/week'
- Standardize property name in Contract Terms sheet
"""

import pandas as pd
from openpyxl import load_workbook

def main():
    print("="*80)
    print("PHASE 1: FIXING CRITICAL DATA CORRUPTION")
    print("="*80)

    master_path = 'Portfolio_Reports/MASTER_Portfolio_Complete_Data.xlsx'

    # Load workbook
    wb = load_workbook(master_path)
    ws_overview = wb['Property Overview']

    print("\n1. PROPERTY OVERVIEW - FIXING DATA CORRUPTION")
    print("="*80)

    # Track changes
    changes_made = []

    # Fix row by row in Property Overview
    for row_idx in range(2, ws_overview.max_row + 1):
        property_name = ws_overview.cell(row_idx, 1).value

        if not property_name or 'PORTFOLIO' in str(property_name).upper():
            continue

        # Get current values
        current_service_type = ws_overview.cell(row_idx, 4).value
        current_container_count = ws_overview.cell(row_idx, 6).value
        current_frequency = ws_overview.cell(row_idx, 8).value

        # Fix Orion McKinney
        if 'McKinney' in str(property_name):
            if current_service_type == 10 or current_service_type == '10':
                ws_overview.cell(row_idx, 4).value = 'Mixed'  # Service Type
                changes_made.append(f"  {property_name}: Service Type '10' -> 'Mixed'")

            if current_container_count == '3x/week':
                ws_overview.cell(row_idx, 6).value = 10  # Container Count
                changes_made.append(f"  {property_name}: Container Count '3x/week' -> 10")

            # Ensure frequency is correct
            if current_frequency != '3x/week':
                ws_overview.cell(row_idx, 8).value = '3x/week'
                changes_made.append(f"  {property_name}: Service Frequency -> '3x/week'")

        # Fix Bella Mirage
        elif 'Bella Mirage' in str(property_name):
            if current_service_type == 6 or current_service_type == '6':
                ws_overview.cell(row_idx, 4).value = 'Dumpster'  # Service Type
                changes_made.append(f"  {property_name}: Service Type '6' -> 'Dumpster'")

            if current_container_count == '4x/week':
                ws_overview.cell(row_idx, 6).value = 6  # Container Count
                changes_made.append(f"  {property_name}: Container Count '4x/week' -> 6")

            # Ensure frequency is correct
            if current_frequency != '4x/week':
                ws_overview.cell(row_idx, 8).value = '4x/week'
                changes_made.append(f"  {property_name}: Service Frequency -> '4x/week'")

        # Fix Tempe Vista
        elif 'Tempe Vista' in str(property_name):
            if current_service_type == 9 or current_service_type == '9':
                ws_overview.cell(row_idx, 4).value = 'Mixed'  # Service Type
                changes_made.append(f"  {property_name}: Service Type '9' -> 'Mixed'")

            if current_container_count == '1x-3x/week':
                ws_overview.cell(row_idx, 6).value = 9  # Container Count
                changes_made.append(f"  {property_name}: Container Count '1x-3x/week' -> 9")

            # Ensure frequency is correct
            if current_frequency != '1x-3x/week':
                ws_overview.cell(row_idx, 8).value = '1x-3x/week'
                changes_made.append(f"  {property_name}: Service Frequency -> '1x-3x/week'")

    if changes_made:
        print(f"Applied {len(changes_made)} corrections to Property Overview:")
        for change in changes_made:
            print(change)
    else:
        print("No changes needed to Property Overview.")

    # Fix Contract Terms property name
    print("\n2. CONTRACT TERMS - STANDARDIZING PROPERTY NAMES")
    print("="*80)

    ws_contract = wb['Contract Terms']
    contract_changes = []

    for row_idx in range(2, ws_contract.max_row + 1):
        property_name = ws_contract.cell(row_idx, 1).value

        if property_name == 'Orion Prosper Lakes (Little Elm)':
            ws_contract.cell(row_idx, 1).value = 'Orion Prosper Lakes'
            contract_changes.append(f"  Row {row_idx}: 'Orion Prosper Lakes (Little Elm)' -> 'Orion Prosper Lakes'")

    if contract_changes:
        print(f"Applied {len(contract_changes)} corrections to Contract Terms:")
        for change in contract_changes:
            print(change)
    else:
        print("No property name changes needed in Contract Terms.")

    # Save changes
    print("\n" + "="*80)
    print("SAVING CHANGES")
    print("="*80)

    wb.save(master_path)
    print(f"Saved: {master_path}")

    # Verify corrections
    print("\n" + "="*80)
    print("VERIFICATION - READING CORRECTED DATA")
    print("="*80)

    df_overview = pd.read_excel(master_path, sheet_name='Property Overview')

    print("\nCorrected Property Overview (3 affected properties):")
    for prop in ['Orion McKinney', 'Bella Mirage', 'Tempe Vista']:
        prop_row = df_overview[df_overview['Property Name'] == prop]
        if not prop_row.empty:
            svc_type = prop_row['Service Type'].values[0]
            cnt = prop_row['Container Count'].values[0]
            freq = prop_row['Service Frequency'].values[0]
            print(f"  {prop}:")
            print(f"    Service Type: {svc_type}")
            print(f"    Container Count: {cnt}")
            print(f"    Service Frequency: {freq}")

    df_contract = pd.read_excel(master_path, sheet_name='Contract Terms')
    print("\nProperty names in Contract Terms:")
    for prop in sorted(df_contract['Property'].unique()):
        print(f"  - {prop}")

    print("\n" + "="*80)
    print("PHASE 1 DATA CORRUPTION FIXES COMPLETE")
    print("="*80)
    print(f"Total Property Overview corrections: {len(changes_made)}")
    print(f"Total Contract Terms corrections: {len(contract_changes)}")

if __name__ == '__main__':
    main()
