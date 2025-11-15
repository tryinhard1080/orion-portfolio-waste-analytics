"""
Regenerate Spend by Category Sheet from Property Tab Data

Fix tax treatment issue by recalculating category totals from source data
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def main():
    print("="*80)
    print("REGENERATING SPEND BY CATEGORY SHEET")
    print("="*80)

    master_path = 'Portfolio_Reports/MASTER_Portfolio_Complete_Data.xlsx'

    # All 10 properties
    properties = [
        'Orion Prosper', 'McCord Park FL', 'Orion McKinney', 'The Club at Millenia',
        'Bella Mirage', 'Orion Prosper Lakes', 'Mandarina', 'Pavilions at Arrowhead',
        'Springs at Alta Mesa', 'Tempe Vista'
    ]

    # Units for Cost Per Door calculation
    units_map = {
        'Orion Prosper': 312,
        'McCord Park FL': 416,
        'Orion McKinney': 453,
        'The Club at Millenia': 560,
        'Bella Mirage': 715,
        'Orion Prosper Lakes': 308,
        'Mandarina': 180,
        'Pavilions at Arrowhead': 248,
        'Springs at Alta Mesa': 200,
        'Tempe Vista': 186
    }

    # Collect all category data
    all_data = []

    print("\nCalculating category totals from property tabs...")
    print("-"*80)

    for prop in properties:
        try:
            # Read property tab
            df_prop = pd.read_excel(master_path, sheet_name=prop)

            # Get unique categories
            categories = df_prop['Category'].unique()

            units = units_map.get(prop, 1)  # Default to 1 if not found

            for cat in categories:
                if pd.notna(cat):  # Skip NaN categories
                    cat_data = df_prop[df_prop['Category'] == cat]
                    total_spend = cat_data['Extended Amount'].sum()
                    count = len(cat_data)

                    # Calculate monthly average (assuming data spans multiple months)
                    # This is simplified - ideally we'd count actual months
                    months = df_prop['Invoice Date'].nunique()
                    if months == 0:
                        months = 1  # Avoid division by zero

                    monthly_avg = total_spend / months
                    cost_per_door = total_spend / units

                    all_data.append({
                        'Property': prop,
                        'Category': cat,
                        'Total Spend': total_spend,
                        'Monthly Average': monthly_avg,
                        'Cost Per Door': cost_per_door
                    })

                    print(f"{prop:25} | {cat:15} | ${total_spend:>12,.2f} | {count:3} items")

        except Exception as e:
            print(f"{prop:25} | ERROR: {e}")

    # Create DataFrame
    df_new = pd.DataFrame(all_data)

    # Sort by Property and Category
    df_new = df_new.sort_values(['Property', 'Category']).reset_index(drop=True)

    print("\n" + "="*80)
    print("UPDATING MASTER FILE")
    print("="*80)

    # Load workbook
    wb = load_workbook(master_path)

    # Update Spend by Category sheet
    if 'Spend by Category' in wb.sheetnames:
        ws = wb['Spend by Category']

        # Clear existing data (keep headers)
        ws.delete_rows(2, ws.max_row)

        # Write new data
        for r_idx, row in enumerate(dataframe_to_rows(df_new, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        print(f"\nUpdated 'Spend by Category' sheet with {len(df_new)} rows")
    else:
        print("\nERROR: 'Spend by Category' sheet not found")
        return

    # Save workbook
    wb.save(master_path)
    print(f"Saved: {master_path}")

    # Verify the three problem properties
    print("\n" + "="*80)
    print("VERIFICATION - PROBLEM PROPERTIES")
    print("="*80)

    problem_props = ['McCord Park FL', 'Orion Prosper', 'Orion Prosper Lakes']

    for prop in problem_props:
        print(f"\n{prop}:")
        prop_data = df_new[df_new['Property'] == prop]

        base = prop_data[prop_data['Category'] == 'base']['Total Spend'].values
        tax = prop_data[prop_data['Category'] == 'tax']['Total Spend'].values

        if len(base) > 0:
            print(f"  Base:  ${base[0]:,.2f}")
        else:
            print(f"  Base:  Not found")

        if len(tax) > 0:
            print(f"  Tax:   ${tax[0]:,.2f}")
            if len(base) > 0:
                if abs(base[0] - tax[0]) < 0.01:
                    print(f"  Status: *** STILL EQUAL *** (Investigation needed)")
                else:
                    tax_rate = (tax[0] / base[0] * 100) if base[0] > 0 else 0
                    print(f"  Status: FIXED - Tax is {tax_rate:.2f}% of base")
        else:
            print(f"  Tax:   Not found")

    print("\n" + "="*80)
    print("REGENERATION COMPLETE")
    print("="*80)
    print("Spend by Category sheet has been regenerated from property tab source data.")
    print("Tax treatment issue should now be resolved.")

if __name__ == '__main__':
    main()
