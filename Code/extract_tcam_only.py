"""
Extract ONLY The Club at Millenia (TCAM) invoices - filtering out mismatched properties.
Uses Claude Vision API for extraction and adds TCAM data to existing workbook.
"""

import os
import sys
import json
import base64
import re
from datetime import datetime
from pathlib import Path
import anthropic
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Set UTF-8 encoding for Windows console
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding='utf-8')

# Configuration
PROPERTY_NAME = "The Club at Millenia"
MODEL = "claude-sonnet-4-20250514"
MAX_TOKENS = 4000

# Invoice file patterns - ONLY TCAM-named files
TCAM_INVOICE_PATTERNS = [
    "TCAM 4.15.25.pdf",
    "TCAM 5.15.25.pdf",
    "TCAM 6.15.25.pdf",
    "TCAM 7.15.25 (1).pdf",
    "TCAM 8.15.25.pdf",
    "TCAM 9.15.25.pdf",
]

# Paths
BASE_DIR = Path(r"C:\Users\Richard\Downloads\Orion Data Part 2")
INVOICES_DIR = BASE_DIR / "Invoices"
EXTRACTION_OUTPUT_DIR = BASE_DIR / "Extraction_Output"
EXCEL_FILE = EXTRACTION_OUTPUT_DIR / "COMPLETE_All_Properties_20251103_094938.xlsx"

# Create timestamp for output file
TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_FILE = EXTRACTION_OUTPUT_DIR / f"COMPLETE_All_Properties_UPDATED_{TIMESTAMP}.xlsx"


def get_api_key():
    """Get Anthropic API key from environment."""
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise ValueError("ANTHROPIC_API_KEY not found in environment variables")
    return api_key


def encode_pdf_to_base64(pdf_path):
    """Encode PDF file to base64."""
    with open(pdf_path, "rb") as pdf_file:
        return base64.standard_b64encode(pdf_file.read()).decode("utf-8")


def extract_invoice_data(client, pdf_path):
    """
    Extract invoice data from PDF using Claude Vision API.

    Returns structured data with invoice details and line items.
    """
    print(f"  Processing: {pdf_path.name}")

    # Encode PDF
    pdf_base64 = encode_pdf_to_base64(pdf_path)

    # Extraction prompt
    extraction_prompt = f"""You are analyzing a waste management invoice for {PROPERTY_NAME}.

Extract the following information from this invoice:

CRITICAL FIELDS (REQUIRED):
1. Invoice Number: The unique invoice identifier
2. Invoice Date: Date invoice was issued (format: YYYY-MM-DD)
3. Service Period: Billing period (format: YYYY-MM-DD to YYYY-MM-DD)
4. Total Amount: Total invoice amount (numeric only, no $)
5. Property Name: Verify this is for "The Club at Millenia" (or similar name)

DETAIL FIELDS:
6. Service Provider: Company providing service
7. Account Number: Customer/account number
8. Due Date: Payment due date (format: YYYY-MM-DD)
9. Previous Balance: Any previous balance (numeric, 0 if none)
10. Payments: Any payments made (numeric, 0 if none)

LINE ITEMS:
For each line item, extract:
- Description: Service description
- Quantity: Number of units/services
- Unit Price: Price per unit (numeric)
- Amount: Line item total (numeric)
- Service Type: Categorize as "Base Service", "Extra Pickup", "Fuel Surcharge", "Environmental Fee", "Other"

VALIDATION:
- CRITICAL: If the property name does NOT match "The Club at Millenia" or similar, flag it with "PROPERTY_MISMATCH: [actual property name]"
- If any CRITICAL field is missing, flag it with "MISSING: [field name]"
- If dates are ambiguous, flag with "AMBIGUOUS: [field name]"

Return the data as a JSON object with this structure:
{{
  "invoice_number": "INV-XXXXX",
  "invoice_date": "2025-04-15",
  "service_period_start": "2025-04-01",
  "service_period_end": "2025-04-30",
  "total_amount": 1234.56,
  "property_name": "{PROPERTY_NAME}",
  "service_provider": "Provider Name",
  "account_number": "ACCT-XXXXX",
  "due_date": "2025-05-15",
  "previous_balance": 0.00,
  "payments": 0.00,
  "line_items": [
    {{
      "description": "Service description",
      "quantity": 1,
      "unit_price": 100.00,
      "amount": 100.00,
      "service_type": "Base Service"
    }}
  ],
  "validation_flags": [],
  "extraction_confidence": "high"
}}

If you cannot extract a field with confidence, use null for that field and add a flag to validation_flags.
Be precise with numbers - extract exact amounts without $ symbols or commas."""

    try:
        # Call Claude API
        response = client.messages.create(
            model=MODEL,
            max_tokens=MAX_TOKENS,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "document",
                            "source": {
                                "type": "base64",
                                "media_type": "application/pdf",
                                "data": pdf_base64,
                            },
                        },
                        {
                            "type": "text",
                            "text": extraction_prompt,
                        },
                    ],
                }
            ],
        )

        # Extract JSON from response
        response_text = response.content[0].text

        # Try to find JSON in response
        json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
        if json_match:
            invoice_data = json.loads(json_match.group())
            invoice_data['source_file'] = pdf_path.name
            return invoice_data
        else:
            print(f"  ⚠️  No JSON found in response for {pdf_path.name}")
            return None

    except Exception as e:
        print(f"  ❌ Error extracting {pdf_path.name}: {str(e)}")
        return None


def extract_all_tcam_invoices():
    """Extract all TCAM invoices from PDFs."""
    print(f"\n{'='*80}")
    print(f"EXTRACTING INVOICES FOR: {PROPERTY_NAME}")
    print(f"{'='*80}\n")

    # Initialize API client
    api_key = get_api_key()
    client = anthropic.Anthropic(api_key=api_key)

    # Collect all invoice files (ONLY TCAM-named files)
    invoice_files = []

    for pattern in TCAM_INVOICE_PATTERNS:
        file_path = INVOICES_DIR / pattern
        if file_path.exists():
            invoice_files.append(file_path)
        else:
            print(f"⚠️  File not found: {pattern}")

    print(f"Found {len(invoice_files)} TCAM invoice files to process\n")

    # Extract data from each invoice
    extracted_data = []
    failed_files = []
    skipped_files = []

    for i, pdf_path in enumerate(invoice_files, 1):
        print(f"[{i}/{len(invoice_files)}]", end=" ")

        invoice_data = extract_invoice_data(client, pdf_path)

        if invoice_data:
            # Check for property mismatch
            validation_flags = invoice_data.get('validation_flags', [])
            has_mismatch = any('PROPERTY' in flag.upper() and 'MISMATCH' in flag.upper()
                              for flag in validation_flags)

            if has_mismatch:
                print(f"  ⚠️  SKIPPED - Property mismatch detected")
                for flag in validation_flags:
                    if 'PROPERTY' in flag.upper() and 'MISMATCH' in flag.upper():
                        print(f"    {flag}")
                skipped_files.append(pdf_path.name)
            else:
                extracted_data.append(invoice_data)
                print(f"  ✓ Extracted successfully")

                # Show other validation flags if any
                if validation_flags:
                    for flag in validation_flags:
                        print(f"    🔴 {flag}")
        else:
            failed_files.append(pdf_path.name)
            print(f"  ❌ Extraction failed")

    print(f"\n{'='*80}")
    print(f"Extraction Summary:")
    print(f"  ✓ Successfully extracted: {len(extracted_data)} invoices")
    print(f"  ⚠️  Skipped (property mismatch): {len(skipped_files)} invoices")
    print(f"  ❌ Failed: {len(failed_files)} invoices")

    if skipped_files:
        print(f"\nSkipped files (wrong property):")
        for file in skipped_files:
            print(f"  - {file}")

    if failed_files:
        print(f"\nFailed files:")
        for file in failed_files:
            print(f"  - {file}")
    print(f"{'='*80}\n")

    return extracted_data


def create_tcam_dataframe(extracted_data):
    """Convert extracted data to DataFrame matching Excel structure."""
    rows = []

    for invoice in extracted_data:
        # Create rows for each line item
        line_items = invoice.get('line_items', [])

        if not line_items:
            # If no line items, create single row with totals
            row = {
                'Property Name': PROPERTY_NAME,
                'Invoice Number': invoice.get('invoice_number'),
                'Invoice Date': invoice.get('invoice_date'),
                'Service Period Start': invoice.get('service_period_start'),
                'Service Period End': invoice.get('service_period_end'),
                'Service Provider': invoice.get('service_provider'),
                'Account Number': invoice.get('account_number'),
                'Due Date': invoice.get('due_date'),
                'Previous Balance': invoice.get('previous_balance', 0),
                'Payments': invoice.get('payments', 0),
                'Line Item Description': 'TOTAL',
                'Quantity': None,
                'Unit Price': None,
                'Line Item Amount': invoice.get('total_amount'),
                'Service Type': 'Total',
                'Total Amount': invoice.get('total_amount'),
                'Validation Flags': ', '.join(invoice.get('validation_flags', [])),
                'Source File': invoice.get('source_file'),
                'Extraction Confidence': invoice.get('extraction_confidence', 'medium'),
            }
            rows.append(row)
        else:
            # Create row for each line item
            for item in line_items:
                row = {
                    'Property Name': PROPERTY_NAME,
                    'Invoice Number': invoice.get('invoice_number'),
                    'Invoice Date': invoice.get('invoice_date'),
                    'Service Period Start': invoice.get('service_period_start'),
                    'Service Period End': invoice.get('service_period_end'),
                    'Service Provider': invoice.get('service_provider'),
                    'Account Number': invoice.get('account_number'),
                    'Due Date': invoice.get('due_date'),
                    'Previous Balance': invoice.get('previous_balance', 0),
                    'Payments': invoice.get('payments', 0),
                    'Line Item Description': item.get('description'),
                    'Quantity': item.get('quantity'),
                    'Unit Price': item.get('unit_price'),
                    'Line Item Amount': item.get('amount'),
                    'Service Type': item.get('service_type'),
                    'Total Amount': invoice.get('total_amount'),
                    'Validation Flags': ', '.join(invoice.get('validation_flags', [])),
                    'Source File': invoice.get('source_file'),
                    'Extraction Confidence': invoice.get('extraction_confidence', 'medium'),
                }
                rows.append(row)

    df = pd.DataFrame(rows)

    # Sort by invoice date
    if 'Invoice Date' in df.columns and not df.empty:
        df = df.sort_values('Invoice Date')

    return df


def update_excel_file(tcam_df):
    """Update existing Excel file with TCAM data."""
    print(f"\n{'='*80}")
    print(f"UPDATING EXCEL FILE")
    print(f"{'='*80}\n")

    if not EXCEL_FILE.exists():
        print(f"❌ Excel file not found: {EXCEL_FILE}")
        return False

    print(f"Reading existing Excel file: {EXCEL_FILE.name}")

    # Load existing workbook
    wb = load_workbook(EXCEL_FILE)

    # Check if TCAM sheet already exists
    if PROPERTY_NAME in wb.sheetnames:
        print(f"⚠️  Sheet '{PROPERTY_NAME}' already exists. Removing old version...")
        wb.remove(wb[PROPERTY_NAME])

    # Create new sheet for TCAM
    ws = wb.create_sheet(PROPERTY_NAME)

    # Write DataFrame to sheet
    for r in dataframe_to_rows(tcam_df, index=False, header=True):
        ws.append(r)

    # Format header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Update Portfolio Summary sheet if it exists
    if "Portfolio Summary" in wb.sheetnames:
        print(f"Updating Portfolio Summary to include {PROPERTY_NAME}...")
        update_portfolio_summary(wb, tcam_df)

    # Save updated workbook
    print(f"Saving updated Excel file: {OUTPUT_FILE.name}")
    wb.save(OUTPUT_FILE)

    print(f"\n✓ Excel file updated successfully!")
    print(f"  - Added sheet: {PROPERTY_NAME}")
    print(f"  - Total rows: {len(tcam_df)}")
    print(f"  - Output file: {OUTPUT_FILE}")
    print(f"{'='*80}\n")

    return True


def update_portfolio_summary(wb, tcam_df):
    """Update Portfolio Summary sheet with TCAM data."""
    ws = wb["Portfolio Summary"]

    # Calculate TCAM summary metrics
    total_invoices = tcam_df['Invoice Number'].nunique()
    total_amount = tcam_df.groupby('Invoice Number')['Total Amount'].first().sum()
    avg_invoice = total_amount / total_invoices if total_invoices > 0 else 0

    # Find next empty row
    next_row = ws.max_row + 1

    # Add TCAM summary row
    ws.append([
        PROPERTY_NAME,
        total_invoices,
        f"${total_amount:,.2f}",
        f"${avg_invoice:,.2f}",
        "See property sheet for details"
    ])

    # Format the new row
    for cell in ws[next_row]:
        cell.alignment = Alignment(horizontal="left", vertical="center")


def main():
    """Main execution function."""
    try:
        # Create output directory if needed
        EXTRACTION_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        # Step 1: Extract TCAM invoices
        print(f"Step 1: Extracting invoice data from PDFs...")
        extracted_data = extract_all_tcam_invoices()

        if not extracted_data:
            print("❌ No invoices extracted successfully. Exiting.")
            print("\nNote: If all files were skipped due to property mismatch,")
            print("verify that the PDF files are actually for The Club at Millenia.")
            return 1

        # Step 2: Convert to DataFrame
        print(f"Step 2: Converting extracted data to DataFrame...")
        tcam_df = create_tcam_dataframe(extracted_data)
        print(f"  ✓ Created DataFrame with {len(tcam_df)} rows")

        # Step 3: Update Excel file
        print(f"Step 3: Updating Excel file...")
        success = update_excel_file(tcam_df)

        if success:
            print(f"\n{'='*80}")
            print(f"✓ TASK COMPLETED SUCCESSFULLY")
            print(f"{'='*80}")
            print(f"\nOutput file: {OUTPUT_FILE}")
            print(f"Property: {PROPERTY_NAME}")
            print(f"Invoices processed: {tcam_df['Invoice Number'].nunique()}")
            print(f"Total rows: {len(tcam_df)}")
            print(f"\nNext steps:")
            print(f"1. Open the Excel file to review TCAM data")
            print(f"2. Check validation flags for any issues")
            print(f"3. Verify Portfolio Summary includes TCAM")
            print(f"{'='*80}\n")
            return 0
        else:
            print("❌ Failed to update Excel file")
            return 1

    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
