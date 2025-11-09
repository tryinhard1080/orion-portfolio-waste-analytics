"""
Update Arizona Properties with Correct Unit Counts and Recalculate YPD

This script updates unit counts for Arizona properties and recalculates YPD values.
"""

import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
import shutil

# Paths
BASE_DIR = Path(__file__).parent.parent
MASTER_FILE = BASE_DIR / "Portfolio_Reports" / "MASTER_Portfolio_Complete_Data.xlsx"
BACKUP_FILE = BASE_DIR / "Portfolio_Reports" / f"MASTER_Portfolio_Complete_Data_BACKUP_UNITS_FIX_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# Corrected unit counts from user
CORRECTED_UNITS = {
    'Pavilions at Arrowhead': 248,  # Was: TBD
    'Springs at Alta Mesa': 200,    # Was: 200 (correct)
    'Mandarina': 180,               # Was: 180 (correct)
    'Tempe Vista': 186              # Was: 150 (INCORRECT - now corrected)
}

# Service details (from previous update)
SERVICE_DETAILS = {
    'Tempe Vista': {
        'total_monthly_yards': 493.62,
        'old_units': 150,
        'new_units': 186
    },
    'Mandarina': {
        'total_monthly_yards': 155.88,
        'old_units': 180,
        'new_units': 180  # No change
    },
    'Springs at Alta Mesa': {
        'total_monthly_yards': 600.63,
        'old_units': 200,
        'new_units': 200  # No change
    },
    'Pavilions at Arrowhead': {
        'total_monthly_yards': 138.56,
        'old_units': None,  # Was TBD
        'new_units': 248
    }
}

def calculate_new_ypd(monthly_yards, units):
    """Calculate YPD"""
    return monthly_yards / units

def show_corrections():
    """Show unit count corrections and YPD recalculations"""
    
    print('=' * 80)
    print('ARIZONA PROPERTIES - UNIT COUNT CORRECTIONS & YPD RECALCULATION')
    print('=' * 80)
    print()
    
    results = {}
    
    for property_name, data in SERVICE_DETAILS.items():
        old_units = data['old_units']
        new_units = data['new_units']
        monthly_yards = data['total_monthly_yards']
        
        print(f'{property_name}:')
        print(f'  Old Units: {old_units if old_units else "TBD"}')
        print(f'  New Units: {new_units}')
        
        if old_units != new_units:
            if old_units:
                print(f'  ⚠️ CORRECTION NEEDED')
            else:
                print(f'  ✅ NEW DATA (was TBD)')
        else:
            print(f'  ✅ No change needed')
        
        print()
        
        # Calculate old YPD if units were known
        if old_units:
            old_ypd = monthly_yards / old_units
            print(f'  Old YPD: {monthly_yards:.2f} / {old_units} = {old_ypd:.2f}')
        else:
            print(f'  Old YPD: TBD (units unknown)')
        
        # Calculate new YPD
        new_ypd = monthly_yards / new_units
        print(f'  New YPD: {monthly_yards:.2f} / {new_units} = {new_ypd:.2f}')
        
        # Show change
        if old_units and old_units != new_units:
            change = new_ypd - old_ypd
            pct_change = (change / old_ypd) * 100
            print(f'  Change: {change:+.2f} ({pct_change:+.1f}%)')
        
        print()
        
        # Compare to target
        target = 2.0
        if new_ypd <= target:
            pct_below = ((target - new_ypd) / target) * 100
            print(f'  ✅ {pct_below:.1f}% below target (2.0)')
        else:
            pct_above = ((new_ypd - target) / target) * 100
            print(f'  ⚠️ {pct_above:.1f}% above target (2.0)')
        
        print()
        print('-' * 80)
        print()
        
        results[property_name] = {
            'old_units': old_units,
            'new_units': new_units,
            'old_ypd': monthly_yards / old_units if old_units else None,
            'new_ypd': new_ypd,
            'monthly_yards': monthly_yards
        }
    
    return results

def create_backup():
    """Create a backup of the master file before updating"""
    
    print('=' * 80)
    print('CREATING BACKUP')
    print('=' * 80)
    print()
    
    if MASTER_FILE.exists():
        shutil.copy2(MASTER_FILE, BACKUP_FILE)
        print(f'✅ Backup created: {BACKUP_FILE.name}')
        print()
        return True
    else:
        print(f'❌ Master file not found: {MASTER_FILE}')
        return False

def update_units_and_ypd(results):
    """Update unit counts and YPD in master file"""
    
    print('=' * 80)
    print('UPDATING MASTER FILE')
    print('=' * 80)
    print()
    
    wb = openpyxl.load_workbook(MASTER_FILE)
    
    for property_name, data in results.items():
        if property_name not in wb.sheetnames:
            print(f'  ⚠️ Sheet not found: {property_name}')
            continue
        
        ws = wb[property_name]
        
        print(f'Updating: {property_name}')
        print('-' * 80)
        
        # Find header row
        header_row = None
        for row in range(1, 10):
            cell_value = ws.cell(row, 1).value
            if cell_value and ('Property' in str(cell_value)):
                header_row = row
                break
        
        if not header_row:
            print(f'  ❌ Could not find header row')
            continue
        
        # Find YPD column
        ypd_col = None
        for col in range(1, 50):
            header = ws.cell(header_row, col).value
            if header == 'YPD':
                ypd_col = col
                break
        
        if not ypd_col:
            print(f'  ⚠️ YPD column not found')
            continue
        
        # Update all data rows with new YPD
        row = header_row + 1
        rows_updated = 0
        
        while ws.cell(row, 1).value:
            ws.cell(row, ypd_col).value = round(data['new_ypd'], 2)
            rows_updated += 1
            row += 1
        
        print(f'  ✓ Updated {rows_updated} rows')

        old_ypd_str = f'{data["old_ypd"]:.2f}' if data["old_ypd"] else "TBD"
        print(f'  ✓ New YPD: {data["new_ypd"]:.2f} (was {old_ypd_str})')
        print(f'  ✅ {property_name} updated successfully')
        print()
    
    wb.save(MASTER_FILE)
    wb.close()
    
    print('✅ Master file updated successfully')
    print()

def verify_updates(results):
    """Verify that updates were applied correctly"""
    
    print('=' * 80)
    print('VERIFYING UPDATES')
    print('=' * 80)
    print()
    
    for property_name, expected_data in results.items():
        df = pd.read_excel(MASTER_FILE, sheet_name=property_name)
        
        if 'YPD' in df.columns:
            actual_ypd = df['YPD'].iloc[0]
            expected_ypd = expected_data['new_ypd']
            
            if isinstance(actual_ypd, (int, float)) and abs(actual_ypd - expected_ypd) < 0.01:
                print(f'  ✅ {property_name}: YPD = {actual_ypd:.2f} (matches expected {expected_ypd:.2f})')
            else:
                print(f'  ⚠️ {property_name}: YPD = {actual_ypd} (expected {expected_ypd:.2f})')
        else:
            print(f'  ⚠️ {property_name}: YPD column not found')
    
    print()

def create_summary_report(results):
    """Create a summary report of the unit count corrections"""
    
    report_path = BASE_DIR / 'Portfolio_Reports' / 'UNIT_COUNT_CORRECTIONS_SUMMARY.md'
    
    content = f"""# Arizona Properties - Unit Count Corrections

**Date:** {datetime.now().strftime('%B %d, %Y at %I:%M %p')}  
**Purpose:** Correct unit counts and recalculate YPD for Arizona properties

---

## UNIT COUNT CORRECTIONS

"""
    
    for property_name, data in results.items():
        old_units = data['old_units']
        new_units = data['new_units']
        old_ypd = data['old_ypd']
        new_ypd = data['new_ypd']
        monthly_yards = data['monthly_yards']
        
        content += f"""### {property_name}

**Unit Count:**
- Old: {old_units if old_units else 'TBD'}
- New: **{new_units}**
"""
        
        if old_units != new_units:
            if old_units:
                content += f"- Status: ⚠️ **CORRECTED**\n\n"
            else:
                content += f"- Status: ✅ **NEW DATA** (was TBD)\n\n"
        else:
            content += f"- Status: ✅ No change needed\n\n"
        
        old_ypd_str = f'{old_ypd:.2f}' if old_ypd else 'TBD'
        old_units_str = str(old_units) if old_units else 'TBD'

        content += f"""**YPD Calculation:**
- Monthly Yards: {monthly_yards:.2f}
- Old YPD: {old_ypd_str} ({monthly_yards:.2f} / {old_units_str})
- New YPD: **{new_ypd:.2f}** ({monthly_yards:.2f} / {new_units})
"""
        
        if old_ypd and old_units != new_units:
            change = new_ypd - old_ypd
            pct_change = (change / old_ypd) * 100
            content += f"- Change: {change:+.2f} ({pct_change:+.1f}%)\n\n"
        else:
            content += "\n"
        
        target = 2.0
        if new_ypd <= target:
            pct_below = ((target - new_ypd) / target) * 100
            content += f"**Performance:** ✅ {pct_below:.1f}% below target (2.0)\n\n"
        else:
            pct_above = ((new_ypd - target) / target) * 100
            content += f"**Performance:** ⚠️ {pct_above:.1f}% above target (2.0)\n\n"
        
        content += "---\n\n"
    
    content += """## SUMMARY TABLE

| Property | Old Units | New Units | Old YPD | New YPD | Change | Status |
|----------|-----------|-----------|---------|---------|--------|--------|
"""
    
    for property_name, data in results.items():
        old_units_str = str(data['old_units']) if data['old_units'] else 'TBD'
        old_ypd_str = f"{data['old_ypd']:.2f}" if data['old_ypd'] else 'TBD'
        
        if data['old_ypd'] and data['old_units'] != data['new_units']:
            change = data['new_ypd'] - data['old_ypd']
            change_str = f"{change:+.2f}"
            status = "⚠️ Corrected"
        elif not data['old_units']:
            change_str = "New"
            status = "✅ Added"
        else:
            change_str = "—"
            status = "✅ No change"
        
        content += f"| {property_name} | {old_units_str} | **{data['new_units']}** | {old_ypd_str} | **{data['new_ypd']:.2f}** | {change_str} | {status} |\n"
    
    content += """
---

## CORRECTED PORTFOLIO TOTALS

**All 8 Properties with Complete Data:**

| Property | State | Units | Containers | Monthly Yards | YPD | Performance |
|----------|-------|-------|------------|---------------|-----|-------------|
| Bella Mirage | AZ | 715 | 6 | 623.52 | 0.87 | ✅ Excellent |
| Mandarina | AZ | 180 | 2 | 155.88 | 0.87 | ✅ Excellent |
| The Club at Millenia | FL | 560 | 6 | 727.44 | 1.30 | ✅ Excellent |
| Orion McKinney | TX | 453 | 10 | 1,091.16 | 2.41 | ⚠️ High |
"""
    
    # Add corrected AZ properties
    for prop in ['Tempe Vista', 'Springs at Alta Mesa', 'Pavilions at Arrowhead']:
        if prop in results:
            data = results[prop]
            state = 'AZ'
            
            # Get container count from SERVICE_DETAILS
            if prop == 'Tempe Vista':
                containers = 8
            elif prop == 'Springs at Alta Mesa':
                containers = 16
            else:  # Pavilions
                containers = 4
            
            monthly_yards = data['monthly_yards']
            ypd = data['new_ypd']
            
            if ypd <= 2.0:
                perf = "✅ Excellent"
            elif ypd <= 2.25:
                perf = "✅ Good"
            else:
                perf = "⚠️ High"
            
            content += f"| {prop} | {state} | {data['new_units']} | {containers} | {monthly_yards:.2f} | {ypd:.2f} | {perf} |\n"
    
    content += f"""| McCord Park FL | FL | 416 | 15 | 1,437.56 | 3.46 | ⚠️ High |

**Portfolio Totals:**
- **Total Units:** {sum(CORRECTED_UNITS.values()) + 715 + 560 + 453 + 416} (8 properties)
- **Total Containers:** 67
- **Portfolio Average YPD:** {(0.87 + 0.87 + 1.30 + 2.41 + results['Tempe Vista']['new_ypd'] + results['Springs at Alta Mesa']['new_ypd'] + results['Pavilions at Arrowhead']['new_ypd'] + 3.46) / 8:.2f}

---

## BACKUP INFORMATION

**Backup File:** """ + BACKUP_FILE.name + """  
**Location:** Portfolio_Reports/

---

**Unit count corrections completed successfully!**
"""
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(content)
    
    print(f'✅ Summary report created: {report_path.name}')
    print()

def main():
    """Main function"""
    
    # Show corrections
    results = show_corrections()
    
    # Create backup
    if not create_backup():
        print('❌ Cannot proceed without backup')
        return
    
    # Update master file
    update_units_and_ypd(results)
    
    # Verify updates
    verify_updates(results)
    
    # Create summary report
    create_summary_report(results)
    
    print('=' * 80)
    print('UNIT COUNT CORRECTIONS COMPLETE')
    print('=' * 80)
    print()
    print('Summary:')
    for property_name, data in results.items():
        print(f'  {property_name}:')
        old_units_str = str(data["old_units"]) if data["old_units"] else "TBD"
        old_ypd_str = f'{data["old_ypd"]:.2f}' if data["old_ypd"] else "TBD"
        print(f'    Units: {old_units_str} → {data["new_units"]}')
        print(f'    YPD: {old_ypd_str} → {data["new_ypd"]:.2f}')
    print()

if __name__ == "__main__":
    main()

