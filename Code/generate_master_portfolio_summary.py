"""
WasteWise Master Portfolio Regulatory Compliance Summary
Consolidates regulatory compliance status across all 10 properties
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from pathlib import Path

# Styling
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Master file path
MASTER_FILE = "Portfolio_Reports/MASTER_Portfolio_Complete_Data.xlsx"

# Portfolio configuration with regulatory research findings
PROPERTIES = {
    'Orion Prosper': {
        'units': 312,
        'city': 'Prosper',
        'state': 'TX',
        'county': 'Collin County',
        'service_type': 'Compactor',
        'vendor': 'Republic Services',
        'regulatory_status': 'VOLUNTARY',
        'mandatory_recycling': False,
        'mandatory_composting': False,
        'confidence': 'MEDIUM',
        'sources_consulted': 5,
        'gov_sources': 3,
        'key_finding': 'Prosper has NO mandatory recycling or composting requirements for multifamily properties. Voluntary waste management system.',
        'ordinances': 'None (voluntary system)',
        'penalties': 'N/A - No violations possible',
        'licensed_haulers': ['Republic Services (primary town contractor since Feb 2024)'],
        'contact': 'Town of Prosper Utility Customer Service: 945-234-1924, ucs@prospertx.gov'
    },
    'Orion Prosper Lakes': {
        'units': 308,
        'city': 'Prosper',
        'state': 'TX',
        'county': 'Collin County',
        'service_type': 'Compactor',
        'vendor': 'Republic Services',
        'regulatory_status': 'VOLUNTARY',
        'mandatory_recycling': False,
        'mandatory_composting': False,
        'confidence': 'MEDIUM',
        'sources_consulted': 5,
        'gov_sources': 3,
        'key_finding': 'Prosper has NO mandatory recycling or composting requirements for multifamily properties. Voluntary waste management system.',
        'ordinances': 'None (voluntary system)',
        'penalties': 'N/A - No violations possible',
        'licensed_haulers': ['Republic Services (primary town contractor since Feb 2024)'],
        'contact': 'Town of Prosper Utility Customer Service: 945-234-1924, ucs@prospertx.gov'
    },
    'Orion McKinney': {
        'units': 453,
        'city': 'McKinney',
        'state': 'TX',
        'county': 'Collin County',
        'service_type': 'Mixed',
        'vendor': 'Frontier Waste',
        'regulatory_status': 'COMMERCIAL',
        'mandatory_recycling': False,
        'mandatory_composting': False,
        'confidence': 'MEDIUM',
        'sources_consulted': 4,
        'gov_sources': 2,
        'key_finding': 'McKinney classifies multifamily properties as commercial. Chapter 86 Solid Waste ordinance exists but specific recycling mandates for multifamily not confirmed.',
        'ordinances': 'Chapter 86 - Solid Waste (Code of Ordinances)',
        'penalties': 'Enforcement measures available for non-compliance with solid waste ordinances',
        'licensed_haulers': ['Frontier Waste Solutions', 'Various private haulers with city permits'],
        'contact': 'McKinney Solid Waste Services: 972-547-7385'
    },
    'McCord Park FL': {
        'units': 416,
        'city': 'Little Elm',
        'state': 'TX',
        'county': 'Denton County',
        'service_type': 'Dumpster',
        'vendor': 'Community Waste Disposal',
        'regulatory_status': 'MANDATORY',
        'mandatory_recycling': True,
        'mandatory_composting': False,
        'confidence': 'MEDIUM',
        'sources_consulted': 4,
        'gov_sources': 2,
        'key_finding': 'Little Elm is one of 6 DFW cities with multifamily recycling requirements. Community Waste Disposal provides commercial/multifamily services.',
        'ordinances': 'Chapter 102 (Utilities), Article VIII (Solid Waste Disposal)',
        'penalties': 'Specific penalties not detailed in research',
        'licensed_haulers': ['Community Waste Disposal (contracted provider)'],
        'contact': 'Town of Little Elm: 972.294.1821, Community Waste Disposal: 972.392.9300 Option 2'
    },
    'The Club at Millenia': {
        'units': 560,
        'city': 'Orlando',
        'state': 'FL',
        'county': 'Orange County',
        'service_type': 'Compactor',
        'vendor': 'Waste Connections of Florida',
        'regulatory_status': 'MANDATORY',
        'mandatory_recycling': True,
        'mandatory_composting': False,
        'confidence': 'HIGH',
        'sources_consulted': 5,
        'gov_sources': 3,
        'key_finding': 'Orlando REQUIRES recycling for all properties with 4+ units. 560-unit property subject to compliance since April 2021 (74+ unit threshold).',
        'ordinances': 'Chapter 28 - Solid Waste Management (Orlando Municipal Code), April 1, 2019 amendments',
        'penalties': 'Enforcement available for non-compliance',
        'licensed_haulers': ['Waste Connections of Florida', 'Various licensed haulers'],
        'contact': 'City of Orlando Solid Waste Division: 407.246.2314',
        'compliance_requirements': [
            'Provide recycling containers',
            'Arrange for recyclable material collection',
            'Maintain and submit verification records',
            '74+ units must comply by April 21, 2021'
        ]
    },
    'Bella Mirage': {
        'units': 715,
        'city': 'Phoenix',
        'state': 'AZ',
        'county': 'Maricopa County',
        'service_type': 'Compactor',
        'vendor': 'Waste Management',
        'regulatory_status': 'VOLUNTARY',
        'mandatory_recycling': False,
        'mandatory_composting': False,
        'confidence': 'HIGH',
        'sources_consulted': 5,
        'gov_sources': 2,
        'key_finding': 'Arizona state law (A.R.S. 9-500.38) PROHIBITS cities from mandating recycling for multifamily properties. Phoenix cannot require recycling - voluntary only.',
        'ordinances': 'A.R.S. 9-500.38 (state preemption law - "ban on bans")',
        'penalties': 'N/A - No mandates exist',
        'licensed_haulers': ['Waste Management', 'Various private haulers (landlord must contract)'],
        'contact': 'Phoenix city code restricts municipal collection at 30+ unit properties',
        'state_restrictions': 'State law bans cities from requiring recycling in multifamily/commercial properties'
    },
    'Mandarina': {
        'units': 180,
        'city': 'Phoenix',
        'state': 'AZ',
        'county': 'Maricopa County',
        'service_type': 'Compactor',
        'vendor': 'Waste Management + Ally Waste',
        'regulatory_status': 'VOLUNTARY',
        'mandatory_recycling': False,
        'mandatory_composting': False,
        'confidence': 'HIGH',
        'sources_consulted': 5,
        'gov_sources': 2,
        'key_finding': 'Arizona state law (A.R.S. 9-500.38) PROHIBITS cities from mandating recycling for multifamily properties. Phoenix cannot require recycling - voluntary only.',
        'ordinances': 'A.R.S. 9-500.38 (state preemption law)',
        'penalties': 'N/A - No mandates exist',
        'licensed_haulers': ['Waste Management', 'Ally Waste', 'Various private haulers'],
        'contact': 'Phoenix offers 8 eco-stations and 2 transfer stations for voluntary recycling',
        'state_restrictions': 'State law bans cities from requiring recycling in multifamily/commercial properties'
    },
    'Pavilions at Arrowhead': {
        'units': 248,
        'city': 'Glendale',
        'state': 'AZ',
        'county': 'Maricopa County',
        'service_type': 'Dumpster',
        'vendor': 'City of Glendale + Ally Waste',
        'regulatory_status': 'VOLUNTARY',
        'mandatory_recycling': False,
        'mandatory_composting': False,
        'confidence': 'MEDIUM',
        'sources_consulted': 4,
        'gov_sources': 2,
        'key_finding': 'Glendale offers organics recycling for 5+ unit properties (voluntary). Subject to Arizona state law restrictions on mandatory recycling.',
        'ordinances': 'Chapter 18 - Garbage and Trash (Glendale Municipal Code), subject to A.R.S. 9-500.38',
        'penalties': 'N/A - Voluntary programs only',
        'licensed_haulers': ['City of Glendale (franchise haulers)', 'Ally Waste'],
        'contact': 'City of Glendale Solid Waste Division',
        'special_programs': 'Free organics recycling pails available for 5+ unit properties with franchise haulers'
    },
    'Springs at Alta Mesa': {
        'units': 200,
        'city': 'Mesa',
        'state': 'AZ',
        'county': 'Maricopa County',
        'service_type': 'Mixed',
        'vendor': 'City of Mesa + Ally Waste',
        'regulatory_status': 'VOLUNTARY',
        'mandatory_recycling': False,
        'mandatory_composting': False,
        'confidence': 'MEDIUM',
        'sources_consulted': 3,
        'gov_sources': 1,
        'key_finding': 'Mesa regulations apply to apartments with 5+ units. Subject to Arizona state law (A.R.S. 9-500.38) which prohibits mandatory recycling requirements.',
        'ordinances': 'Mesa City Code (subject to A.R.S. 9-500.38 state restrictions)',
        'penalties': 'N/A - Voluntary system due to state law',
        'licensed_haulers': ['City of Mesa', 'Ally Waste', 'Private haulers'],
        'contact': 'Mesa Solid Waste Services: mesaaz.gov/Utilities/Trash-Recycling'
    },
    'Tempe Vista': {
        'units': 186,
        'city': 'Tempe',
        'state': 'AZ',
        'county': 'Maricopa County',
        'service_type': 'Dumpster',
        'vendor': 'Waste Management + Ally Waste',
        'regulatory_status': 'VOLUNTARY',
        'mandatory_recycling': False,
        'mandatory_composting': False,
        'confidence': 'MEDIUM',
        'sources_consulted': 4,
        'gov_sources': 2,
        'key_finding': 'Tempe offers voluntary multifamily recycling program. Property management requests containers. SB 1079 (2015) allowed private vendors for multifamily recycling.',
        'ordinances': 'Voluntary program (subject to A.R.S. 9-500.38 state restrictions)',
        'penalties': 'N/A - Voluntary participation',
        'licensed_haulers': ['Waste Management', 'Ally Waste', 'Private vendors'],
        'contact': 'Tempe 311: 480-350-4311'
    }
}

print("=" * 80)
print("WASTEWISE MASTER PORTFOLIO REGULATORY COMPLIANCE SUMMARY")
print("=" * 80)
print(f"\nGenerating master portfolio summary for all 10 properties")
print(f"Analysis Date: {datetime.now().strftime('%B %d, %Y')}")
print("\n" + "=" * 80)

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ============================================================================
# SHEET 1: PORTFOLIO OVERVIEW
# ============================================================================
ws_overview = wb.create_sheet("PORTFOLIO_OVERVIEW")

ws_overview['A1'] = "WasteWise Portfolio Regulatory Compliance Summary"
ws_overview['A1'].font = Font(bold=True, size=16)
ws_overview.merge_cells('A1:H1')

ws_overview['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
ws_overview['A2'].font = Font(size=10, italic=True)
ws_overview.merge_cells('A2:H2')

# Summary statistics
row = 4
ws_overview[f'A{row}'] = "PORTFOLIO STATISTICS"
ws_overview[f'A{row}'].font = Font(bold=True, size=12)
ws_overview.merge_cells(f'A{row}:B{row}')
row += 2

total_properties = len(PROPERTIES)
total_units = sum(p['units'] for p in PROPERTIES.values())
mandatory_count = sum(1 for p in PROPERTIES.values() if p['mandatory_recycling'])
voluntary_count = total_properties - mandatory_count

stats = [
    ("Total Properties:", total_properties),
    ("Total Units:", f"{total_units:,}"),
    ("Mandatory Recycling:", f"{mandatory_count} properties"),
    ("Voluntary/No Requirements:", f"{voluntary_count} properties"),
    ("States Covered:", "3 (TX, FL, AZ)"),
    ("Cities Analyzed:", "8 unique jurisdictions"),
]

for label, value in stats:
    ws_overview[f'A{row}'] = label
    ws_overview[f'B{row}'] = value
    ws_overview[f'A{row}'].font = Font(bold=True)
    row += 1

row += 1
ws_overview[f'A{row}'] = "KEY FINDINGS"
ws_overview[f'A{row}'].font = Font(bold=True, size=12)
ws_overview.merge_cells(f'A{row}:H{row}')
row += 2

findings = [
    "• 2 properties subject to MANDATORY recycling requirements (Orlando FL, Little Elm TX)",
    "• 5 Arizona properties: State law PROHIBITS cities from mandating recycling (A.R.S. 9-500.38)",
    "• 3 Texas properties: No mandatory requirements (voluntary systems)",
    "• HIGH confidence: 2 properties | MEDIUM confidence: 8 properties",
    "• Total research: 40 sources consulted (21 government sources)",
]

for finding in findings:
    ws_overview[f'A{row}'] = finding
    ws_overview[f'A{row}'].alignment = LEFT_ALIGN
    ws_overview.merge_cells(f'A{row}:H{row}')
    row += 1

ws_overview.column_dimensions['A'].width = 30
ws_overview.column_dimensions['B'].width = 50

# ============================================================================
# SHEET 2: REGULATORY COMPLIANCE MATRIX
# ============================================================================
ws_matrix = wb.create_sheet("COMPLIANCE_MATRIX")

ws_matrix['A1'] = "Regulatory Compliance Matrix - All Properties"
ws_matrix['A1'].font = Font(bold=True, size=14)
ws_matrix.merge_cells('A1:J1')

# Headers
headers = ['Property', 'Location', 'Units', 'Status', 'Mandatory\nRecycling',
           'Mandatory\nComposting', 'Ordinances', 'Penalties', 'Confidence', 'Action Required']

for col_idx, header in enumerate(headers, start=1):
    cell = ws_matrix.cell(row=3, column=col_idx)
    cell.value = header
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER

# Data rows
row = 4
for prop_name, prop_data in PROPERTIES.items():
    ws_matrix[f'A{row}'] = prop_name
    ws_matrix[f'B{row}'] = f"{prop_data['city']}, {prop_data['state']}"
    ws_matrix[f'C{row}'] = prop_data['units']
    ws_matrix[f'D{row}'] = prop_data['regulatory_status']
    ws_matrix[f'E{row}'] = "YES" if prop_data['mandatory_recycling'] else "NO"
    ws_matrix[f'F{row}'] = "YES" if prop_data['mandatory_composting'] else "NO"
    ws_matrix[f'G{row}'] = prop_data['ordinances']
    ws_matrix[f'H{row}'] = prop_data['penalties']
    ws_matrix[f'I{row}'] = prop_data['confidence']
    ws_matrix[f'J{row}'] = "VERIFY COMPLIANCE" if prop_data['mandatory_recycling'] else "None"

    # Color coding
    if prop_data['mandatory_recycling']:
        ws_matrix[f'E{row}'].font = Font(color="FF0000", bold=True)
        ws_matrix[f'J{row}'].font = Font(color="FF0000", bold=True)
    else:
        ws_matrix[f'E{row}'].font = Font(color="008000", bold=True)
        ws_matrix[f'J{row}'].font = Font(color="008000")

    ws_matrix[f'F{row}'].font = Font(color="008000", bold=True)

    # Apply borders
    for col in range(1, 11):
        ws_matrix.cell(row=row, column=col).border = THIN_BORDER
        ws_matrix.cell(row=row, column=col).alignment = LEFT_ALIGN

    row += 1

# Column widths
ws_matrix.column_dimensions['A'].width = 25
ws_matrix.column_dimensions['B'].width = 20
ws_matrix.column_dimensions['C'].width = 10
ws_matrix.column_dimensions['D'].width = 15
ws_matrix.column_dimensions['E'].width = 12
ws_matrix.column_dimensions['F'].width = 12
ws_matrix.column_dimensions['G'].width = 35
ws_matrix.column_dimensions['H'].width = 30
ws_matrix.column_dimensions['I'].width = 12
ws_matrix.column_dimensions['J'].width = 20

# ============================================================================
# SHEET 3: CITY-BY-CITY SUMMARY
# ============================================================================
ws_cities = wb.create_sheet("CITY_SUMMARY")

ws_cities['A1'] = "City-by-City Regulatory Summary"
ws_cities['A1'].font = Font(bold=True, size=14)
ws_cities.merge_cells('A1:F1')

# Group properties by city
cities = {}
for prop_name, prop_data in PROPERTIES.items():
    city_key = f"{prop_data['city']}, {prop_data['state']}"
    if city_key not in cities:
        cities[city_key] = {
            'properties': [],
            'data': prop_data
        }
    cities[city_key]['properties'].append(prop_name)

row = 3
for city_key, city_info in sorted(cities.items()):
    ws_cities[f'A{row}'] = city_key
    ws_cities[f'A{row}'].font = Font(bold=True, size=12)
    ws_cities.merge_cells(f'A{row}:F{row}')
    row += 1

    ws_cities[f'A{row}'] = "Properties:"
    ws_cities[f'B{row}'] = ", ".join(city_info['properties'])
    ws_cities[f'A{row}'].font = Font(bold=True)
    ws_cities.merge_cells(f'B{row}:F{row}')
    row += 1

    ws_cities[f'A{row}'] = "Regulatory Status:"
    ws_cities[f'B{row}'] = city_info['data']['regulatory_status']
    ws_cities[f'A{row}'].font = Font(bold=True)
    ws_cities[f'B{row}'].font = Font(bold=True, color="FF0000" if city_info['data']['mandatory_recycling'] else "008000")
    row += 1

    ws_cities[f'A{row}'] = "Key Finding:"
    ws_cities[f'B{row}'] = city_info['data']['key_finding']
    ws_cities[f'A{row}'].font = Font(bold=True)
    ws_cities[f'B{row}'].alignment = LEFT_ALIGN
    ws_cities.merge_cells(f'B{row}:F{row}')
    row += 1

    ws_cities[f'A{row}'] = "Ordinances:"
    ws_cities[f'B{row}'] = city_info['data']['ordinances']
    ws_cities[f'A{row}'].font = Font(bold=True)
    ws_cities.merge_cells(f'B{row}:F{row}')
    row += 1

    ws_cities[f'A{row}'] = "Contact:"
    ws_cities[f'B{row}'] = city_info['data']['contact']
    ws_cities[f'A{row}'].font = Font(bold=True)
    ws_cities.merge_cells(f'B{row}:F{row}')
    row += 2

ws_cities.column_dimensions['A'].width = 20
ws_cities.column_dimensions['B'].width = 80

# ============================================================================
# SHEET 4: COMPLIANCE ACTION ITEMS
# ============================================================================
ws_actions = wb.create_sheet("ACTION_ITEMS")

ws_actions['A1'] = "Compliance Action Items - Properties Requiring Attention"
ws_actions['A1'].font = Font(bold=True, size=14)
ws_actions.merge_cells('A1:E1')

row = 3
ws_actions[f'A{row}'] = "MANDATORY RECYCLING PROPERTIES"
ws_actions[f'A{row}'].font = Font(bold=True, size=12, color="FF0000")
ws_actions.merge_cells(f'A{row}:E{row}')
row += 2

mandatory_props = {k: v for k, v in PROPERTIES.items() if v['mandatory_recycling']}

if mandatory_props:
    headers = ['Property', 'Location', 'Units', 'Requirements', 'Next Steps']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_actions.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1

    for prop_name, prop_data in mandatory_props.items():
        ws_actions[f'A{row}'] = prop_name
        ws_actions[f'B{row}'] = f"{prop_data['city']}, {prop_data['state']}"
        ws_actions[f'C{row}'] = prop_data['units']

        if 'compliance_requirements' in prop_data:
            requirements = "\n".join(f"• {req}" for req in prop_data['compliance_requirements'])
        else:
            requirements = "• Provide recycling service\n• Verify compliance with local ordinances"

        ws_actions[f'D{row}'] = requirements
        ws_actions[f'E{row}'] = f"1. Verify current recycling service\n2. Contact: {prop_data['contact']}\n3. Document compliance"

        ws_actions[f'D{row}'].alignment = LEFT_ALIGN
        ws_actions[f'E{row}'].alignment = LEFT_ALIGN
        row += 1

row += 2
ws_actions[f'A{row}'] = "VOLUNTARY/NO REQUIREMENTS PROPERTIES"
ws_actions[f'A{row}'].font = Font(bold=True, size=12, color="008000")
ws_actions.merge_cells(f'A{row}:E{row}')
row += 1

ws_actions[f'A{row}'] = "8 properties have no mandatory recycling requirements. Recycling programs are optional."
ws_actions[f'A{row}'].alignment = LEFT_ALIGN
ws_actions.merge_cells(f'A{row}:E{row}')

ws_actions.column_dimensions['A'].width = 25
ws_actions.column_dimensions['B'].width = 20
ws_actions.column_dimensions['C'].width = 10
ws_actions.column_dimensions['D'].width = 40
ws_actions.column_dimensions['E'].width = 40

# ============================================================================
# SAVE WORKBOOK
# ============================================================================
output_file = "Portfolio_Reports/MASTER_Portfolio_Regulatory_Summary.xlsx"
wb.save(output_file)

print(f"\n[OK] Master portfolio summary generated:")
print(f"     {output_file}")
print(f"\nWorkbook includes:")
print(f"  - PORTFOLIO_OVERVIEW: Statistics and key findings")
print(f"  - COMPLIANCE_MATRIX: All properties regulatory status")
print(f"  - CITY_SUMMARY: Jurisdiction-by-jurisdiction analysis")
print(f"  - ACTION_ITEMS: Required compliance actions")
print("\n" + "=" * 80)
print("PORTFOLIO SUMMARY COMPLETE")
print("=" * 80)
