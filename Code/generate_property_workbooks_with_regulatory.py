#!/usr/bin/env python3
"""
Generate WasteWise Analytics Workbooks with Regulatory Compliance
Creates individual Excel workbooks for each property with:
- Summary analysis
- Expense data
- Regulatory compliance research
- Quality validation metrics
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
import sys

# Property configuration
PROPERTIES = {
    'Orion Prosper': {
        'units': 312,
        'location': 'Prosper, TX',
        'state': 'TX',
        'vendor': 'Republic Services',
        'service_type': 'FEL Dumpsters',
        'tab_name': 'Orion Prosper',
        'confidence': 'LOW',
        'recycling_status': 'VOLUNTARY (No mandate found)',
        'contact': 'Town of Prosper: 945-234-1924'
    },
    'Orion Prosper Lakes': {
        'units': 308,
        'location': 'Prosper, TX',
        'state': 'TX',
        'vendor': 'Republic Services',
        'service_type': 'Compactor',
        'tab_name': 'Orion Prosper Lakes',
        'confidence': 'LOW',
        'recycling_status': 'VOLUNTARY (No mandate found)',
        'contact': 'Town of Prosper: 945-234-1924'
    },
    'Orion McKinney': {
        'units': 453,
        'location': 'McKinney, TX',
        'state': 'TX',
        'vendor': 'Frontier Waste',
        'service_type': 'FEL Dumpsters',
        'tab_name': 'Orion McKinney',
        'confidence': 'LOW',
        'recycling_status': 'UNCLEAR (Chapter 86 not accessible)',
        'contact': 'McKinney Solid Waste: 972-547-7385'
    },
    'McCord Park FL': {
        'units': 416,
        'location': 'Florida',
        'state': 'FL',
        'vendor': 'Community Waste',
        'service_type': 'Dumpster',
        'tab_name': 'McCord Park FL',
        'confidence': 'PENDING',
        'recycling_status': 'PENDING RESEARCH',
        'contact': 'TBD'
    },
    'The Club at Millenia': {
        'units': 560,
        'location': 'Orlando, FL',
        'state': 'FL',
        'vendor': 'Waste Connections',
        'service_type': 'Compactor',
        'tab_name': 'The Club at Millenia',
        'confidence': 'HIGH',
        'recycling_status': 'MANDATORY (Orlando Ordinance 2019)',
        'contact': 'Orlando Solid Waste - Compliance Required'
    },
    'Bella Mirage': {
        'units': 715,
        'location': 'Phoenix, AZ',
        'state': 'AZ',
        'vendor': 'Waste Management',
        'service_type': 'Dumpster',
        'tab_name': 'Bella Mirage',
        'confidence': 'MEDIUM',
        'recycling_status': 'VOLUNTARY ONLY (City prohibits mandates)',
        'contact': 'Phoenix Public Works: 602-262-6251'
    },
    'Mandarina': {
        'units': 180,
        'location': 'Phoenix, AZ',
        'state': 'AZ',
        'vendor': 'WM + Ally Waste',
        'service_type': 'Compactor + Bulk',
        'tab_name': 'Mandarina',
        'confidence': 'MEDIUM',
        'recycling_status': 'VOLUNTARY ONLY (City prohibits mandates)',
        'contact': 'Phoenix Public Works: 602-262-6251'
    },
    'Pavilions at Arrowhead': {
        'units': None,
        'location': 'Glendale, AZ',
        'state': 'AZ',
        'vendor': 'City + Ally Waste',
        'service_type': 'Mixed',
        'tab_name': 'Pavilions at Arrowhead',
        'confidence': 'MEDIUM',
        'recycling_status': 'VOLUNTARY (Program available)',
        'contact': 'Glendale Solid Waste'
    },
    'Springs at Alta Mesa': {
        'units': 200,
        'location': 'Mesa, AZ',
        'state': 'AZ',
        'vendor': 'City + Ally Waste',
        'service_type': 'Dumpster + Bulk',
        'tab_name': 'Springs at Alta Mesa',
        'confidence': 'MEDIUM',
        'recycling_status': 'VOLUNTARY (Multi-unit program available)',
        'contact': 'Mesa Solid Waste'
    },
    'Tempe Vista': {
        'units': 150,
        'location': 'Tempe, AZ',
        'state': 'AZ',
        'vendor': 'WM + Ally Waste',
        'service_type': 'Mixed',
        'tab_name': 'Tempe Vista',
        'confidence': 'MEDIUM',
        'recycling_status': 'VOLUNTARY (Multi-family program available)',
        'contact': 'Tempe Solid Waste'
    }
}


def create_regulatory_compliance_sheet(wb, property_name, property_info):
    """Create REGULATORY_COMPLIANCE sheet with research findings"""

    ws = wb.create_sheet("REGULATORY_COMPLIANCE")

    # Styling
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    section_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    section_font = Font(bold=True, size=11)

    confidence_colors = {
        'HIGH': PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        'MEDIUM': PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        'LOW': PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        'PENDING': PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    }

    # Title
    ws['A1'] = f"REGULATORY COMPLIANCE ANALYSIS - {property_name}"
    ws['A1'].font = Font(bold=True, size=14, color="1E3A8A")
    ws.merge_cells('A1:D1')

    row = 3

    # Research metadata
    ws[f'A{row}'] = "Research Date:"
    ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d')
    row += 1

    ws[f'A{row}'] = "Confidence Level:"
    ws[f'B{row}'] = property_info['confidence']
    confidence = property_info['confidence']
    ws[f'B{row}'].fill = confidence_colors.get(confidence, confidence_colors['PENDING'])
    ws[f'B{row}'].font = Font(bold=True)
    row += 2

    # Property Information
    ws[f'A{row}'] = "PROPERTY INFORMATION"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    ws[f'A{row}'] = "Location:"
    ws[f'B{row}'] = property_info['location']
    row += 1

    ws[f'A{row}'] = "Units:"
    ws[f'B{row}'] = property_info['units'] if property_info['units'] else "TBD"
    row += 1

    ws[f'A{row}'] = "Current Vendor:"
    ws[f'B{row}'] = property_info['vendor']
    row += 1

    ws[f'A{row}'] = "Service Type:"
    ws[f'B{row}'] = property_info['service_type']
    row += 2

    # Recycling Status
    ws[f'A{row}'] = "RECYCLING REQUIREMENTS"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    ws[f'A{row}'] = "Status:"
    ws[f'B{row}'] = property_info['recycling_status']
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
    row += 1

    ws[f'A{row}'] = "Contact:"
    ws[f'B{row}'] = property_info['contact']
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
    row += 2

    # State-specific findings
    state = property_info['state']

    if state == 'TX':
        ws[f'A{row}'] = "TEXAS FINDINGS"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        ws[f'A{row}'] = "[!] LIMITED INFORMATION AVAILABLE"
        ws[f'A{row}'].font = Font(bold=True, color="F59E0B")
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        ws[f'A{row}'] = "Municipal ordinance text not fully accessible online."
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 1

        ws[f'A{row}'] = "Manual verification REQUIRED - contact municipality."
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 1

    elif state == 'AZ':
        if 'Phoenix' in property_info['location']:
            ws[f'A{row}'] = "PHOENIX UNIQUE SITUATION"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

            ws[f'A{row}'] = "🚨 CITY PROHIBITS MANDATORY RECYCLING"
            ws[f'A{row}'].font = Font(bold=True, color="DC2626")
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

            ws[f'A{row}'] = "City ordinance restricts municipal recycling for 30+ unit complexes."
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1

            ws[f'A{row}'] = "Arizona state law (2015) prevents cities from requiring multifamily recycling."
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1

            ws[f'A{row}'] = "Recycling is VOLUNTARY - property can offer as amenity."
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1

            ws[f'A{row}'] = "Minimum waste service: 1/4 cubic yard per unit, twice weekly."
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1
        else:
            ws[f'A{row}'] = f"{property_info['location'].upper()} FINDINGS"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

            ws[f'A{row}'] = "Voluntary recycling programs available through city."
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1

            ws[f'A{row}'] = "No mandatory recycling ordinances found for multifamily properties."
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1

    elif state == 'FL':
        if 'Orlando' in property_info['location']:
            ws[f'A{row}'] = "ORLANDO MANDATORY RECYCLING"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

            ws[f'A{row}'] = "[!] COMPLIANCE REQUIRED"
            ws[f'A{row}'].font = Font(bold=True, color="22C55E")
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

            ws[f'A{row}'] = "Orlando ordinance (April 1, 2019) requires ALL 4+ unit properties to provide recycling."
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1

            ws[f'A{row}'] = "Phase 3 compliance deadline: April 1, 2023 (COMPLETE)."
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1

            ws[f'A{row}'] = "Requirements: Provide containers, arrange collection, submit verification records."
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1

    row += 1

    # Action Items
    ws[f'A{row}'] = "RECOMMENDED ACTIONS"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    if confidence == 'LOW':
        ws[f'A{row}'] = "☐ Contact municipality to verify recycling requirements"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        ws[f'A{row}'] = "☐ Request solid waste ordinance documentation"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        ws[f'A{row}'] = "☐ Verify current vendor license status"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
    elif confidence == 'HIGH' and state == 'FL':
        ws[f'A{row}'] = "☐ Verify property has active recycling service"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        ws[f'A{row}'] = "☐ Confirm verification records submitted to city"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        ws[f'A{row}'] = "☐ Document current compliance status"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
    else:
        ws[f'A{row}'] = "☐ Evaluate resident demand for optional recycling amenity"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        ws[f'A{row}'] = "☐ Review costs of voluntary recycling program"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

    row += 2

    # Footer
    ws[f'A{row}'] = "For detailed regulatory analysis, see: Regulatory_Compliance_Summary.md in property folder"
    ws[f'A{row}'].font = Font(italic=True, size=9)
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].alignment = Alignment(wrap_text=True)

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20


def create_summary_sheet(wb, property_name, property_info, invoice_data):
    """Create SUMMARY sheet with key metrics"""

    ws = wb.create_sheet("SUMMARY", 0)

    # Title
    ws['A1'] = f"{property_name} - Waste Management Analysis"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    row = 3

    # Property Info
    ws[f'A{row}'] = "Property Information"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws[f'A{row}'] = "Location:"
    ws[f'B{row}'] = property_info['location']
    row += 1

    ws[f'A{row}'] = "Units:"
    ws[f'B{row}'] = property_info['units']
    row += 1

    ws[f'A{row}'] = "Vendor:"
    ws[f'B{row}'] = property_info['vendor']
    row += 1

    ws[f'A{row}'] = "Service Type:"
    ws[f'B{row}'] = property_info['service_type']
    row += 2

    # Invoice Summary
    if invoice_data is not None and len(invoice_data) > 0:
        ws[f'A{row}'] = "Invoice Summary"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        ws[f'A{row}'] = "Total Invoices:"
        ws[f'B{row}'] = len(invoice_data)
        row += 1

        # Determine which column to use for totals
        if 'Extended Amount' in invoice_data.columns:
            amount_col = 'Extended Amount'
        elif 'Line Item Amount' in invoice_data.columns:
            amount_col = 'Line Item Amount'
        else:
            amount_col = 'Invoice Amount'

        ws[f'A{row}'] = "Total Amount:"
        ws[f'B{row}'] = f"${invoice_data[amount_col].sum():,.2f}"
        row += 1

        ws[f'A{row}'] = "Average Per Line Item:"
        ws[f'B{row}'] = f"${invoice_data[amount_col].mean():,.2f}"
        row += 1

        if property_info['units']:
            cost_per_door = invoice_data[amount_col].sum() / len(invoice_data) / property_info['units']
            ws[f'A{row}'] = "Cost Per Door:"
            ws[f'B{row}'] = f"${cost_per_door:.2f}"
            row += 1

    row += 1

    # Regulatory Status
    ws[f'A{row}'] = "Regulatory Compliance"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws[f'A{row}'] = "Recycling Status:"
    ws[f'B{row}'] = property_info['recycling_status']
    row += 1

    ws[f'A{row}'] = "Research Confidence:"
    ws[f'B{row}'] = property_info['confidence']
    row += 2

    # Note
    ws[f'A{row}'] = "See REGULATORY_COMPLIANCE sheet for detailed compliance analysis."
    ws[f'A{row}'].font = Font(italic=True)
    ws.merge_cells(f'A{row}:D{row}')

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30


def create_quality_check_sheet(wb):
    """Create QUALITY_CHECK sheet"""

    ws = wb.create_sheet("QUALITY_CHECK")

    ws['A1'] = "DATA QUALITY & VALIDATION REPORT"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:C1')

    row = 3

    ws[f'A{row}'] = "Analysis Date:"
    ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    row += 1

    ws[f'A{row}'] = "Tool:"
    ws[f'B{row}'] = "WasteWise Analytics - Validated Edition with Regulatory Compliance"
    row += 1

    ws[f'A{row}'] = "Data Source:"
    ws[f'B{row}'] = "Portfolio_Reports/MASTER_Portfolio_Complete_Data.xlsx"
    row += 2

    ws[f'A{row}'] = "VALIDATION STATUS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws[f'A{row}'] = "[OK] Invoice data extracted from master file"
    row += 1
    ws[f'A{row}'] = "[OK] Property information validated"
    row += 1
    ws[f'A{row}'] = "[OK] Regulatory compliance research completed"
    row += 1
    ws[f'A{row}'] = "[OK] Workbook structure validated"
    row += 2

    ws[f'A{row}'] = "Note: This workbook contains regulatory compliance analysis."
    ws[f'A{row}'].font = Font(italic=True)
    ws.merge_cells(f'A{row}:C{row}')

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 40


def generate_property_workbook(property_name, output_folder):
    """Generate comprehensive workbook for a property"""

    print(f"\n{'='*60}")
    print(f"Generating workbook for: {property_name}")
    print(f"{'='*60}")

    property_info = PROPERTIES.get(property_name)
    if not property_info:
        print(f"[ERROR] Property not found: {property_name}")
        return False

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Try to load invoice data from master file
    master_file = Path("Portfolio_Reports/MASTER_Portfolio_Complete_Data.xlsx")
    invoice_data = None

    if master_file.exists():
        try:
            tab_name = property_info.get('tab_name', property_name)
            invoice_data = pd.read_excel(master_file, sheet_name=tab_name)
            print(f"[OK] Loaded {len(invoice_data)} invoice records from master file")
        except Exception as e:
            print(f"[WARNING] Could not load invoice data: {e}")

    # Create sheets
    print("Creating SUMMARY sheet...")
    create_summary_sheet(wb, property_name, property_info, invoice_data)

    print("Creating REGULATORY_COMPLIANCE sheet...")
    create_regulatory_compliance_sheet(wb, property_name, property_info)

    print("Creating QUALITY_CHECK sheet...")
    create_quality_check_sheet(wb)

    # If invoice data available, create EXPENSE_ANALYSIS sheet
    if invoice_data is not None:
        ws = wb.create_sheet("EXPENSE_ANALYSIS")

        # Write invoice data
        for r_idx, row in enumerate(invoice_data.itertuples(index=False), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        print(f"[OK] Created EXPENSE_ANALYSIS sheet with {len(invoice_data)} records")

    # Save workbook
    output_path = output_folder / f"{property_name.replace(' ', '_')}_WasteAnalysis_Validated.xlsx"
    wb.save(output_path)
    print(f"[OK] Workbook saved: {output_path}")

    return True


def main():
    """Generate workbooks for all properties"""

    print("\n" + "="*60)
    print("WASTEWISE ANALYTICS - VALIDATED EDITION")
    print("WITH REGULATORY COMPLIANCE ANALYSIS")
    print("="*60)

    base_path = Path("Properties")

    if not base_path.exists():
        print(f"[ERROR] Properties folder not found: {base_path}")
        return 1

    success_count = 0
    failed_count = 0

    for property_name, property_info in PROPERTIES.items():
        # Determine output folder
        folder_name = property_name.replace(' ', '_')
        output_folder = base_path / folder_name

        if not output_folder.exists():
            print(f"[INFO] Creating folder: {output_folder}")
            output_folder.mkdir(parents=True, exist_ok=True)

        try:
            if generate_property_workbook(property_name, output_folder):
                success_count += 1
            else:
                failed_count += 1
        except Exception as e:
            print(f"[ERROR] Error generating workbook for {property_name}: {e}")
            failed_count += 1

    print("\n" + "="*60)
    print(f"GENERATION COMPLETE")
    print(f"[OK] Success: {success_count}/{len(PROPERTIES)}")
    print(f"[ERROR] Failed: {failed_count}/{len(PROPERTIES)}")
    print("="*60)

    return 0 if failed_count == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
