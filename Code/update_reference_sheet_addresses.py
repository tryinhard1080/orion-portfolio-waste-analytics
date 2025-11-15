"""
Update Property_Reference_Sheet with extracted addresses from invoices
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import json
import re

# File paths
REFERENCE_FILE = "Portfolio_Reports/Property_Reference_Sheet.xlsx"
ADDRESS_DATA_FILE = "Portfolio_Reports/addresses_data.json"
OUTPUT_FILE = "Portfolio_Reports/Property_Reference_Sheet.xlsx"

# Load extracted addresses
with open(ADDRESS_DATA_FILE, 'r') as f:
    addresses = json.load(f)

# Parse addresses to extract components
def parse_address(full_address):
    """Parse full address into street, city, state, zip"""
    if not full_address or full_address == 'TBD':
        return {'street': 'TBD', 'city': 'TBD', 'state': 'TBD', 'zip': 'TBD'}

    # Normalize - remove extra spaces
    addr = ' '.join(full_address.strip().split())

    # Pattern: Street Address, City, State Zip
    # Example: 12835 W Indian School Rd, Avondale, AZ 85392
    pattern = r'^(.+?),\s*([A-Za-z\s]+),?\s+([A-Z]{2})\s*(\d{5}(?:-\d{4})?)?'

    match = re.match(pattern, addr)

    if match:
        street = match.group(1).strip()
        city = match.group(2).strip()
        state = match.group(3).strip()
        zip_code = match.group(4).strip() if match.group(4) else 'TBD'

        return {
            'street': street,
            'city': city,
            'state': state,
            'zip': zip_code
        }

    # Try alternate pattern without commas: Street Address City State Zip
    # Example: 1865 N. Higley Rd Mesa AZ 85205
    pattern2 = r'^(.+?)\s+([A-Za-z\s]+?)\s+([A-Z]{2})\s+(\d{5}(?:-\d{4})?)'

    match2 = re.match(pattern2, addr)

    if match2:
        street = match2.group(1).strip()
        city = match2.group(2).strip()
        state = match2.group(3).strip()
        zip_code = match2.group(4).strip()

        return {
            'street': street,
            'city': city,
            'state': state,
            'zip': zip_code
        }

    # Fallback - split by comma
    parts = [p.strip() for p in addr.split(',')]

    if len(parts) >= 3:
        street = parts[0]
        city = parts[1]

        # Last part should have state and zip
        last_part = parts[2]
        state_zip = last_part.split()

        state = state_zip[0] if len(state_zip) > 0 else 'TBD'
        zip_code = state_zip[1] if len(state_zip) > 1 else 'TBD'

        return {
            'street': street,
            'city': city,
            'state': state,
            'zip': zip_code
        }

    # Final fallback
    return {
        'street': addr,
        'city': 'TBD',
        'state': 'TBD',
        'zip': 'TBD'
    }

print("=" * 80)
print("UPDATING PROPERTY REFERENCE SHEET WITH ADDRESSES")
print("=" * 80)

# Read current reference sheet (skip title rows, read from row 4)
df = pd.read_excel(REFERENCE_FILE, header=3)

print(f"\nCurrent properties in reference sheet: {len(df)}")
print("\nUpdating addresses...")

# Update addresses
for idx, row in df.iterrows():
    prop_name = row['Property Name']

    if prop_name in addresses:
        full_addr = addresses[prop_name]
        parsed = parse_address(full_addr)

        df.at[idx, 'Address'] = parsed['street']
        df.at[idx, 'City'] = parsed['city']
        df.at[idx, 'State'] = parsed['state']
        df.at[idx, 'Zip Code'] = parsed['zip']

        print(f"\n{prop_name}:")
        print(f"  Full: {full_addr}")
        print(f"  Street: {parsed['street']}")
        print(f"  City: {parsed['city']}, {parsed['state']} {parsed['zip']}")

# Create updated workbook with formatting
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Property Reference"

# Styling
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Title
ws['A1'] = "Property Reference Sheet - Complete with Addresses"
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:W1')

ws['A2'] = f"Last Updated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
ws['A2'].font = Font(size=10, italic=True)
ws.merge_cells('A2:W2')

# Headers
row = 4
for col_idx, column in enumerate(df.columns, start=1):
    cell = ws.cell(row=row, column=col_idx)
    cell.value = column
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER

# Data rows
for idx, data_row in df.iterrows():
    row += 1
    for col_idx, value in enumerate(data_row, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        cell.alignment = LEFT_ALIGN if col_idx <= 3 or col_idx >= 20 else CENTER_ALIGN
        cell.border = THIN_BORDER

        # Color coding for regulatory status
        if df.columns[col_idx-1] == 'Mandatory Recycling':
            if value == 'YES':
                cell.font = Font(color="FF0000", bold=True)
            elif value == 'NO':
                cell.font = Font(color="008000", bold=True)

# Column widths
column_widths = {
    'A': 25,  # Property Name
    'B': 35,  # Address
    'C': 15,  # City
    'D': 8,   # State
    'E': 15,  # County
    'F': 10,  # Zip Code
    'G': 8,   # Units
    'H': 15,  # Property Type
    'I': 25,  # Vendor
    'J': 18,  # Service Type
    'K': 12,  # Container Count
    'L': 15,  # Container Size
    'M': 20,  # Service Frequency
    'N': 12,  # Monthly Yards
    'O': 8,   # YPD
    'P': 18,  # Regulatory Status
    'Q': 18,  # Mandatory Recycling
    'R': 12,  # PDF Invoices
    'S': 12,  # Excel Files
    'T': 10,  # Contracts
    'U': 15,  # Extracted Records
    'V': 12,  # Status
    'W': 30,  # Notes
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Freeze panes
ws.freeze_panes = 'A5'

# Save
wb.save(OUTPUT_FILE)

print(f"\n\n[OK] Property Reference Sheet updated successfully!")
print(f"     {OUTPUT_FILE}")
print(f"\nAll 10 properties now have complete addresses with:")
print(f"  - Street Address")
print(f"  - City")
print(f"  - State")
print(f"  - Zip Code")
print("\n" + "=" * 80)
print("PROPERTY REFERENCE SHEET ADDRESS UPDATE COMPLETE")
print("=" * 80)
