"""
Add REGULATORY_COMPLIANCE tab to existing WasteAnalysis_Validated workbooks
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from datetime import datetime


# Property configuration
PROPERTIES = {
    'OrionProsper': {
        'name': 'Orion Prosper',
        'units': 312,
        'location': 'Prosper, TX',
        'state': 'TX',
        'confidence': 'LOW',
        'recycling_status': 'VERIFICATION REQUIRED',
        'contact': 'Town of Prosper: 945-234-1924',
        'summary_file': 'Properties/Orion_Prosper/Regulatory_Compliance_Summary.md'
    },
    'OrionProsperLakes': {
        'name': 'Orion Prosper Lakes',
        'units': 308,
        'location': 'Prosper, TX',
        'state': 'TX',
        'confidence': 'LOW',
        'recycling_status': 'VERIFICATION REQUIRED',
        'contact': 'Town of Prosper: 945-234-1924',
        'summary_file': 'Properties/Orion_Prosper_Lakes/Regulatory_Compliance_Summary.md'
    },
    'OrionMcKinney': {
        'name': 'Orion McKinney',
        'units': 453,
        'location': 'McKinney, TX',
        'state': 'TX',
        'confidence': 'LOW',
        'recycling_status': 'VERIFICATION REQUIRED',
        'contact': 'McKinney Solid Waste: 972-547-7385',
        'summary_file': 'Properties/Orion_McKinney/Regulatory_Compliance_Summary.md'
    },
    'McCordParkFL': {
        'name': 'McCord Park FL',
        'units': 416,
        'location': 'Florida',
        'state': 'FL',
        'confidence': 'PENDING',
        'recycling_status': 'RESEARCH NEEDED',
        'contact': 'TBD',
        'summary_file': None
    },
    'TheClubAtMillenia': {
        'name': 'The Club at Millenia',
        'units': 560,
        'location': 'Orlando, FL',
        'state': 'FL',
        'confidence': 'HIGH',
        'recycling_status': 'MANDATORY (City ordinance - April 2019)',
        'contact': 'Orlando Solid Waste',
        'summary_file': None  # Research completed but summary pending
    },
    'BellaMirage': {
        'name': 'Bella Mirage',
        'units': 715,
        'location': 'Phoenix, AZ',
        'state': 'AZ',
        'confidence': 'MEDIUM',
        'recycling_status': 'VOLUNTARY ONLY (State law prohibits mandates)',
        'contact': 'Phoenix Public Works: 602-262-6251',
        'summary_file': 'Properties/Bella_Mirage/Regulatory_Compliance_Summary.md'
    },
    'Mandarina': {
        'name': 'Mandarina',
        'units': 180,
        'location': 'Phoenix, AZ',
        'state': 'AZ',
        'confidence': 'MEDIUM',
        'recycling_status': 'VOLUNTARY ONLY (State law prohibits mandates)',
        'contact': 'Phoenix Public Works: 602-262-6251',
        'summary_file': 'Properties/Mandarina/Regulatory_Compliance_Summary.md'
    },
    'PavilionsAtArrowhead': {
        'name': 'Pavilions at Arrowhead',
        'units': None,
        'location': 'Glendale, AZ',
        'state': 'AZ',
        'confidence': 'MEDIUM',
        'recycling_status': 'VOLUNTARY (Program available)',
        'contact': 'Glendale Solid Waste',
        'summary_file': None
    },
    'SpringsAtAltaMesa': {
        'name': 'Springs at Alta Mesa',
        'units': 200,
        'location': 'Mesa, AZ',
        'state': 'AZ',
        'confidence': 'MEDIUM',
        'recycling_status': 'VOLUNTARY (Multi-unit program available)',
        'contact': 'Mesa Solid Waste',
        'summary_file': None
    },
    'TempeVista': {
        'name': 'Tempe Vista',
        'units': 150,
        'location': 'Tempe, AZ',
        'state': 'AZ',
        'confidence': 'MEDIUM',
        'recycling_status': 'VOLUNTARY (Multi-family program available)',
        'contact': 'Tempe Solid Waste',
        'summary_file': None
    }
}


def create_regulatory_compliance_sheet(wb, property_name, property_info):
    """Create REGULATORY_COMPLIANCE sheet"""

    # Check if sheet already exists
    if 'REGULATORY_COMPLIANCE' in wb.sheetnames:
        print(f"  [INFO] REGULATORY_COMPLIANCE sheet already exists, removing old version")
        wb.remove(wb['REGULATORY_COMPLIANCE'])

    ws = wb.create_sheet("REGULATORY_COMPLIANCE")

    # Styling
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    section_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    section_font = Font(bold=True, size=11)

    confidence_colors = {
        'HIGH': PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        'MEDIUM': PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        'LOW': PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        'PENDING': PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    }

    # Title
    ws['A1'] = f"REGULATORY COMPLIANCE ANALYSIS - {property_name}"
    ws['A1'].font = Font(bold=True, size=14, color="1E3A8A")
    ws.merge_cells('A1:D1')

    row = 3

    # Research metadata
    ws[f'A{row}'] = "Research Date:"
    ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d')
    row += 1

    ws[f'A{row}'] = "Confidence Level:"
    ws[f'B{row}'] = property_info['confidence']
    confidence = property_info['confidence']
    ws[f'B{row}'].fill = confidence_colors.get(confidence, confidence_colors['PENDING'])
    ws[f'B{row}'].font = Font(bold=True)
    row += 2

    # Property Information
    ws[f'A{row}'] = "PROPERTY INFORMATION"
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    ws[f'A{row}'] = "Location:"
    ws[f'B{row}'] = property_info['location']
    row += 1

    ws[f'A{row}'] = "State:"
    ws[f'B{row}'] = property_info['state']
    row += 1

    ws[f'A{row}'] = "Units:"
    ws[f'B{row}'] = property_info['units'] if property_info['units'] else 'TBD'
    row += 2

    # Recycling Requirements
    ws[f'A{row}'] = "RECYCLING REQUIREMENTS"
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    ws[f'A{row}'] = "Status:"
    ws[f'B{row}'] = property_info['recycling_status']
    ws[f'B{row}'].font = Font(bold=True)
    ws.merge_cells(f'B{row}:D{row}')
    row += 2

    # State-specific findings
    if property_info['state'] == 'TX':
        ws[f'A{row}'] = "TEXAS FINDINGS"
        ws[f'A{row}'].fill = section_fill
        ws[f'A{row}'].font = section_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        ws[f'A{row}'] = "[!] LIMITED INFORMATION AVAILABLE"
        ws[f'A{row}'].font = Font(bold=True, color="F59E0B")
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        ws[f'A{row}'] = "Municipal ordinances not publicly accessible online"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        ws[f'A{row}'] = "Manual verification required with city solid waste department"
        ws.merge_cells(f'A{row}:D{row}')
        row += 2

        ws[f'A{row}'] = "Contact for Verification:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = property_info['contact']
        ws.merge_cells(f'A{row}:D{row}')
        row += 2

    elif property_info['state'] == 'FL':
        if 'Orlando' in property_info['location']:
            ws[f'A{row}'] = "ORLANDO, FL - MANDATORY RECYCLING"
            ws[f'A{row}'].fill = section_fill
            ws[f'A{row}'].font = section_font
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

            ws[f'A{row}'] = "[!] COMPLIANCE REQUIRED"
            ws[f'A{row}'].font = Font(bold=True, color="22C55E")
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

            ws[f'A{row}'] = "City ordinance (April 1, 2019) requires ALL properties with 4+ units to provide recycling"
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

            ws[f'A{row}'] = "Phase 3 implementation complete (April 1, 2023)"
            ws.merge_cells(f'A{row}:D{row}')
            row += 2

            ws[f'A{row}'] = "Required Actions:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = "1. Provide recycling containers for residents"
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            ws[f'A{row}'] = "2. Arrange collection service with licensed provider"
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            ws[f'A{row}'] = "3. Submit verification records to city"
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            ws[f'A{row}'] = "4. Maintain ongoing compliance documentation"
            ws.merge_cells(f'A{row}:D{row}')
            row += 2

        else:
            ws[f'A{row}'] = "FLORIDA PROPERTY - RESEARCH NEEDED"
            ws[f'A{row}'].fill = section_fill
            ws[f'A{row}'].font = section_font
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

            ws[f'A{row}'] = "Specific city/county regulations require research"
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

            ws[f'A{row}'] = "Contact local solid waste department for requirements"
            ws.merge_cells(f'A{row}:D{row}')
            row += 2

    elif property_info['state'] == 'AZ':
        if 'Phoenix' in property_info['location']:
            ws[f'A{row}'] = "PHOENIX, AZ - UNIQUE REGULATORY SITUATION"
            ws[f'A{row}'].fill = section_fill
            ws[f'A{row}'].font = section_font
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

            ws[f'A{row}'] = "[!] NO RECYCLING MANDATES PERMITTED"
            ws[f'A{row}'].font = Font(bold=True, color="3B82F6")
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

            ws[f'A{row}'] = "Phoenix City Code Section 27-21(c): Prohibits city from offering recycling at 30+ unit complexes"
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

            ws[f'A{row}'] = "Arizona state law (2015): Prevents cities from requiring multifamily recycling"
            ws.merge_cells(f'A{row}:D{row}')
            row += 2

            ws[f'A{row}'] = "Required Waste Service:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = "- Minimum 1/4 cubic yard per unit"
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            ws[f'A{row}'] = "- Twice weekly minimum collection"
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            ws[f'A{row}'] = "- Private hauler required (city service not available)"
            ws.merge_cells(f'A{row}:D{row}')
            row += 2

            ws[f'A{row}'] = "Recycling Status:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = "- VOLUNTARY ONLY (no compliance required)"
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            ws[f'A{row}'] = "- Can offer as amenity via private contract if desired"
            ws.merge_cells(f'A{row}:D{row}')
            row += 2

        else:
            # Other AZ cities
            city = property_info['location'].split(',')[0]
            ws[f'A{row}'] = f"{city.upper()}, AZ - VOLUNTARY PROGRAMS"
            ws[f'A{row}'].fill = section_fill
            ws[f'A{row}'].font = section_font
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

            ws[f'A{row}'] = "No mandatory multifamily recycling ordinance found"
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

            ws[f'A{row}'] = "Voluntary recycling programs available through city or private haulers"
            ws.merge_cells(f'A{row}:D{row}')
            row += 2

            ws[f'A{row}'] = "Contact for Information:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = property_info['contact']
            ws.merge_cells(f'A{row}:D{row}')
            row += 2

    # Next Steps / Action Items
    ws[f'A{row}'] = "RECOMMENDED NEXT STEPS"
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    if property_info['confidence'] == 'LOW':
        ws[f'A{row}'] = "1. Contact city solid waste department (contact info above)"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        ws[f'A{row}'] = "2. Request solid waste ordinance and multifamily requirements"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        ws[f'A{row}'] = "3. Verify current waste hauler licensing status"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        ws[f'A{row}'] = "4. Update regulatory compliance summary with findings"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

    elif property_info['confidence'] == 'HIGH' and 'MANDATORY' in property_info['recycling_status']:
        ws[f'A{row}'] = "1. Verify property has active recycling service in place"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        ws[f'A{row}'] = "2. Confirm recycling containers are accessible to residents"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        ws[f'A{row}'] = "3. Check that verification records have been submitted to city"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        ws[f'A{row}'] = "4. Maintain compliance documentation for audits"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

    elif property_info['confidence'] == 'MEDIUM':
        ws[f'A{row}'] = "1. Verify current waste service meets minimum requirements"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        ws[f'A{row}'] = "2. Confirm waste hauler is properly licensed"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        ws[f'A{row}'] = "3. Evaluate resident demand for optional recycling amenity"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        ws[f'A{row}'] = "4. Consider voluntary certification programs if applicable"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

    else:  # PENDING
        ws[f'A{row}'] = "1. Research specific city/county solid waste ordinances"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        ws[f'A{row}'] = "2. Contact local solid waste department for requirements"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        ws[f'A{row}'] = "3. Document findings in regulatory compliance summary"
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

    row += 2

    # Footer
    ws[f'A{row}'] = "Generated by: WasteWise Regulatory Compliance Analysis"
    ws[f'A{row}'].font = Font(italic=True, size=9)
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    ws[f'A{row}'] = f"For detailed findings, see: {property_info.get('summary_file', 'Regulatory_Compliance_Summary.md')}"
    ws[f'A{row}'].font = Font(italic=True, size=9)
    ws.merge_cells(f'A{row}:D{row}')

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20


def add_regulatory_tab(property_key, workbook_path):
    """Add REGULATORY_COMPLIANCE tab to existing workbook"""

    print(f"\nProcessing: {property_key}")
    print(f"  Workbook: {workbook_path}")

    if not workbook_path.exists():
        print(f"  [ERROR] Workbook not found: {workbook_path}")
        return False

    property_info = PROPERTIES.get(property_key)
    if not property_info:
        print(f"  [ERROR] Property configuration not found: {property_key}")
        return False

    try:
        # Load existing workbook
        wb = load_workbook(workbook_path)
        print(f"  [OK] Loaded workbook with {len(wb.sheetnames)} sheets: {wb.sheetnames}")

        # Create regulatory compliance sheet
        print(f"  Creating REGULATORY_COMPLIANCE sheet...")
        create_regulatory_compliance_sheet(wb, property_info['name'], property_info)

        # Save workbook
        wb.save(workbook_path)
        print(f"  [OK] Saved workbook with REGULATORY_COMPLIANCE tab")

        return True

    except Exception as e:
        print(f"  [ERROR] Failed to add regulatory tab: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Add REGULATORY_COMPLIANCE tab to all existing workbooks"""

    print("="*60)
    print("ADD REGULATORY_COMPLIANCE TAB TO EXISTING WORKBOOKS")
    print("="*60)

    base_path = Path("Properties")

    if not base_path.exists():
        print(f"[ERROR] Properties folder not found: {base_path}")
        return 1

    success_count = 0
    failed_count = 0

    for property_key, property_info in PROPERTIES.items():
        # Construct path to existing workbook (without underscores in property name)
        folder_name = property_info['name'].replace(' ', '_')
        workbook_path = base_path / folder_name / f"{property_key}_WasteAnalysis_Validated.xlsx"

        if add_regulatory_tab(property_key, workbook_path):
            success_count += 1
        else:
            failed_count += 1

    print("\n" + "="*60)
    print("PROCESSING COMPLETE")
    print(f"[OK] Success: {success_count}/{len(PROPERTIES)}")
    print(f"[ERROR] Failed: {failed_count}/{len(PROPERTIES)}")
    print("="*60)

    return 0 if failed_count == 0 else 1


if __name__ == '__main__':
    import sys
    sys.exit(main())
